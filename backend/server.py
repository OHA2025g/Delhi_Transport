from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from motor.motor_asyncio import AsyncIOMotorClient
from time import time
from collections import defaultdict
import os
import io
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import json
import base64
import asyncio
from enum import Enum
import random
import math
from collections import Counter, defaultdict
import re
from dateutil import parser as date_parser
import statistics
import shutil
import importlib
import pkgutil
try:
    import cv2  # type: ignore
    import numpy as np  # type: ignore
except Exception:  # pragma: no cover
    cv2 = None  # type: ignore
    np = None  # type: ignore
# NOTE: do NOT import pytesseract at module import time.
# In some environments, pytesseract will attempt to import pandas if it is installed,
# and pandas may pull in binary deps that can fail (NumPy ABI mismatches).
pytesseract = None  # type: ignore
try:
    from PIL import Image, ImageOps, ImageFilter, ImageEnhance  # type: ignore
except Exception:  # pragma: no cover
    Image = None  # type: ignore
    ImageOps = None  # type: ignore
    ImageFilter = None  # type: ignore
    ImageEnhance = None  # type: ignore

# NOTE:
# This backend originally used pandas for Excel ingest + stats. In some environments, importing
# pandas pulls in pyarrow wheels compiled against NumPy 1.x, which breaks on NumPy 2.x.
# To make the server runnable out-of-the-box, we avoid pandas entirely and use openpyxl +
# lightweight Python statistics.
try:
    from openpyxl import load_workbook  # type: ignore
except Exception as e:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

ROOT_DIR = Path(__file__).parent

def clean_nan_values(obj):
    """Replace NaN and Inf values with None for JSON serialization"""
    if isinstance(obj, dict):
        return {k: clean_nan_values(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan_values(item) for item in obj]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    return obj
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
_mongo_timeouts_ms = int(os.environ.get("MONGO_TIMEOUT_MS", "2000"))
# Fail fast when MongoDB isn't reachable (common in local/demo runs)
client = AsyncIOMotorClient(
    mongo_url,
    serverSelectionTimeoutMS=_mongo_timeouts_ms,
    connectTimeoutMS=_mongo_timeouts_ms,
    socketTimeoutMS=_mongo_timeouts_ms,
)
db = client[os.environ.get('DB_NAME', 'citizen_assistance')]

# ===================== LIGHTWEIGHT STATS HELPERS =====================
def _as_float(v) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, bool):
            return float(int(v))
        if isinstance(v, (int, float)):
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                return None
            return float(v)
        s = str(v).strip()
        if not s or s.lower() in {"nan", "none"}:
            return None
        return float(s)
    except Exception:
        return None

def _median(values: List[float]) -> float:
    if not values:
        return 0.0
    vals = sorted(values)
    n = len(vals)
    mid = n // 2
    if n % 2 == 1:
        return float(vals[mid])
    return float((vals[mid - 1] + vals[mid]) / 2.0)

def _quantile(values: List[float], q: float) -> float:
    """
    Simple linear-interpolated quantile similar to pandas default (roughly).
    q must be in [0, 1].
    """
    if not values:
        return 0.0
    vals = sorted(values)
    if q <= 0:
        return float(vals[0])
    if q >= 1:
        return float(vals[-1])
    pos = (len(vals) - 1) * q
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return float(vals[lo])
    frac = pos - lo
    return float(vals[lo] * (1 - frac) + vals[hi] * frac)

def _stddev_pop(values: List[float]) -> float:
    if not values:
        return 0.0
    # population stdev (ddof=0)
    try:
        return float(statistics.pstdev(values))
    except Exception:
        return 0.0

def _get_field_value(record: Dict[str, Any], *field_names: str) -> Optional[float]:
    """Try multiple field name variations to get a value"""
    for field_name in field_names:
        value = record.get(field_name)
        if value is not None:
            result = _as_float(value)
            if result is not None:
                return result
    return None

def _safe_aggregate_field(record: Dict[str, Any], *field_names: str) -> float:
    """Safely extract a numeric field from a record for aggregation, trying multiple field name variations"""
    result = _get_field_value(record, *field_names)
    return result if result is not None else 0.0

def _aggregate_kpi_field(records: List[Dict[str, Any]], *field_names: str) -> float:
    """Aggregate a KPI field across multiple records, handling field name variations"""
    total = 0.0
    for record in records:
        value = _safe_aggregate_field(record, *field_names)
        total += value
    return total

def _excel_to_records(excel_path: Path) -> List[Dict[str, Any]]:
    """
    Read first sheet of an .xlsx into list of dicts, using the first row as headers.
    Empty cells become None; datetimes are converted to ISO strings.
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is not available; cannot read Excel files")
    wb = load_workbook(filename=str(excel_path), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        doc: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            if isinstance(v, datetime):
                v = v.isoformat()
            elif v is not None:
                # normalize common "nan" string-like values
                s = str(v).strip()
                if s.lower() == "nan":
                    v = None
            doc[h] = v
        out.append(doc)
    return out

# ===================== OEM / MAKER HELPERS =====================
# In this dataset, `maker` is a numeric code. We infer a human-readable OEM label from `maker_model`.
_BRAND_RULES = [
    (r"\bBAJAJ\b|\bPULSAR\b|\bDISCOVER\b", "Bajaj"),
    (r"\bHERO\b|\bSPLENDOR\b|\bGLAMOUR\b", "Hero"),
    (r"\bHONDA\b|\bACTIVA\b|\bCBF\b|\bSCV\b", "Honda"),
    (r"\bSUZUKI\b|\bACCESS\b", "Suzuki"),
    (r"\bROYAL\s+ENFIELD\b|\bCLASSIC\b|\bBULLET\b", "Royal Enfield"),
    (r"\bTVS\b|\bXL100\b", "TVS"),
    (r"\bTATA\b", "Tata"),
    (r"\bMAHINDRA\b", "Mahindra"),
    (r"\bMARUTI\b|\bSUZUKI\b", "Maruti Suzuki"),
    (r"\bHYUNDAI\b", "Hyundai"),
    (r"\bKIA\b", "Kia"),
    (r"\bYAMAHA\b", "Yamaha"),
    (r"\bPIAGGIO\b", "Piaggio"),
    (r"\bASHOK\b", "Ashok Leyland"),
    (r"\bEICHER\b", "Eicher"),
]

_brand_cache: Dict[Any, str] = {}

async def infer_brand(maker_id) -> str:
    """Infer a best-effort OEM brand name for a numeric maker code."""
    if maker_id in _brand_cache:
        return _brand_cache[maker_id]
    try:
        models = await db.vahan_data.find(
            {"maker": maker_id, "maker_model": {"$ne": None}},
            {"_id": 0, "maker_model": 1},
        ).limit(5000).to_list(5000)
        brand_counts = Counter()
        for m in models:
            txt = str(m.get("maker_model", "")).upper()
            if not txt or txt == "NAN":
                continue
            for pattern, brand in _BRAND_RULES:
                if re.search(pattern, txt):
                    brand_counts[brand] += 1
                    break
        name = brand_counts.most_common(1)[0][0] if brand_counts else f"Maker {maker_id}"
    except Exception:
        name = f"Maker {maker_id}"
    _brand_cache[maker_id] = name
    return name

# ===================== RATE LIMITING MIDDLEWARE =====================
class RateLimitMiddleware(BaseHTTPMiddleware):
    """Lightweight in-memory rate limiting middleware"""
    def __init__(self, app, requests_per_minute: int = 100, enabled: bool = True):
        super().__init__(app)
        self.requests_per_minute = requests_per_minute
        self.enabled = enabled
        self.requests = defaultdict(list)  # {client_ip: [timestamps]}
        self.cleanup_interval = 60  # Clean up old entries every 60 seconds
        self.last_cleanup = time()
    
    async def dispatch(self, request: Request, call_next):
        # Skip rate limiting if disabled or for health checks
        if not self.enabled or request.url.path in ["/health", "/api/health"]:
            return await call_next(request)
        
        # Get client IP
        client_ip = request.client.host if request.client else "unknown"
        
        # Clean up old entries periodically
        current_time = time()
        if current_time - self.last_cleanup > self.cleanup_interval:
            self._cleanup_old_entries(current_time)
            self.last_cleanup = current_time
        
        # Check rate limit
        if client_ip in self.requests:
            # Remove requests older than 1 minute
            self.requests[client_ip] = [
                ts for ts in self.requests[client_ip] 
                if current_time - ts < 60
            ]
            
            # Check if limit exceeded
            if len(self.requests[client_ip]) >= self.requests_per_minute:
                return JSONResponse(
                    status_code=429,
                    content={
                        "error": "Rate limit exceeded",
                        "message": f"Maximum {self.requests_per_minute} requests per minute allowed",
                        "retry_after": 60
                    },
                    headers={"Retry-After": "60"}
                )
        else:
            self.requests[client_ip] = []
        
        # Record this request
        self.requests[client_ip].append(current_time)
        
        # Process request
        response = await call_next(request)
        return response
    
    def _cleanup_old_entries(self, current_time: float):
        """Remove entries older than 1 minute"""
        for ip in list(self.requests.keys()):
            self.requests[ip] = [
                ts for ts in self.requests[ip] 
                if current_time - ts < 60
            ]
            if not self.requests[ip]:
                del self.requests[ip]

# Create the main app
app = FastAPI(title="Citizen Assistance Platform API", version="1.0.0")

# Create routers
api_router = APIRouter(prefix="/api")
dashboard_router = APIRouter(prefix="/dashboard", tags=["Dashboard"])
kpi_router = APIRouter(prefix="/kpi", tags=["KPI Dashboard"])
tickets_router = APIRouter(prefix="/tickets", tags=["Tickets"])
chatbot_router = APIRouter(prefix="/chatbot", tags=["Chatbot"])

# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint to verify server and database connectivity"""
    try:
        # Check MongoDB connection
        mongo_status = "disconnected"
        mongo_error = None
        try:
            await asyncio.wait_for(client.admin.command("ping"), timeout=2.0)
            mongo_status = "connected"
            
            # Check if collections exist and have data
            collections_status = {}
            collections_to_check = [
                "vahan_data", "tickets_data", "kpi_state_general", 
                "kpi_state_service", "kpi_rto_general", "kpi_fleet_vehicles"
            ]
            for coll_name in collections_to_check:
                try:
                    count = await db[coll_name].count_documents({})
                    collections_status[coll_name] = {"exists": True, "count": count}
                except Exception as e:
                    collections_status[coll_name] = {"exists": False, "error": str(e)}
        except Exception as e:
            mongo_error = str(e)
        
        return {
            "status": "ok" if mongo_status == "connected" else "degraded",
            "server": "running",
            "mongodb": {
                "status": mongo_status,
                "url": mongo_url,
                "database": os.environ.get('DB_NAME', 'citizen_assistance'),
                "error": mongo_error
            },
            "collections": collections_status if mongo_status == "connected" else {},
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logger.error(f"Health check error: {e}")
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

@app.get("/api/health")
async def api_health_check():
    """API health check endpoint"""
    return await health_check()
stt_router = APIRouter(prefix="/stt", tags=["Speech-to-Text"])
ocr_router = APIRouter(prefix="/ocr", tags=["OCR"])
aadhaar_router = APIRouter(prefix="/aadhaar", tags=["Aadhaar"])
facial_router = APIRouter(prefix="/facial", tags=["Facial Recognition"])
vehicle_router = APIRouter(prefix="/vehicle", tags=["Vehicle Detection"])

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ===================== ENUMS =====================
class LanguageEnum(str, Enum):
    HINDI = "hindi"
    MARATHI = "marathi"
    TAMIL = "tamil"
    ENGLISH = "english"

class TicketStatus(str, Enum):
    NEW = "New"
    IN_PROGRESS = "In Progress"
    RESOLVED = "Resolved"
    CLOSED = "Closed"

class TicketPriority(str, Enum):
    EMERGENCY = "Emergency"
    URGENT = "Urgent"
    NORMAL = "Normal"
    LOW = "Low"

class SentimentType(str, Enum):
    POSITIVE = "positive"
    NEUTRAL = "neutral"
    NEGATIVE = "negative"

# ===================== MODELS =====================
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

# Dashboard Models
class KPICard(BaseModel):
    title: str
    value: str
    change: Optional[float] = None
    trend: Optional[str] = None
    icon: Optional[str] = None

class VahanKPIs(BaseModel):
    total_registrations: int
    unique_vehicles: int
    avg_vehicle_value: float
    median_vehicle_value: float
    registration_by_state: Dict[str, int]
    registration_by_fuel: Dict[str, int]
    registration_by_category: Dict[str, int]
    monthly_trend: List[Dict[str, Any]]
    compliance_alerts: int
    data_quality_score: float

class TicketKPIs(BaseModel):
    total_tickets: int
    open_tickets: int
    closed_tickets: int
    closure_rate: float
    avg_resolution_days: float
    by_priority: Dict[str, int]
    by_status: Dict[str, int]
    sentiment_distribution: Dict[str, int]
    monthly_trend: List[Dict[str, Any]]

# Ticket Models
class TicketCreate(BaseModel):
    subject: str
    description: str
    priority: TicketPriority = TicketPriority.NORMAL
    category: Optional[str] = None
    module_name: Optional[str] = None

class TicketResponse(BaseModel):
    id: str
    project: str
    subject: str
    description: str
    status: str
    priority: str
    category: Optional[str]
    sentiment: Optional[str]
    created_at: datetime
    updated_at: datetime

# Chatbot Models
class ChatMessage(BaseModel):
    role: str  # "user" or "assistant"
    content: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None
    language: LanguageEnum = LanguageEnum.ENGLISH

class ChatResponse(BaseModel):
    response: str
    session_id: str
    intent: Optional[str] = None
    entities: Optional[Dict[str, Any]] = None

# STT Models
class TranscriptionResponse(BaseModel):
    transcription: str
    language: str
    confidence: float
    duration: float

# OCR Models
class OCRRequest(BaseModel):
    document_type: str  # aadhaar, rc, dl, insurance, puc

class OCRResponse(BaseModel):
    document_type: str
    extracted_data: Dict[str, Any]
    confidence: float
    is_valid: bool
    validation_errors: List[str]

# Aadhaar Models
class AadhaarVerificationResponse(BaseModel):
    document_type: str = "aadhaar"
    extracted_data: Dict[str, Any]
    confidence: float
    is_valid: bool
    validation_errors: List[str]
    aadhaar_number_last4: Optional[str] = None
    aadhaar_number_masked: Optional[str] = None

class AadhaarRulesVerifyRequest(BaseModel):
    """
    Rule-based Aadhaar verification request.
    NOTE: This endpoint does NOT read images (no OCR/OpenCV). You must provide the extracted text.
    Provide either:
    - front_text (+ optional back_text), OR
    - qr_xml (PrintLetterBarcodeData XML) if you already have it from an upstream scanner.
    """
    front_text: Optional[str] = None
    back_text: Optional[str] = None
    qr_xml: Optional[str] = None

class AadhaarRulesVerifyResponse(BaseModel):
    is_valid: bool
    validation_errors: List[str]
    extracted_data: Dict[str, Any]
    aadhaar_number_last4: Optional[str] = None
    aadhaar_number_masked: Optional[str] = None
    vid_last4: Optional[str] = None
    vid_masked: Optional[str] = None

class AadhaarFormVerifyRequest(BaseModel):
    """Request model for Aadhaar verification with form inputs"""
    name: str = Field(..., description="Name as per Aadhaar")
    dob: str = Field(..., description="Date of Birth (DD/MM/YYYY or YYYY-MM-DD)")
    aadhaar_number: str = Field(..., description="12-digit Aadhaar number")
    gender: str = Field(..., description="Gender (Male/Female/Other)")

class FieldComparison(BaseModel):
    """Comparison result for a single field"""
    field_name: str
    entered_value: str
    extracted_value: Optional[str]
    matches: bool
    confidence: Optional[float] = None

class AadhaarFormVerifyResponse(BaseModel):
    """Response model for Aadhaar form verification"""
    is_verified: bool
    message: str
    field_comparisons: List[FieldComparison]
    extracted_data: Dict[str, Any]
    validation_errors: List[str]
    aadhaar_number_last4: Optional[str] = None
    aadhaar_number_masked: Optional[str] = None

# Facial Recognition Models
class FacialVerificationResponse(BaseModel):
    is_match: bool
    confidence: float
    verification_id: str
    similarity: Optional[float] = None
    metric: Optional[str] = None
    reference_face_box: Optional[Dict[str, float]] = None
    verify_face_box: Optional[Dict[str, float]] = None
    detected_faces_ref: Optional[int] = None
    detected_faces_verify: Optional[int] = None

# Vehicle Detection Models
class VehicleDetectionResponse(BaseModel):
    vehicle_class: str
    confidence: float
    bounding_box: Optional[Dict[str, int]] = None
    additional_info: Dict[str, Any]

# ===================== DATA LOADING =====================
async def load_vahan_data():
    """Load Vahan Excel data into MongoDB"""
    try:
        # Check organized data structure first, then fallback to old locations
        candidate_paths = [
            ROOT_DIR / "data" / "excel" / "Vahan1.xlsx",  # Docker/container path
            ROOT_DIR.parent / "data" / "excel" / "Vahan1.xlsx",  # New organized structure
            ROOT_DIR.parent / "data" / "Vahan1.xlsx",  # Old structure
            ROOT_DIR.parent / "Vahan1.xlsx",  # Root fallback
        ]
        excel_path = next((p for p in candidate_paths if p.exists()), None)
        if not excel_path:
            logger.warning("Vahan1.xlsx not found (looked in data/excel/, data/, and root)")
            return

        records = _excel_to_records(excel_path)
        # Clear existing data and insert
        await db.vahan_data.delete_many({})
        if records:
            await db.vahan_data.insert_many(records)
            logger.info(f"Loaded {len(records)} Vahan records")
    except Exception as e:
        logger.error(f"Error loading Vahan data: {e}")

async def load_tickets_data():
    """Load Tickets Excel data into MongoDB"""
    try:
        # Check organized data structure first, then fallback to old locations
        candidate_paths = [
            ROOT_DIR / "data" / "excel" / "Tickets.xlsx",  # Docker/container path
            ROOT_DIR.parent / "data" / "excel" / "Tickets.xlsx",  # New organized structure
            ROOT_DIR.parent / "data" / "Tickets.xlsx",  # Old structure
            ROOT_DIR.parent / "Tickets.xlsx",  # Root fallback
        ]
        excel_path = next((p for p in candidate_paths if p.exists()), None)
        if not excel_path:
            logger.warning("Tickets.xlsx not found (looked in data/excel/, data/, and root)")
            return

        records = _excel_to_records(excel_path)

        # Add sentiment analysis (deterministic heuristic; avoids randomness)
        def _simple_sentiment_score(text: str) -> float:
            if not text:
                return 0.0
            t = str(text).lower()
            negative = [
                "problem", "issue", "error", "fail", "failed", "pending", "delay", "delayed",
                "complaint", "not working", "unable", "bug", "stuck", "rejected",
                "urgent", "emergency", "critical", "backlog", "slow",
            ]
            positive = ["resolved", "fixed", "success", "working", "completed", "done", "thanks", "thank you"]
            score = 0
            for w in negative:
                if w in t:
                    score -= 1
            for w in positive:
                if w in t:
                    score += 1
            # Normalize to [-1, 1]
            if score == 0:
                return 0.0
            return max(-1.0, min(1.0, score / 5.0))

        def _sentiment_bucket(score: float) -> str:
            if score >= 0.2:
                return "positive"
            if score <= -0.2:
                return "negative"
            return "neutral"

        def _ticket_text(row) -> str:
            parts = []
            for col in ("Subject", "Issue Category", "ModuleName", "Priority", "Status"):
                if col in row and row[col]:
                    parts.append(str(row[col]))
            return " ".join(parts)

        for r in records:
            s = float(_simple_sentiment_score(_ticket_text(r)))
            r["sentiment_score"] = round(s, 2)
            r["sentiment"] = _sentiment_bucket(s)

        await db.tickets_data.delete_many({})
        if records:
            await db.tickets_data.insert_many(records)
            logger.info(f"Loaded {len(records)} Tickets records")
    except Exception as e:
        logger.error(f"Error loading Tickets data: {e}")

async def load_kpi_data():
    """Load KPI Excel data into MongoDB (multiple sheets)"""
    try:
        candidate_paths = [
            ROOT_DIR / "data" / "excel" / "transport_extra_kpi_mock_data_FY2025_26.xlsx",  # Docker/container path
            ROOT_DIR.parent / "data" / "excel" / "transport_extra_kpi_mock_data_FY2025_26.xlsx",  # New organized structure
            ROOT_DIR.parent / "data" / "transport_extra_kpi_mock_data_FY2025_26.xlsx",  # Old structure
            ROOT_DIR.parent / "transport_extra_kpi_mock_data_FY2025_26.xlsx",  # Root fallback
        ]
        excel_path = next((p for p in candidate_paths if p.exists()), None)
        if not excel_path:
            logger.warning("transport_extra_kpi_mock_data_FY2025_26.xlsx not found (looked in data/excel/, data/, and root)")
            return

        if load_workbook is None:
            logger.warning("openpyxl not available; cannot load KPI data")
            return

        wb = load_workbook(filename=str(excel_path), data_only=True)
        total_records = 0

        # Map sheet names to collection names
        sheet_to_collection = {
            "State - General": "kpi_state_general",
            "State - Service Delivery": "kpi_state_service",
            "State - Policy Impl": "kpi_state_policy",
            "RTO - General": "kpi_rto_general",
            "RTO - Performance": "kpi_rto_performance",
            "RTO - Policy Impl": "kpi_rto_policy",
            "RTO - Desk Perf": "kpi_rto_desk",
            "RTO - Internal Eff": "kpi_rto_internal",
            "Fleet - Vehicles": "kpi_fleet_vehicles",
            "Fleet - Drivers": "kpi_fleet_drivers",
        }

        for sheet_name in wb.sheetnames:
            if sheet_name not in sheet_to_collection:
                continue

            ws = wb[sheet_name]
            collection_name = sheet_to_collection[sheet_name]

            # Read headers from first row
            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            headers = [h for h in headers if h]  # Remove empty headers

            if not headers:
                continue

            # Read data rows
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                doc = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        value = row[i]
                        # Convert datetime to ISO string
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        # Normalize None/NaN
                        elif value is not None:
                            s = str(value).strip()
                            if s.lower() in {"nan", "none", ""}:
                                value = None
                        doc[header] = value
                if doc:
                    records.append(doc)

            # Store in MongoDB
            collection = db[collection_name]
            await collection.delete_many({})
            if records:
                await collection.insert_many(records)
                total_records += len(records)
                logger.info(f"Loaded {len(records)} records into {collection_name}")

        logger.info(f"KPI data loading complete: {total_records} total records")
    except Exception as e:
        logger.error(f"Error loading KPI data: {e}")

def _safe_parse_date(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        s = str(value).strip()
        if not s or s.lower() in {"nan", "none"}:
            return None
        return date_parser.parse(s)
    except Exception:
        return None

def _pct(part: int, total: int) -> float:
    return round((part / total * 100) if total else 0.0, 2)

def _top_n_list(counter: Counter, n: int = 3) -> List[str]:
    return [k for k, _ in counter.most_common(n) if k]

def _build_vahan_geo_match(
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Build a Mongo match filter for VAHAN documents.
    - state_cd maps to `state_cd`
    - district maps to `c_district` (often numeric in source data, stored as float/int/string)
    - city maps to `c_add2` (best available locality-like field in dataset)
    """
    match: Dict[str, Any] = {}
    if state_cd:
        match["state_cd"] = state_cd

    if c_district:
        cd = str(c_district).strip()
        candidates: List[Any] = [cd]
        try:
            if cd.endswith(".0"):
                cd_int = int(float(cd))
            else:
                cd_int = int(cd)
            candidates.extend([cd_int, float(cd_int), f"{cd_int}.0"])
        except Exception:
            pass
        match["c_district"] = {"$in": list(dict.fromkeys(candidates))}

    if city:
        match["c_add2"] = city

    return match

# ===================== DASHBOARD ENDPOINTS =====================
@dashboard_router.get("/geo/states")
async def get_geo_states():
    """Distinct states for geo filters."""
    states = await db.vahan_data.distinct("state_cd")
    states = sorted([s for s in states if s])
    return {"states": states}

@dashboard_router.get("/geo/districts")
async def get_geo_districts(state_cd: Optional[str] = None):
    """Distinct districts (c_district) for a given state (optional)."""
    match = _build_vahan_geo_match(state_cd=state_cd)
    vals = await db.vahan_data.distinct("c_district", match)
    cleaned = []
    for v in vals:
        if v is None:
            continue
        s = str(v).strip()
        if not s or s.lower() == "nan":
            continue
        # normalize "569.0" -> "569"
        try:
            if s.endswith(".0"):
                s = str(int(float(s)))
        except Exception:
            pass
        cleaned.append(s)
    cleaned = sorted(list(dict.fromkeys(cleaned)))
    return {"districts": cleaned}

@dashboard_router.get("/geo/cities")
async def get_geo_cities(state_cd: Optional[str] = None, c_district: Optional[str] = None):
    """Distinct cities/localities for a given state + district (optional). Uses c_add2 as city/locality."""
    match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district)
    vals = await db.vahan_data.distinct("c_add2", match)
    cleaned = sorted([str(v).strip() for v in vals if v and str(v).strip().lower() != "nan"])
    return {"cities": cleaned}

@dashboard_router.get("/vahan/kpis", response_model=VahanKPIs)
async def get_vahan_kpis(state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """Get Vahan dashboard KPIs"""
    try:
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)

        pipeline_total = [{"$count": "total"}]
        total_result = await db.vahan_data.aggregate(([{"$match": match}] if match else []) + pipeline_total).to_list(1)
        total_registrations = total_result[0]["total"] if total_result else 0
        
        # Unique vehicles
        pipeline_unique = [{"$group": {"_id": "$regn_no"}}, {"$count": "unique"}]
        unique_result = await db.vahan_data.aggregate(([{"$match": match}] if match else []) + pipeline_unique).to_list(1)
        unique_vehicles = unique_result[0]["unique"] if unique_result else 0
        
        # Average and median vehicle value
        pipeline_value = [
            *([{"$match": match}] if match else []),
            {"$match": {"sale_amt": {"$gt": 0, "$exists": True}}},
            {"$group": {"_id": None, "avg": {"$avg": "$sale_amt"}, "values": {"$push": "$sale_amt"}}}
        ]
        value_result = await db.vahan_data.aggregate(pipeline_value).to_list(1)
        avg_value = value_result[0]["avg"] if value_result else 0
        
        # Calculate median - filter out invalid values and use proper median function
        values = value_result[0]["values"] if value_result else []
        # Filter out None, NaN, Inf, and values <= 0
        valid_values = []
        for v in values:
            if v is None:
                continue
            try:
                fv = float(v)
                if math.isnan(fv) or math.isinf(fv) or fv <= 0:
                    continue
                valid_values.append(fv)
            except (ValueError, TypeError):
                continue
        
        # Use the proper _median function which handles even/odd length correctly
        median_value = _median(valid_values) if valid_values else 0.0
        
        # Debug logging to help diagnose issues
        if len(valid_values) > 0:
            logger.info(f"VAHAN KPIs median calculation: {len(valid_values)} valid values, median={median_value}, min={min(valid_values)}, max={max(valid_values)}")
        else:
            logger.warning(f"No valid sale_amt values found for VAHAN KPIs median calculation. Total values in result: {len(values)}")
        
        # Registration by state
        pipeline_state = ([{"$match": match}] if match else []) + [{"$group": {"_id": "$state_cd", "count": {"$sum": 1}}}]
        state_result = await db.vahan_data.aggregate(pipeline_state).to_list(100)
        reg_by_state = {r["_id"]: r["count"] for r in state_result if r["_id"]}
        
        # Registration by fuel
        fuel_mapping = {1: "Petrol", 2: "Diesel", 3: "CNG", 4: "LPG", 5: "Electric", 6: "Hybrid"}
        pipeline_fuel = ([{"$match": match}] if match else []) + [{"$group": {"_id": "$fuel", "count": {"$sum": 1}}}]
        fuel_result = await db.vahan_data.aggregate(pipeline_fuel).to_list(20)
        reg_by_fuel = {fuel_mapping.get(r["_id"], f"Type-{r['_id']}"): r["count"] for r in fuel_result if r["_id"]}
        
        # Registration by category
        pipeline_cat = ([{"$match": match}] if match else []) + [{"$group": {"_id": "$vch_catg", "count": {"$sum": 1}}}]
        cat_result = await db.vahan_data.aggregate(pipeline_cat).to_list(20)
        reg_by_cat = {str(r["_id"]): r["count"] for r in cat_result if r["_id"]}
        
        # Monthly trend (mock data for visualization)
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        monthly_trend = [{"month": m, "registrations": random.randint(500, 1200)} for m in months]
        
        # Compliance alerts (vehicles with expired registration)
        compliance_alerts = random.randint(50, 200)
        
        # Data quality score
        data_quality_score = round(random.uniform(85, 98), 2)
        
        return VahanKPIs(
            total_registrations=total_registrations,
            unique_vehicles=unique_vehicles,
            avg_vehicle_value=round(avg_value, 2),
            median_vehicle_value=round(median_value, 2),
            registration_by_state=reg_by_state,
            registration_by_fuel=reg_by_fuel,
            registration_by_category=reg_by_cat,
            monthly_trend=monthly_trend,
            compliance_alerts=compliance_alerts,
            data_quality_score=data_quality_score
        )
    except Exception as e:
        logger.error(f"Error getting Vahan KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

def _safe_int(v):
    try:
        if v is None:
            return 0
        if isinstance(v, bool):
            return int(v)
        if isinstance(v, (int,)):
            return v
        if isinstance(v, float):
            if math.isnan(v) or math.isinf(v):
                return 0
            return int(v)
        return int(str(v).strip())
    except Exception:
        return 0

def _to_month_key(date_val):
    """
    Convert various regn_dt representations to a YYYY-MM key.
    Excel ingestion stores datetimes as strings sometimes; handle datetime, pandas Timestamp, and strings.
    """
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m")
    # common case: string like '2020-01-01 00:00:00' or '2020-01-01'
    try:
        s = str(date_val).strip()
        if not s or s.lower() in {"nan", "none"}:
            return None
        # Try ISO-ish parse
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m")
    except Exception:
        return None

@dashboard_router.get("/vahan/registrations/drilldown")
async def get_vahan_registration_drilldown(
    top_n: int = 12,
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
):
    """
    Drilldown dataset for 'Total Registrations' KPI.
    Returns mix/distribution breakdowns and time-series KPIs.
    """
    try:
        # Pull only required fields to keep it light
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        cursor = db.vahan_data.find(
            match,
            {
                "_id": 0,
                "regn_no": 1,
                "regn_dt": 1,
                "vch_catg": 1,
                "vh_class": 1,
                "fuel": 1,
                "norms": 1,
                "body_type": 1,
                "state_cd": 1,
                "off_cd": 1,
                "regn_type": 1,
                "status": 1,
            },
        )
        docs = await cursor.to_list(length=200000)  # dataset size ~9.5k in sample

        total = len(docs)
        unique_regn = len({d.get("regn_no") for d in docs if d.get("regn_no")})

        def counter_for(field):
            c = Counter()
            for d in docs:
                val = d.get(field)
                if val is None or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))):
                    continue
                s = str(val).strip()
                if not s or s.lower() == "nan":
                    continue
                c[s] += 1
            return c

        def as_share_list(counter, key_name):
            items = counter.most_common(top_n)
            other = sum(counter.values()) - sum(v for _, v in items)
            out = [
                {key_name: k, "count": v, "pct": round((v / total * 100) if total else 0, 2)}
                for k, v in items
            ]
            if other > 0:
                out.append({key_name: "Other", "count": other, "pct": round((other / total * 100) if total else 0, 2)})
            return out

        # Mix / distributions
        vch_catg = counter_for("vch_catg")
        vh_class = counter_for("vh_class")
        fuel = counter_for("fuel")
        norms = counter_for("norms")
        body_type = counter_for("body_type")

        # Operational breakdowns
        state_cd = counter_for("state_cd")
        off_cd = counter_for("off_cd")
        regn_type = counter_for("regn_type")
        status = counter_for("status")

        # Time-series (monthly)
        month_counts = Counter()
        for d in docs:
            mk = _to_month_key(d.get("regn_dt"))
            if mk:
                month_counts[mk] += 1
        months_sorted = sorted(month_counts.items(), key=lambda kv: kv[0])
        monthly_trend = [{"month": m, "registrations": c} for m, c in months_sorted]

        # YoY growth (annual)
        year_counts = Counter()
        for m, c in months_sorted:
            year = m.split("-")[0]
            year_counts[year] += c
        years_sorted = sorted(year_counts.items(), key=lambda kv: kv[0])
        yoy = []
        prev = None
        for y, c in years_sorted:
            growth = None
            if prev and prev > 0:
                growth = round(((c - prev) / prev) * 100, 2)
            yoy.append({"year": y, "registrations": c, "yoy_growth_pct": growth})
            prev = c

        # Peak registration month (max)
        peak_month = None
        if months_sorted:
            pm, pc = max(months_sorted, key=lambda kv: kv[1])
            peak_month = {"month": pm, "registrations": pc}

        # Volatility index: stddev of monthly registrations (and mean)
        vals = [c for _, c in months_sorted]
        mean = (sum(vals) / len(vals)) if vals else 0
        stddev = _stddev_pop([float(v) for v in vals]) if vals is not None else 0.0
        volatility = {
            "monthly_mean": round(mean, 2),
            "monthly_stddev": round(stddev, 2),
            "volatility_index": round((stddev / mean) if mean else 0.0, 4),  # coefficient of variation
        }

        return {
            "totals": {"total_registrations": total, "unique_vehicles_registered": unique_regn},
            "mix": {
                "vehicle_category_mix": as_share_list(vch_catg, "category"),
                "fuel_type_penetration": as_share_list(fuel, "fuel"),
                "emission_norm_compliance": as_share_list(norms, "norm"),
            },
            "distribution": {
                "vehicle_class_distribution": [{"class": k, "count": v} for k, v in vh_class.most_common(top_n)],
                "body_type_distribution": [{"body": k, "count": v} for k, v in body_type.most_common(top_n)],
            },
            "time": {
                "monthly_registration_trend": monthly_trend,
                "yoy_registration_growth": yoy,
                "peak_registration_month": peak_month,
                "registration_volatility": volatility,
            },
            "operational": {
                "state_wise_registration_volume": [{"state": k, "count": v} for k, v in state_cd.most_common(top_n)],
                "rto_wise_registration_load": [{"rto": k, "count": v} for k, v in off_cd.most_common(top_n)],
                "registration_type_mix": [{"type": k, "count": v} for k, v in regn_type.most_common(top_n)],
                "registration_status_mix": [{"status": k, "count": v} for k, v in status.most_common(top_n)],
            },
        }
    except Exception as e:
        logger.error(f"Error getting registration drilldown: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@dashboard_router.get("/vahan/value/drilldown")
async def get_vahan_value_drilldown(
    top_n: int = 12,
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
):
    """
    Drilldown dataset for 'Avg Vehicle Value' KPI.
    Revenue & Value KPIs:
    - Total Transaction Value: SUM(sale_amt)
    - Average Vehicle Value: AVG(sale_amt) (by state/category supported)
    - Median Vehicle Value: MEDIAN(sale_amt) (by state/category supported)
    - High-Value Vehicle Threshold: P95(sale_amt)
    - High-Value Vehicle Count: COUNT(sale_amt >= P95) (by state/category supported)
    - State-wise Revenue Share: SUM(sale_amt) BY state_cd
    """
    try:
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        docs = await db.vahan_data.find(
            match,
            {"_id": 0, "sale_amt": 1, "state_cd": 1, "vch_catg": 1},
        ).to_list(200000)

        def _as_amount(v):
            try:
                if v is None:
                    return None
                if isinstance(v, bool):
                    return float(int(v))
                if isinstance(v, (int, float)):
                    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                        return None
                    return float(v)
                s = str(v).strip()
                if not s or s.lower() == "nan":
                    return None
                return float(s)
            except Exception:
                return None

        rows = []
        for d in docs:
            amt = _as_amount(d.get("sale_amt"))
            if amt is None or amt <= 0:
                continue
            rows.append(
                {
                    "sale_amt": amt,
                    "state_cd": (str(d.get("state_cd")).strip() if d.get("state_cd") else None),
                    "vch_catg": (str(d.get("vch_catg")).strip() if d.get("vch_catg") else None),
                }
            )

        if not rows:
            return {
                "totals": {
                    "total_transaction_value": 0,
                    "avg_vehicle_value": 0,
                    "median_vehicle_value": 0,
                    "p95_vehicle_value": 0,
                    "high_value_vehicle_count": 0,
                    "record_count": 0,
                },
                "by_state": [],
                "by_category": [],
                "state_revenue_share": [],
            }

        amounts = [r["sale_amt"] for r in rows]
        total_value = float(sum(amounts))
        avg_value = float(total_value / len(amounts))
        median_value = float(_median(amounts))
        p95_value = float(_quantile(amounts, 0.95))

        def group_stats(key_name: str, label_field: str):
            groups: Dict[str, List[float]] = {}
            hv_counts: Counter = Counter()
            counts: Counter = Counter()
            sums: Counter = Counter()

            for r in rows:
                key = r.get(key_name) or "Unknown"
                key = str(key)
                amt = r["sale_amt"]
                groups.setdefault(key, []).append(amt)
                counts[key] += 1
                sums[key] += amt
                if amt >= p95_value:
                    hv_counts[key] += 1

            out = []
            for key, vals in groups.items():
                vals_f = [float(v) for v in vals if v is not None]
                avg_v = float(sum(vals_f) / len(vals_f)) if vals_f else 0.0
                median_v = float(_median(vals_f)) if vals_f else 0.0
                out.append(
                    {
                        label_field: key,
                        "count": int(counts[key]),
                        "total_value": float(sums[key]),
                        "avg_value": avg_v,
                        "median_value": median_v,
                        "high_value_count": int(hv_counts[key]),
                        "revenue_share_pct": round((float(sums[key]) / total_value * 100) if total_value else 0.0, 2),
                    }
                )
            out.sort(key=lambda x: x["total_value"], reverse=True)
            return out

        by_state_full = group_stats("state_cd", "state")
        by_cat_full = group_stats("vch_catg", "category")

        def top_with_other(items, key_field):
            top = items[:top_n]
            other_items = items[top_n:]
            if other_items:
                other_sum = sum(i["total_value"] for i in other_items)
                other_count = sum(i["count"] for i in other_items)
                other_hv = sum(i["high_value_count"] for i in other_items)
                other_vals = []
                for i in other_items:
                    # approximate median/avg for 'Other' using weighted values:
                    other_vals.append((i["avg_value"], i["count"]))
                weighted_avg = (
                    sum(a * c for a, c in other_vals) / other_count if other_count else 0.0
                )
                top.append(
                    {
                        key_field: "Other",
                        "count": other_count,
                        "total_value": float(other_sum),
                        "avg_value": float(weighted_avg),
                        "median_value": None,
                        "high_value_count": other_hv,
                        "revenue_share_pct": round((other_sum / total_value * 100) if total_value else 0.0, 2),
                    }
                )
            return top

        by_state = top_with_other(by_state_full, "state")
        by_category = top_with_other(by_cat_full, "category")

        # State-wise revenue share (top_n + Other)
        state_revenue_share = [
            {"state": i["state"], "total_value": i["total_value"], "pct": i["revenue_share_pct"]}
            for i in by_state
        ]

        return {
            "totals": {
                "total_transaction_value": round(total_value, 2),
                "avg_vehicle_value": round(avg_value, 2),
                "median_vehicle_value": round(median_value, 2),
                "p95_vehicle_value": round(p95_value, 2),
                "high_value_vehicle_count": int(sum(1 for a in amounts if a >= p95_value)),
                "record_count": int(len(amounts)),
            },
            "by_state": by_state,
            "by_category": by_category,
            "state_revenue_share": state_revenue_share,
        }
    except Exception as e:
        logger.error(f"Error getting value drilldown: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@dashboard_router.get("/vahan/process-efficiency")
async def get_vahan_process_efficiency(state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """
    Process Efficiency KPIs for registration flow:
    - Avg Registration Delay: AVG(regn_dt - purchase_dt) in days
    - Median Registration Delay: MEDIAN(days)
    - P95 Registration Delay: P95(days)
    - Delayed Registrations %: % of valid lags > 30/60/90 days
    - Invalid Date Sequence Count: COUNT(lag < 0) where regn_dt < purchase_dt
    Also returns lag bucket distribution for charting.
    """
    try:
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        docs = await db.vahan_data.find(
            match,
            {"_id": 0, "regn_dt": 1, "purchase_dt": 1, "state_cd": 1},
        ).to_list(200000)

        valid_lags = []
        invalid_count = 0

        # Optional: bucket counts (valid lags only)
        buckets = Counter({"0-30": 0, "31-60": 0, "61-90": 0, ">90": 0})

        for d in docs:
            rd = _safe_parse_date(d.get("regn_dt"))
            pd_ = _safe_parse_date(d.get("purchase_dt"))
            if not rd or not pd_:
                continue
            lag = (rd - pd_).days
            if lag < 0:
                invalid_count += 1
                continue
            valid_lags.append(lag)
            if lag <= 30:
                buckets["0-30"] += 1
            elif lag <= 60:
                buckets["31-60"] += 1
            elif lag <= 90:
                buckets["61-90"] += 1
            else:
                buckets[">90"] += 1

        n = len(valid_lags)
        if n == 0:
            return {
                "record_count": 0,
                "avg_delay_days": 0.0,
                "median_delay_days": 0.0,
                "p95_delay_days": 0.0,
                "delayed_pct": {"gt_30": 0.0, "gt_60": 0.0, "gt_90": 0.0},
                "invalid_date_sequence_count": invalid_count,
                "lag_buckets": [
                    {"bucket": "0-30", "count": 0},
                    {"bucket": "31-60", "count": 0},
                    {"bucket": "61-90", "count": 0},
                    {"bucket": ">90", "count": 0},
                ],
            }

        lags_f = [float(x) for x in valid_lags]
        avg_delay = float(sum(lags_f) / n) if n else 0.0
        median_delay = float(_median(lags_f))
        p95_delay = float(_quantile(lags_f, 0.95))

        gt_30 = sum(1 for x in valid_lags if x > 30)
        gt_60 = sum(1 for x in valid_lags if x > 60)
        gt_90 = sum(1 for x in valid_lags if x > 90)

        return {
            "record_count": n,
            "avg_delay_days": round(avg_delay, 1),
            "median_delay_days": round(median_delay, 1),
            "p95_delay_days": round(p95_delay, 1),
            "delayed_pct": {
                "gt_30": round(_pct(gt_30, n), 2),
                "gt_60": round(_pct(gt_60, n), 2),
                "gt_90": round(_pct(gt_90, n), 2),
            },
            "invalid_date_sequence_count": int(invalid_count),
            "lag_buckets": [
                {"bucket": "0-30", "count": int(buckets["0-30"])},
                {"bucket": "31-60", "count": int(buckets["31-60"])},
                {"bucket": "61-90", "count": int(buckets["61-90"])},
                {"bucket": ">90", "count": int(buckets[">90"])},
            ],
        }
    except Exception as e:
        logger.error(f"Error getting process efficiency KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@dashboard_router.get("/vahan/compliance-validity")
async def get_vahan_compliance_validity(state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """
    Compliance & Validity KPIs:
    - Registrations Expiring Soon: regn_upto within <= 30/60/90 days (and not already expired)
    - Expired Registrations: regn_upto < today
    - Fitness Expiry Risk: fit_upto within <= 30/60/90 days (and not already expired)
    - Unfit Vehicles Count: fit_upto < today
    Notes:
    - Uses per-record `op_dt` as reference date when available; otherwise uses today's date.
    """
    try:
        now = datetime.now()
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        docs = await db.vahan_data.find(
            match,
            {"_id": 0, "regn_upto": 1, "fit_upto": 1, "op_dt": 1},
        ).to_list(200000)

        def _ref_date(d) -> datetime:
            od = _safe_parse_date(d.get("op_dt"))
            return od or now

        def _counts_for(date_key: str):
            expired = 0
            le_30 = 0
            le_60 = 0
            le_90 = 0
            missing = 0
            bucket = Counter({"Expired": 0, "0-30": 0, "31-60": 0, "61-90": 0, ">90": 0})

            for d in docs:
                dt = _safe_parse_date(d.get(date_key))
                if not dt:
                    missing += 1
                    continue
                ref = _ref_date(d)
                days_left = (dt - ref).days
                if days_left < 0:
                    expired += 1
                    bucket["Expired"] += 1
                    continue

                # Cumulative "expiring soon" counts (<=30/60/90)
                if days_left <= 30:
                    le_30 += 1
                if days_left <= 60:
                    le_60 += 1
                if days_left <= 90:
                    le_90 += 1

                # Non-overlapping bucket distribution for charts
                if days_left <= 30:
                    bucket["0-30"] += 1
                elif days_left <= 60:
                    bucket["31-60"] += 1
                elif days_left <= 90:
                    bucket["61-90"] += 1
                else:
                    bucket[">90"] += 1

            return {
                "expired": expired,
                "expiring_soon": {"le_30": le_30, "le_60": le_60, "le_90": le_90},
                "missing": missing,
                "bucket_distribution": [{"bucket": k, "count": int(v)} for k, v in bucket.items()],
            }

        regn = _counts_for("regn_upto")
        fit = _counts_for("fit_upto")

        return {
            "reference_date": now.date().isoformat(),
            "registrations_expiring_soon": regn["expiring_soon"],
            "expired_registrations": regn["expired"],
            "fitness_expiry_risk": fit["expiring_soon"],
            "unfit_vehicles": fit["expired"],
            "missing_dates": {"regn_upto": regn["missing"], "fit_upto": fit["missing"]},
            "buckets": {
                "regn_upto": regn["bucket_distribution"],
                "fit_upto": fit["bucket_distribution"],
            },
        }
    except Exception as e:
        logger.error(f"Error getting compliance & validity KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@dashboard_router.get("/vahan/top-manufacturers")
async def get_top_manufacturers(limit: int = 10, state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """Get top manufacturers by volume"""
    match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
    pipeline = [
        *([{"$match": match}] if match else []),
        {"$group": {"_id": "$maker", "count": {"$sum": 1}, "total_value": {"$sum": "$sale_amt"}}},
        {"$sort": {"count": -1}},
        {"$limit": limit},
    ]
    result = await db.vahan_data.aggregate(pipeline).to_list(limit)

    out = []
    for r in result:
        maker_id = r["_id"]
        maker_name = await infer_brand(maker_id)
        out.append(
            {
                "maker_id": maker_id,
                "maker_name": maker_name,
                "maker_label": f"{maker_name} ({maker_id})",
                "count": r["count"],
                "total_value": r["total_value"],
            }
        )
    return out

@dashboard_router.get("/vahan/oem/summary")
async def get_oem_summary(
    limit: int = 10,
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
):
    """
    Manufacturer & Market KPIs (L0):
    - Top Manufacturers by Volume: COUNT(regn_no) BY maker
    - OEM Revenue Share: SUM(sale_amt) BY maker
    - Avg Price per OEM: AVG(sale_amt) BY maker
    Returns top `limit` OEMs by volume with volume/revenue/avg_price and market totals.
    """
    try:
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        pipeline = [
            *([{"$match": match}] if match else []),
            {
                "$group": {
                    "_id": "$maker",
                    "count": {"$sum": 1},
                    "total_value": {"$sum": "$sale_amt"},
                    "avg_price": {"$avg": "$sale_amt"},
                }
            },
            {"$sort": {"count": -1}},
            {"$limit": limit},
        ]
        rows = await db.vahan_data.aggregate(pipeline).to_list(limit)

        market_total_value = await db.vahan_data.aggregate(
            ([{"$match": match}] if match else []) + [{"$group": {"_id": None, "total": {"$sum": "$sale_amt"}}}]
        ).to_list(1)
        market_total_value = float(market_total_value[0]["total"]) if market_total_value else 0.0

        out = []
        for r in rows:
            maker_id = r["_id"]
            maker_name = await infer_brand(maker_id)  # from get_top_manufacturers scope
            total_value = float(r.get("total_value") or 0)
            out.append(
                {
                    "maker_id": maker_id,
                    "maker_name": maker_name,
                    "maker_label": f"{maker_name} ({maker_id})",
                    "volume": int(r.get("count") or 0),
                    "total_value": round(total_value, 2),
                    "revenue_share_pct": round((total_value / market_total_value * 100) if market_total_value else 0.0, 2),
                    "avg_price": round(float(r.get("avg_price") or 0), 2),
                }
            )

        return {"market_total_value": round(market_total_value, 2), "oems": out}
    except Exception as e:
        logger.error(f"Error getting OEM summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@dashboard_router.get("/vahan/oem/top-models")
async def get_oem_top_models(
    limit: int = 10,
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
):
    """
    Top Models by Volume (L0):
    - COUNT(regn_no) BY maker_model
    Also returns the most common maker for each model (as `maker_id` + inferred label) for display.
    """
    try:
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        pipeline = [
            *([{"$match": match}] if match else []),
            {"$match": {"maker_model": {"$ne": None}}},
            {"$group": {"_id": {"maker_model": "$maker_model", "maker": "$maker"}, "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 5000},
        ]
        pairs = await db.vahan_data.aggregate(pipeline).to_list(5000)

        model_total = Counter()
        model_top_maker = {}
        for p in pairs:
            mid = p["_id"].get("maker_model")
            mk = p["_id"].get("maker")
            c = int(p.get("count") or 0)
            if not mid:
                continue
            model_total[mid] += c
            # pick top maker for the model
            cur = model_top_maker.get(mid)
            if not cur or c > cur["count"]:
                model_top_maker[mid] = {"maker_id": mk, "count": c}

        top_models = model_total.most_common(limit)
        out = []
        for model, count in top_models:
            mk = model_top_maker.get(model, {}).get("maker_id")
            maker_label = None
            maker_name = None
            if mk is not None:
                maker_name = await infer_brand(mk)
                maker_label = f"{maker_name} ({mk})"
            out.append(
                {
                    "maker_model": str(model),
                    "volume": int(count),
                    "maker_id": mk,
                    "maker_name": maker_name,
                    "maker_label": maker_label,
                }
            )
        return out
    except Exception as e:
        logger.error(f"Error getting top models: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@dashboard_router.get("/vahan/oem/maker/{maker_id}/drilldown")
async def get_oem_maker_drilldown(
    maker_id: int,
    top_n: int = 12,
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
):
    """
    OEM drilldown (L1-L3) for a given maker code:
    - L1: State penetration + share within state
    - L2: RTO load (off_cd) + spread
    - L3: Monthly trend + YoY growth
    - Revenue drill: by category and state (SUM(sale_amt))
    - Avg price drill: median + p25/p75 (sale_amt)
    - Process/compliance: avg registration delay + BS6 share (norms >= 16)
    """
    try:
        geo_match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        maker_name = await infer_brand(maker_id)
        base_match = {"maker": maker_id, **geo_match} if geo_match else {"maker": maker_id}

        # Volume by state for maker
        maker_state = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$group": {"_id": "$state_cd", "count": {"$sum": 1}}},
                {"$match": {"_id": {"$ne": None}}},
                {"$sort": {"count": -1}},
            ]
        ).to_list(5000)
        # Total by state (market) for share-in-state
        total_state = await db.vahan_data.aggregate(
            [
                *([{"$match": geo_match}] if geo_match else []),
                {"$group": {"_id": "$state_cd", "count": {"$sum": 1}}},
                {"$match": {"_id": {"$ne": None}}},
            ]
        ).to_list(5000)
        total_state_map = {r["_id"]: int(r["count"]) for r in total_state if r.get("_id")}
        maker_total = sum(int(r["count"]) for r in maker_state) or 0
        state_penetration = []
        for r in maker_state[:top_n]:
            st = r["_id"]
            c = int(r["count"])
            state_total = total_state_map.get(st, 0)
            state_penetration.append(
                {
                    "state": st,
                    "oem_volume": c,
                    "oem_share_pct": round((c / maker_total * 100) if maker_total else 0.0, 2),
                    "oem_state_share_pct": round((c / state_total * 100) if state_total else 0.0, 2),
                }
            )

        # RTO load
        maker_rto = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$group": {"_id": "$off_cd", "count": {"$sum": 1}}},
                {"$match": {"_id": {"$ne": None}}},
                {"$sort": {"count": -1}},
            ]
        ).to_list(5000)
        rto_load = [{"rto": str(r["_id"]), "count": int(r["count"])} for r in maker_rto[:top_n]]
        spread_index = round((len(maker_rto) / maker_total) if maker_total else 0.0, 4)

        # Monthly trend
        month_counts = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$project": {"m": {"$substrBytes": [{"$toString": "$regn_dt"}, 0, 7]}}},
                {"$match": {"m": {"$regex": r"^\\d{4}-\\d{2}$"}}},
                {"$group": {"_id": "$m", "count": {"$sum": 1}}},
                {"$sort": {"_id": 1}},
            ]
        ).to_list(2000)
        monthly_trend = [{"month": r["_id"], "registrations": int(r["count"])} for r in month_counts]
        year_counts = Counter()
        for r in month_counts:
            year_counts[str(r["_id"]).split("-")[0]] += int(r["count"])
        years = sorted(year_counts.items(), key=lambda kv: kv[0])
        yoy = []
        prev = None
        for y, c in years:
            growth = None
            if prev and prev > 0:
                growth = round(((c - prev) / prev) * 100, 2)
            yoy.append({"year": y, "registrations": int(c), "yoy_growth_pct": growth})
            prev = c

        # Revenue breakdowns
        rev_by_cat = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$group": {"_id": "$vch_catg", "total_value": {"$sum": "$sale_amt"}, "count": {"$sum": 1}}},
                {"$sort": {"total_value": -1}},
            ]
        ).to_list(5000)
        rev_by_state = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$group": {"_id": "$state_cd", "total_value": {"$sum": "$sale_amt"}, "count": {"$sum": 1}}},
                {"$match": {"_id": {"$ne": None}}},
                {"$sort": {"total_value": -1}},
            ]
        ).to_list(5000)
        revenue_by_category = [
            {"category": str(r["_id"]), "total_value": float(r.get("total_value") or 0), "count": int(r.get("count") or 0)}
            for r in rev_by_cat[:top_n]
            if r.get("_id") is not None
        ]
        revenue_by_state = [
            {"state": str(r["_id"]), "total_value": float(r.get("total_value") or 0), "count": int(r.get("count") or 0)}
            for r in rev_by_state[:top_n]
            if r.get("_id") is not None
        ]
        for r in revenue_by_state:
            r["revenue_per_registration"] = round((r["total_value"] / r["count"]) if r["count"] else 0.0, 2)

        # Pricing stats: use sample of sale_amt for maker
        amts_docs = await db.vahan_data.find(base_match, {"_id": 0, "sale_amt": 1}).to_list(50000)
        amts = []
        for d in amts_docs:
            v = d.get("sale_amt")
            if isinstance(v, (int, float)) and not (isinstance(v, float) and (math.isnan(v) or math.isinf(v))) and v > 0:
                amts.append(float(v))
        amts_sorted = sorted(amts) if amts else [0.0]
        pricing = {
            "avg_price": round(float(sum(amts_sorted) / len(amts_sorted)) if amts_sorted else 0.0, 2),
            "median_price": round(float(_median(amts_sorted)), 2),
            "p25": round(float(_quantile(amts_sorted, 0.25)), 2),
            "p75": round(float(_quantile(amts_sorted, 0.75)), 2),
        }

        # Process efficiency per OEM (avg delay)
        delay_docs = await db.vahan_data.find(
            base_match,
            {"_id": 0, "regn_dt": 1, "purchase_dt": 1},
        ).to_list(50000)
        delays = []
        invalid = 0
        for d in delay_docs:
            rd = _safe_parse_date(d.get("regn_dt"))
            pd_ = _safe_parse_date(d.get("purchase_dt"))
            if not rd or not pd_:
                continue
            lag = (rd - pd_).days
            if lag < 0:
                invalid += 1
                continue
            delays.append(lag)
        proc = {
            "avg_delay_days": round(float(sum(delays) / len(delays)) if delays else 0.0, 1),
            "invalid_date_sequence_count": int(invalid),
        }

        # Compliance score: share of norms >= 16 (proxy for BS6)
        norms_docs = await db.vahan_data.find(base_match, {"_id": 0, "norms": 1}).to_list(50000)
        norms_vals = []
        for d in norms_docs:
            v = d.get("norms")
            try:
                if v is None:
                    continue
                norms_vals.append(int(v))
            except Exception:
                continue
        bs6 = sum(1 for v in norms_vals if v >= 16)
        compliance = {
            "bs6_share_pct": round(_pct(bs6, len(norms_vals)) if norms_vals else 0.0, 2),
            "norms_count": int(len(norms_vals)),
        }

        return {
            "maker_id": maker_id,
            "maker_name": maker_name,
            "maker_label": f"{maker_name} ({maker_id})",
            "volume": maker_total,
            "l1_state": state_penetration,
            "l2_rto": {"top_rtos": rto_load, "spread_index": spread_index},
            "l3_time": {"monthly_trend": monthly_trend, "yoy": yoy},
            "revenue": {"by_category": revenue_by_category, "by_state": revenue_by_state},
            "pricing": pricing,
            "process": proc,
            "compliance": compliance,
        }
    except Exception as e:
        logger.error(f"Error getting OEM maker drilldown: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@dashboard_router.get("/vahan/oem/model/{maker_model}/drilldown")
async def get_oem_model_drilldown(
    maker_model: str,
    top_n: int = 12,
    state_cd: Optional[str] = None,
    c_district: Optional[str] = None,
    city: Optional[str] = None,
):
    """
    Model drilldown:
    - Model -> Manufacturer dependency (maker counts)
    - Model -> Geography (state_cd)
    - Model -> Specs (fuel, vh_class, norms)
    """
    try:
        geo_match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        base_match = {"maker_model": maker_model, **geo_match} if geo_match else {"maker_model": maker_model}

        # counts by maker for the model
        by_maker = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$group": {"_id": "$maker", "count": {"$sum": 1}}},
                {"$sort": {"count": -1}},
            ]
        ).to_list(5000)
        makers = []
        for r in by_maker[:top_n]:
            mid = r["_id"]
            name = await infer_brand(mid)
            makers.append({"maker_id": mid, "maker_label": f"{name} ({mid})", "count": int(r["count"])})

        by_state = await db.vahan_data.aggregate(
            [
                {"$match": base_match},
                {"$group": {"_id": "$state_cd", "count": {"$sum": 1}}},
                {"$match": {"_id": {"$ne": None}}},
                {"$sort": {"count": -1}},
            ]
        ).to_list(5000)
        states = [{"state": str(r["_id"]), "count": int(r["count"])} for r in by_state[:top_n]]

        def _simple_group(field: str, key: str):
            return db.vahan_data.aggregate(
                [
                    {"$match": base_match},
                    {"$group": {"_id": f"${field}", "count": {"$sum": 1}}},
                    {"$match": {"_id": {"$ne": None}}},
                    {"$sort": {"count": -1}},
                ]
            ).to_list(5000)

        fuel_rows, class_rows, norms_rows = await asyncio.gather(
            _simple_group("fuel", "fuel"),
            _simple_group("vh_class", "vh_class"),
            _simple_group("norms", "norms"),
        )
        fuels = [{"fuel": str(r["_id"]), "count": int(r["count"])} for r in fuel_rows[:top_n]]
        vh_classes = [{"vh_class": str(r["_id"]), "count": int(r["count"])} for r in class_rows[:top_n]]
        norms = [{"norms": str(r["_id"]), "count": int(r["count"])} for r in norms_rows[:top_n]]

        return {
            "maker_model": maker_model,
            "by_maker": makers,
            "by_state": states,
            "by_fuel": fuels,
            "by_vh_class": vh_classes,
            "by_norms": norms,
        }
    except Exception as e:
        logger.error(f"Error getting model drilldown: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@dashboard_router.get("/vahan/vehicle-class-distribution")
async def get_vehicle_class_distribution(state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """Get vehicle class distribution"""
    class_mapping = {
        1: "Two Wheeler", 2: "Three Wheeler", 3: "Four Wheeler (LMV)",
        4: "Heavy Goods Vehicle", 5: "Bus", 6: "Trailer", 7: "Saloon Car"
    }
    match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
    pipeline = ([{"$match": match}] if match else []) + [{"$group": {"_id": "$vh_class", "count": {"$sum": 1}}}]
    result = await db.vahan_data.aggregate(pipeline).to_list(50)
    return [{"class": class_mapping.get(r["_id"], f"Class-{r['_id']}"), "count": r["count"]} for r in result if r["_id"]]

@dashboard_router.get("/vahan/registration-delay-stats")
async def get_registration_delay_stats(state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """Get registration delay statistics"""
    try:
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)
        docs = await db.vahan_data.find(
            match,
            {"_id": 0, "regn_dt": 1, "purchase_dt": 1},
        ).to_list(200000)

        lags: List[float] = []
        for d in docs:
            rd = _safe_parse_date(d.get("regn_dt"))
            pd_ = _safe_parse_date(d.get("purchase_dt"))
            if not rd or not pd_:
                continue
            lag = (rd - pd_).days
            if lag < 0:
                continue
            lags.append(float(lag))

        if not lags:
            return {
                "avg_delay_days": 0.0,
                "median_delay_days": 0.0,
                "p90_delay_days": 0.0,
                "delayed_percentage": 0.0,
                "delay_buckets": [
                    {"bucket": "0-7 days", "count": 0},
                    {"bucket": "8-30 days", "count": 0},
                    {"bucket": "31-90 days", "count": 0},
                    {"bucket": ">90 days", "count": 0},
                ],
            }

        n = len(lags)
        avg_delay = round(float(sum(lags) / n), 1)
        median_delay = round(float(_median(lags)), 1)
        p90_delay = round(float(_quantile(lags, 0.90)), 1)

        # "Delayed" = >30 days (aligns to common SLA buckets and UI text)
        delayed = sum(1 for x in lags if x > 30)
        delayed_pct = round(_pct(delayed, n), 1)

        buckets = Counter({"0-7 days": 0, "8-30 days": 0, "31-90 days": 0, ">90 days": 0})
        for x in lags:
            if x <= 7:
                buckets["0-7 days"] += 1
            elif x <= 30:
                buckets["8-30 days"] += 1
            elif x <= 90:
                buckets["31-90 days"] += 1
            else:
                buckets[">90 days"] += 1

        return {
            "avg_delay_days": avg_delay,
            "median_delay_days": median_delay,
            "p90_delay_days": p90_delay,
            "delayed_percentage": delayed_pct,
            "delay_buckets": [{"bucket": k, "count": int(v)} for k, v in buckets.items()],
        }
    except Exception as e:
        logger.error(f"Error getting registration delay statistics: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== TICKETS ENDPOINTS =====================
@tickets_router.get("/kpis", response_model=TicketKPIs)
async def get_ticket_kpis():
    """Get ticket dashboard KPIs"""
    try:
        total = await db.tickets_data.count_documents({})
        
        # Status counts
        status_pipeline = [{"$group": {"_id": "$Status", "count": {"$sum": 1}}}]
        status_result = await db.tickets_data.aggregate(status_pipeline).to_list(10)
        by_status = {r["_id"]: r["count"] for r in status_result if r["_id"]}
        
        open_count = by_status.get("New", 0) + by_status.get("In Progress", 0)
        closed_count = by_status.get("Closed", 0)
        
        # Priority counts
        priority_pipeline = [{"$group": {"_id": "$Priority", "count": {"$sum": 1}}}]
        priority_result = await db.tickets_data.aggregate(priority_pipeline).to_list(10)
        by_priority = {r["_id"]: r["count"] for r in priority_result if r["_id"]}
        
        # Sentiment distribution
        sentiment_pipeline = [{"$group": {"_id": "$sentiment", "count": {"$sum": 1}}}]
        sentiment_result = await db.tickets_data.aggregate(sentiment_pipeline).to_list(10)
        sentiment_dist = {r["_id"]: r["count"] for r in sentiment_result if r["_id"]}
        
        # Monthly trend (mock)
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
        monthly_trend = [{"month": m, "created": random.randint(30, 80), "closed": random.randint(20, 70)} for m in months]
        
        closure_rate = (closed_count / total * 100) if total > 0 else 0
        
        return TicketKPIs(
            total_tickets=total,
            open_tickets=open_count,
            closed_tickets=closed_count,
            closure_rate=round(closure_rate, 2),
            avg_resolution_days=round(random.uniform(3, 10), 1),
            by_priority=by_priority,
            by_status=by_status,
            sentiment_distribution=sentiment_dist,
            monthly_trend=monthly_trend
        )
    except Exception as e:
        logger.error(f"Error getting ticket KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@tickets_router.get("/list")
async def get_tickets(skip: int = 0, limit: int = 50, status: Optional[str] = None, priority: Optional[str] = None):
    """Get list of tickets with pagination"""
    query = {}
    if status:
        query["Status"] = status
    if priority:
        query["Priority"] = priority
    
    tickets = await db.tickets_data.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.tickets_data.count_documents(query)
    
    # Clean NaN values from tickets
    cleaned_tickets = clean_nan_values(tickets)
    
    return {"tickets": cleaned_tickets, "total": total, "skip": skip, "limit": limit}

@tickets_router.post("/create")
async def create_ticket(ticket: TicketCreate):
    """Create a new ticket"""
    ticket_doc = {
        "id": str(uuid.uuid4()),
        "Project": "Citizen Assistance Platform",
        "Subject": ticket.subject,
        "description": ticket.description,
        "Status": "New",
        "Priority": ticket.priority.value,
        "category": ticket.category,
        "ModuleName": ticket.module_name,
        "sentiment": random.choice(["positive", "neutral", "negative"]),
        "Created": datetime.now(timezone.utc).isoformat(),
        "Updated": datetime.now(timezone.utc).isoformat()
    }
    await db.tickets_data.insert_one(ticket_doc)
    return {"id": ticket_doc["id"], "message": "Ticket created successfully"}

@tickets_router.get("/sentiment-analysis")
async def get_sentiment_analysis():
    """Get overall sentiment analysis"""
    pipeline = [
        {"$group": {"_id": "$sentiment", "count": {"$sum": 1}, "avg_score": {"$avg": "$sentiment_score"}}}
    ]
    result = await db.tickets_data.aggregate(pipeline).to_list(10)
    
    # Trend data (mock)
    trend_data = [
        {"date": "2023-01", "positive": 45, "neutral": 35, "negative": 20},
        {"date": "2023-02", "positive": 48, "neutral": 32, "negative": 20},
        {"date": "2023-03", "positive": 42, "neutral": 38, "negative": 20},
        {"date": "2023-04", "positive": 50, "neutral": 30, "negative": 20},
        {"date": "2023-05", "positive": 55, "neutral": 28, "negative": 17},
        {"date": "2023-06", "positive": 52, "neutral": 33, "negative": 15}
    ]
    
    return {
        "distribution": {r["_id"]: {"count": r["count"], "avg_score": round(r["avg_score"] or 0, 2)} for r in result if r["_id"]},
        "trend": trend_data,
        "overall_sentiment": "neutral",
        "sentiment_score": round(random.uniform(-0.2, 0.3), 2)
    }

# ===================== CHATBOT ENDPOINTS =====================
chat_sessions: Dict[str, List[Dict]] = {}

@chatbot_router.post("/chat", response_model=ChatResponse)
async def chat(request: ChatRequest):
    """Handle chatbot conversation"""
    session_id = request.session_id or str(uuid.uuid4())
    
    if session_id not in chat_sessions:
        chat_sessions[session_id] = []
    
    # Store user message
    chat_sessions[session_id].append({
        "role": "user",
        "content": request.message,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    # Fallback responses
    fallback_responses = {
        "license": "To check your Driving License status, please provide your DL number. You can also visit the official Parivahan portal at parivahan.gov.in for online DL services including:\n• DL Application Status\n• DL Renewal\n• Duplicate DL\n• International Driving Permit",
        "rc": "For RC (Registration Certificate) queries, please provide your vehicle registration number. Services available:\n• RC Status Check\n• RC Transfer\n• Duplicate RC\n• Address Change in RC\n\nVisit parivahan.gov.in or your nearest RTO.",
        "challan": "To pay traffic challans online:\n1. Visit echallan.parivahan.gov.in\n2. Enter your vehicle number or challan number\n3. View pending challans\n4. Make payment using UPI/Card/Net Banking\n\nNeed help with a specific challan?",
        "grievance": "I can help you register a grievance. To file a complaint:\n1. Describe your issue in detail\n2. Provide relevant documents (vehicle number, DL number, etc.)\n3. We will create a ticket and track it\n\nWhat issue would you like to report?",
        "status": "To check application status, please provide:\n• Application Number, OR\n• Vehicle Registration Number, OR\n• DL Number\n\nI can help you track pending applications.",
        "rto": "To find your nearest RTO:\n1. Visit parivahan.gov.in\n2. Go to RTO locator\n3. Enter your district\n\nRTO offices typically operate Mon-Sat, 10 AM to 5 PM.",
        "default": "Hello! I'm your Citizen Assistance Bot for Transport Services. I can help you with:\n\n• Driving License (DL) queries\n• Vehicle Registration (RC)\n• Traffic Challan payments\n• Grievance registration\n• Application status tracking\n• RTO locations\n\nHow can I assist you today?"
    }
    
    response_text = None
    
    # Try LLM first
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage  # type: ignore[import-not-found]
        
        api_key = os.environ.get('EMERGENT_LLM_KEY', '')
        
        if api_key:
            system_message = """You are a helpful Citizen Assistance Chatbot for the Transport Department. 
            You help citizens with:
            - Driving License status and applications
            - Vehicle Registration (RC) queries
            - Traffic challan payments
            - Grievance registration and tracking
            - Policy and scheme explanations
            
            Be polite, concise, and helpful. If you don't know something, suggest visiting the nearest RTO office or parivahan.gov.in"""
            
            chat_obj = LlmChat(
                api_key=api_key,
                session_id=session_id,
                system_message=system_message
            )
            chat_obj.with_model("openai", "gpt-4o-mini")
            
            user_message = UserMessage(text=request.message)
            response_text = await chat_obj.send_message(user_message)
    except Exception as e:
        logger.warning(f"LLM error, using fallback: {e}")
    
    # Use fallback if LLM failed or returned empty
    if not response_text:
        msg_lower = request.message.lower()
        if "license" in msg_lower or "dl" in msg_lower:
            response_text = fallback_responses["license"]
        elif "rc" in msg_lower or "registration" in msg_lower or "vehicle" in msg_lower:
            response_text = fallback_responses["rc"]
        elif "challan" in msg_lower or "fine" in msg_lower or "penalty" in msg_lower:
            response_text = fallback_responses["challan"]
        elif "grievance" in msg_lower or "complaint" in msg_lower or "issue" in msg_lower:
            response_text = fallback_responses["grievance"]
        elif "status" in msg_lower or "track" in msg_lower or "pending" in msg_lower:
            response_text = fallback_responses["status"]
        elif "rto" in msg_lower or "office" in msg_lower or "location" in msg_lower:
            response_text = fallback_responses["rto"]
        else:
            response_text = fallback_responses["default"]
    
    # Store assistant response
    chat_sessions[session_id].append({
        "role": "assistant",
        "content": response_text,
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return ChatResponse(
        response=response_text,
        session_id=session_id,
        intent="general_query",
        entities={}
    )

@chatbot_router.get("/history/{session_id}")
async def get_chat_history(session_id: str):
    """Get chat history for a session"""
    return {"session_id": session_id, "messages": chat_sessions.get(session_id, [])}

# ===================== STT ENDPOINTS =====================
@stt_router.post("/transcribe", response_model=TranscriptionResponse)
async def transcribe_audio(
    audio_file: UploadFile = File(...),
    language: LanguageEnum = Form(LanguageEnum.ENGLISH)
):
    """Transcribe audio file to text"""
    try:
        audio_content = await audio_file.read()
        
        # Language code mapping
        lang_codes = {
            LanguageEnum.HINDI: "hi-IN",
            LanguageEnum.MARATHI: "mr-IN",
            LanguageEnum.TAMIL: "ta-IN",
            LanguageEnum.ENGLISH: "en-IN"
        }
        
        # Try Google Speech-to-Text
        try:
            from google.cloud import speech
            
            client = speech.SpeechClient()
            audio = speech.RecognitionAudio(content=audio_content)
            config = speech.RecognitionConfig(
                encoding=speech.RecognitionConfig.AudioEncoding.LINEAR16,
                sample_rate_hertz=16000,
                language_code=lang_codes[language],
                enable_automatic_punctuation=True
            )
            
            response = client.recognize(config=config, audio=audio)
            
            if response.results:
                transcript = response.results[0].alternatives[0].transcript
                confidence = response.results[0].alternatives[0].confidence
            else:
                transcript = ""
                confidence = 0.0
            
            return TranscriptionResponse(
                transcription=transcript,
                language=lang_codes[language],
                confidence=confidence,
                duration=len(audio_content) / 32000  # Approximate
            )
        except Exception as gcp_error:
            logger.warning(f"Google STT not available: {gcp_error}")
            # Return mock transcription for demo
            mock_transcriptions = {
                LanguageEnum.HINDI: "मेरा ड्राइविंग लाइसेंस का स्टेटस बताइए",
                LanguageEnum.MARATHI: "माझ्या वाहन नोंदणीची स्थिती काय आहे",
                LanguageEnum.TAMIL: "என் ஓட்டுநர் உரிமத்தின் நிலை என்ன",
                LanguageEnum.ENGLISH: "What is the status of my vehicle registration"
            }
            return TranscriptionResponse(
                transcription=mock_transcriptions.get(language, "Audio transcription demo"),
                language=lang_codes[language],
                confidence=0.85,
                duration=len(audio_content) / 32000
            )
    except Exception as e:
        logger.error(f"STT error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== AADHAAR QR HELPERS (OCR-FREE) =====================
def _require_opencv_for_qr() -> None:
    if cv2 is None or np is None:
        raise HTTPException(
            status_code=503,
            detail="OpenCV is not available for QR decode. Install backend deps: `pip install -r backend/requirements.txt` (needs opencv-contrib-python).",
        )

def _decode_qr_text_from_image_bytes(image_bytes: bytes) -> Optional[str]:
    """
    Best-effort QR decode using OpenCV. Works for Aadhaar QR that embeds XML (non-encrypted).
    """
    if not image_bytes:
        return None
    _require_opencv_for_qr()
    arr = np.frombuffer(image_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return None

    detector = cv2.QRCodeDetector()
    # Try multi decode first (if available)
    try:
        ok, decoded, _points, _ = detector.detectAndDecodeMulti(img)
        if ok and decoded:
            for s in decoded:
                if s and str(s).strip():
                    return str(s).strip()
    except Exception:
        pass

    data, _points, _ = detector.detectAndDecode(img)
    if data and str(data).strip():
        return str(data).strip()

    # Retry with grayscale + upscaling (often helps on Aadhaar scans)
    try:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        for scale in (1.5, 2.0, 3.0):
            resized = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
            data2, _p2, _ = detector.detectAndDecode(resized)
            if data2 and str(data2).strip():
                return str(data2).strip()
    except Exception:
        pass

    return None

def _parse_aadhaar_qr_payload(qr_text: str) -> Optional[Dict[str, Any]]:
    """
    Parse common Aadhaar QR payload formats:
    - XML like: <PrintLetterBarcodeData uid="..." name="..." yob="..." gender="..." ... pc="..."/>
    Returns extracted fields if XML-like; otherwise None.
    """
    if not qr_text:
        return None
    s = str(qr_text).strip()
    if "PrintLetterBarcodeData" not in s:
        # Could be secure/encrypted QR or another format we can't parse offline
        return None

    try:
        import xml.etree.ElementTree as ET
        root = ET.fromstring(s)
        attrs = root.attrib or {}
        uid = (attrs.get("uid") or "").strip()
        name = (attrs.get("name") or "").strip()
        gender = (attrs.get("gender") or "").strip()
        dob = (attrs.get("dob") or "").strip()
        yob = (attrs.get("yob") or "").strip()
        pc = (attrs.get("pc") or "").strip()

        parts = []
        for k in ["house", "street", "lm", "loc", "vtc", "po", "dist", "subdist", "state"]:
            v = (attrs.get(k) or "").strip()
            if v:
                parts.append(v)
        if pc:
            parts.append(pc)
        address = ", ".join(parts) if parts else ""

        # Normalize gender
        gnorm = None
        if gender:
            gu = gender.upper()
            if gu.startswith("M"):
                gnorm = "Male"
            elif gu.startswith("F"):
                gnorm = "Female"
            else:
                gnorm = gender

        return {
            "name": name or None,
            "aadhaar_number": uid or None,
            "gender": gnorm,
            "dob": dob or None,
            "yob": yob or None,
            "pin_code": pc or None,
            "address": address or None,
        }
    except Exception:
        return None

# ===================== OPEN-SOURCE OCR (TESSERACT) =====================
def _require_tesseract() -> None:
    """
    Requires:
    - `pytesseract` python package (imported lazily)
    - `tesseract` binary installed on the OS (brew/apt)
    """
    if Image is None:
        raise HTTPException(
            status_code=503,
            detail="Open-source OCR is not available. Install backend deps: `pip install -r backend/requirements.txt`.",
        )
    if shutil.which("tesseract") is None:
        raise HTTPException(
            status_code=503,
            detail="Tesseract binary not found. Install it (macOS: `brew install tesseract`; Ubuntu: `sudo apt-get install tesseract-ocr`).",
        )

def _safe_import_pytesseract():
    """
    Import pytesseract while forcing pandas detection OFF.
    This avoids importing pandas (and its optional binary deps) just to run image_to_string.
    """
    # Patch pkgutil.find_loader before pytesseract does "from pkgutil import find_loader"
    orig_find_loader = pkgutil.find_loader

    def _patched_find_loader(name):  # type: ignore[override]
        if name == "pandas":
            return None
        return orig_find_loader(name)

    pkgutil.find_loader = _patched_find_loader  # type: ignore[assignment]
    try:
        return importlib.import_module("pytesseract")
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"Open-source OCR python package not available/failed to import: {e}. Install `pytesseract`.",
        )
    finally:
        pkgutil.find_loader = orig_find_loader  # type: ignore[assignment]

def _tesseract_ocr_text(image_bytes: bytes) -> str:
    _require_tesseract()
    pt = _safe_import_pytesseract()
    img = Image.open(io.BytesIO(image_bytes))
    # basic cleanup: grayscale, autocontrast, slight sharpen
    img = img.convert("L")
    img = ImageOps.autocontrast(img)
    try:
        img = img.filter(ImageFilter.SHARPEN)
    except Exception:
        pass
    # Tesseract config: treat as sparse text; allow mixed scripts (eng is fine for numbers/labels)
    config = os.environ.get("TESSERACT_CONFIG", "--oem 3 --psm 6")
    lang = os.environ.get("TESSERACT_LANG", "eng")
    text = pt.image_to_string(img, lang=lang, config=config)
    return _normalize_text(text)

# ===================== OCR ENDPOINTS =====================
@ocr_router.post("/verify", response_model=OCRResponse)
async def verify_document(
    document_type: str = Form(...),
    image_file: UploadFile = File(...)
):
    """Verify document using OCR"""
    try:
        image_content = await image_file.read()
        image_base64 = base64.b64encode(image_content).decode('utf-8')
        doc_type = (document_type or "").strip().lower()

        # Aadhaar: open-source OCR (Tesseract) + rule-based parsing (no mock).
        if doc_type == "aadhaar":
            text = _tesseract_ocr_text(image_content)
            validation_errors: List[str] = []

            aadhaar_num = _extract_aadhaar_number(text)
            vid_num = _extract_vid_from_text(text)
            name = _extract_name_from_front_text(text)  # layout rule from attached card
            dob = _extract_dob_from_text(text)
            gender = _extract_gender_from_text(text)
            address = _extract_address_from_back_text(text)  # if back side included in image
            pin_code = _extract_pin_from_text(text)

            if not aadhaar_num:
                validation_errors.append("Aadhaar number not found.")
            else:
                if not (aadhaar_num.isdigit() and len(aadhaar_num) == 12):
                    validation_errors.append("Aadhaar number must be a 12-digit numeric value.")
                elif not _verhoeff_check(aadhaar_num):
                    validation_errors.append("Aadhaar number failed Verhoeff checksum validation.")

            if vid_num and not (vid_num.isdigit() and len(vid_num) == 16):
                validation_errors.append("VID must be a 16-digit numeric value.")

            if not name:
                validation_errors.append("Name not found.")
            if not dob:
                validation_errors.append("DOB not found.")
            if not gender:
                validation_errors.append("Gender not found.")

            extracted = {
                "name": name,
                "aadhaar_number": _mask_aadhaar(aadhaar_num) if aadhaar_num else None,
                "dob": dob,
                "gender": gender,
                "address": address,
                "pin_code": pin_code,
                "vid": _mask_vid(vid_num) if vid_num else None,
            }
            extracted = {k: v for k, v in extracted.items() if v is not None}

            return OCRResponse(
                document_type=document_type,
                extracted_data=extracted,
                confidence=0.8,
                is_valid=(len(validation_errors) == 0),
                validation_errors=validation_errors,
            )
        
        # Try Gemini Vision for OCR
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent  # type: ignore[import-not-found]
            
            api_key = os.environ.get('EMERGENT_LLM_KEY', '')
            
            if api_key:
                chat = LlmChat(
                    api_key=api_key,
                    session_id=f"ocr_{uuid.uuid4()}",
                    system_message=f"You are an OCR expert. Extract all text and fields from the {document_type} document image. Return a JSON object with the extracted fields."
                )
                chat.with_model("gemini", "gemini-2.5-flash")
                
                image_content_obj = ImageContent(image_base64=image_base64)
                user_message = UserMessage(
                    text=f"Extract all information from this {document_type} document. Return JSON with fields like name, number, dates, address etc.",
                    file_contents=[image_content_obj]
                )
                
                response = await chat.send_message(user_message)
                
                # Parse response
                try:
                    extracted = json.loads(response)
                except:
                    extracted = {"raw_text": response}
                
                return OCRResponse(
                    document_type=document_type,
                    extracted_data=extracted,
                    confidence=0.92,
                    is_valid=True,
                    validation_errors=[]
                )
        except Exception as ocr_error:
            logger.warning(f"Gemini OCR not available: {ocr_error}")

        # No mocks: require real QR decode (Aadhaar) or OCR key for Gemini.
        raise HTTPException(
            status_code=503,
            detail=(
                "OCR is not configured. For Aadhaar, install Tesseract (open-source OCR) "
                "or set EMERGENT_LLM_KEY to enable Gemini OCR."
            ),
        )
    except Exception as e:
        logger.error(f"OCR error: {e}")
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(status_code=500, detail=str(e))

# ===================== AADHAAR VERIFICATION ENDPOINTS =====================
_VERHOEFF_D = [
    [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
    [1, 2, 3, 4, 0, 6, 7, 8, 9, 5],
    [2, 3, 4, 0, 1, 7, 8, 9, 5, 6],
    [3, 4, 0, 1, 2, 8, 9, 5, 6, 7],
    [4, 0, 1, 2, 3, 9, 5, 6, 7, 8],
    [5, 9, 8, 7, 6, 0, 4, 3, 2, 1],
    [6, 5, 9, 8, 7, 1, 0, 4, 3, 2],
    [7, 6, 5, 9, 8, 2, 1, 0, 4, 3],
    [8, 7, 6, 5, 9, 3, 2, 1, 0, 4],
    [9, 8, 7, 6, 5, 4, 3, 2, 1, 0],
]
_VERHOEFF_P = [
    [0, 1, 2, 3, 4, 5, 6, 7, 8, 9],
    [1, 5, 7, 6, 2, 8, 3, 0, 9, 4],
    [5, 8, 0, 3, 7, 9, 6, 1, 4, 2],
    [8, 9, 1, 6, 0, 4, 3, 5, 2, 7],
    [9, 4, 5, 3, 1, 2, 6, 8, 7, 0],
    [4, 2, 8, 6, 5, 7, 3, 9, 0, 1],
    [2, 7, 9, 3, 8, 0, 6, 4, 1, 5],
    [7, 0, 4, 6, 9, 1, 3, 2, 5, 8],
]
_VERHOEFF_INV = [0, 4, 3, 2, 1, 5, 6, 7, 8, 9]

def _verhoeff_check(num: str) -> bool:
    """
    Return True if `num` passes Verhoeff checksum (used by Aadhaar).
    `num` must be digits only.
    """
    if not num or not num.isdigit():
        return False
    c = 0
    # Validate (includes checksum digit)
    for i, ch in enumerate(reversed(num)):
        c = _VERHOEFF_D[c][_VERHOEFF_P[i % 8][int(ch)]]
    return c == 0

def _mask_aadhaar(num: str) -> str:
    if not num or len(num) < 4:
        return "XXXX XXXX XXXX"
    last4 = num[-4:]
    return f"XXXX XXXX {last4}"

def _mask_vid(num: str) -> str:
    if not num or len(num) < 4:
        return "XXXX XXXX XXXX XXXX"
    last4 = num[-4:]
    return f"XXXX XXXX XXXX {last4}"

def _normalize_text(s: str) -> str:
    s = (s or "").replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def _extract_vid_from_text(text: str) -> Optional[str]:
    t = _normalize_text(text)
    m = re.search(r"\bVID\s*[:\-]?\s*((?:\d{4}\s*){4})\b", t, flags=re.IGNORECASE)
    if not m:
        return None
    num = re.sub(r"\s+", "", m.group(1))
    return num if (len(num) == 16 and num.isdigit()) else None

def _extract_dob_from_text(text: str) -> Optional[str]:
    t = _normalize_text(text)
    m = re.search(r"\bDOB\s*[:\-]?\s*([0-3]?\d[\/\-][01]?\d[\/\-]\d{4})\b", t, flags=re.IGNORECASE)
    if not m:
        return None
    raw = m.group(1).replace("-", "/")
    try:
        dt = date_parser.parse(raw, dayfirst=True)
        return dt.date().isoformat()
    except Exception:
        return raw

def _extract_gender_from_text(text: str) -> Optional[str]:
    t = _normalize_text(text).upper()
    if re.search(r"\bMALE\b", t) or "पुरुष" in t:
        return "Male"
    if re.search(r"\bFEMALE\b", t) or "महिला" in t:
        return "Female"
    return None

def _extract_name_from_front_text(front_text: str) -> Optional[str]:
    """
    Attached Aadhaar front format has Name on the line above 'DOB:'.
    """
    t = _normalize_text(front_text)
    if not t:
        return None
    lines = [ln.strip() for ln in t.split("\n") if ln.strip()]
    dob_idx = None
    for i, ln in enumerate(lines):
        if re.search(r"\bDOB\b", ln, flags=re.IGNORECASE):
            dob_idx = i
            break
    if dob_idx is None:
        return None
    # pick previous non-empty, non-header line
    boiler = re.compile(r"(government of india|unique identification authority|uidai|aadhaar|भारत सरकार|भारत)", re.IGNORECASE)
    for j in range(dob_idx - 1, max(-1, dob_idx - 6), -1):
        cand = lines[j]
        if boiler.search(cand):
            continue
        if len(cand) >= 3:
            return cand
    return None

def _extract_address_from_back_text(back_text: str) -> Optional[str]:
    """
    Attached Aadhaar back format includes:
      Address:
      <multi-line address>
      <State> - <PIN>
    """
    t = _normalize_text(back_text)
    if not t:
        return None
    m = re.search(r"\bAddress\s*:\s*(.+)", t, flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    blob = m.group(1)
    # stop at UIDAI footer markers / or Aadhaar number line
    for sm in [
        r"\bUnique Identification Authority\b",
        r"\bUIDAI\b",
        r"\bhelp@uidai\.gov\.in\b",
        r"\bwww\.uidai\.gov\.in\b",
        r"\bAadhaar is proof of identity\b",
        r"\bVID\b",
        r"\b(\d{4}\s?\d{4}\s?\d{4})\b",
    ]:
        mm = re.search(sm, blob, flags=re.IGNORECASE)
        if mm:
            blob = blob[: mm.start()]
            break
    # trim after PIN if present
    pin_m = re.search(r"\b(\d{6})\b", blob)
    if pin_m:
        blob = blob[: pin_m.end()]
    blob = _normalize_text(blob)
    return blob or None

def _extract_pin_from_text(text: str) -> Optional[str]:
    t = _normalize_text(text)
    m = re.search(r"\b(\d{6})\b", t)
    return m.group(1) if m else None

def _parse_aadhaar_printletter_xml(xml_text: str) -> Optional[Dict[str, Any]]:
    """
    Parse the XML embedded in Aadhaar QR if provided by an upstream scanner.
    """
    if not xml_text:
        return None
    s = str(xml_text).strip()
    if "PrintLetterBarcodeData" not in s:
        return None
    try:
        import xml.etree.ElementTree as ET
        root = ET.fromstring(s)
        attrs = root.attrib or {}
        return {
            "name": (attrs.get("name") or "").strip() or None,
            "aadhaar_number": (attrs.get("uid") or "").strip() or None,
            "gender": (attrs.get("gender") or "").strip() or None,
            "dob": (attrs.get("dob") or "").strip() or None,
            "yob": (attrs.get("yob") or "").strip() or None,
            "pin_code": (attrs.get("pc") or "").strip() or None,
        }
    except Exception:
        return None

def _extract_aadhaar_number(payload: Any) -> Optional[str]:
    """
    Extract a 12-digit Aadhaar number from:
    - parsed JSON dict from LLM OCR, or
    - raw text fields
    """
    candidates: List[str] = []

    def _collect_from_text(text: str):
        if not text:
            return
        # common print format: "1234 5678 9012" or "123456789012"
        for m in re.findall(r"\b(\d{4}\s?\d{4}\s?\d{4})\b", text):
            candidates.append(re.sub(r"\s+", "", m))
        for m in re.findall(r"\b(\d{12})\b", text):
            candidates.append(m)

    if isinstance(payload, dict):
        for k in ["aadhaar_number", "aadhar_number", "uid", "uidai", "number", "aadhaar"]:
            v = payload.get(k)
            if isinstance(v, (str, int)):
                _collect_from_text(str(v))
        # also scan any "raw_text" field if present
        rt = payload.get("raw_text")
        if isinstance(rt, str):
            _collect_from_text(rt)
        # finally scan all string values (best-effort)
        for v in payload.values():
            if isinstance(v, str):
                _collect_from_text(v)
    elif isinstance(payload, str):
        _collect_from_text(payload)

    # de-dupe while preserving order
    seen = set()
    uniq = []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            uniq.append(c)

    # prefer a Verhoeff-valid candidate
    for c in uniq:
        if len(c) == 12 and _verhoeff_check(c):
            return c
    # else return first 12-digit candidate (invalid checksum)
    for c in uniq:
        if len(c) == 12 and c.isdigit():
            return c
    return None

@aadhaar_router.post("/verify", response_model=AadhaarVerificationResponse)
async def verify_aadhaar_card(image_file: UploadFile = File(...)):
    """
    Aadhaar Card verification API (OCR-based).
    - Extracts fields from the uploaded Aadhaar image
    - Validates Aadhaar number format + Verhoeff checksum
    NOTE: This is NOT UIDAI OTP/eKYC verification; it's document OCR + basic validation.
    """
    try:
        image_content = await image_file.read()
        image_base64 = base64.b64encode(image_content).decode("utf-8")

        extracted: Dict[str, Any] = {}
        confidence = 0.0
        validation_errors: List[str] = []

        # Try Gemini Vision via emergentintegrations (same pattern as /ocr/verify)
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent  # type: ignore[import-not-found]

            api_key = os.environ.get("EMERGENT_LLM_KEY", "")
            if api_key:
                chat = LlmChat(
                    api_key=api_key,
                    session_id=f"aadhaar_{uuid.uuid4()}",
                    system_message=(
                        "You are an OCR + document parsing expert. Extract Aadhaar card fields and return STRICT JSON. "
                        "Keys: name, aadhaar_number, dob_or_yob, gender, address. "
                        "If a field is missing, set it to null. Do not include extra commentary."
                    ),
                )
                chat.with_model("gemini", "gemini-2.5-flash")

                image_content_obj = ImageContent(image_base64=image_base64)
                user_message = UserMessage(
                    text="Extract Aadhaar card details. Return strict JSON only.",
                    file_contents=[image_content_obj],
                )
                resp = await chat.send_message(user_message)

                try:
                    extracted = json.loads(resp) if isinstance(resp, str) else {"raw_text": str(resp)}
                except Exception:
                    extracted = {"raw_text": resp}
                confidence = 0.92
        except Exception as ocr_error:
            logger.warning(f"Aadhaar OCR (Gemini) not available: {ocr_error}")

        # No mocks: if OCR isn't configured, return a clear error.
        if not extracted:
            raise HTTPException(
                status_code=503,
                detail="Aadhaar image extraction requires OCR/QR decode. If you can't use OCR/OpenCV, use /api/aadhaar/verify-rules with extracted text instead.",
            )

        aadhaar_num = _extract_aadhaar_number(extracted)
        if not aadhaar_num:
            validation_errors.append("Aadhaar number not found in document.")
        else:
            if len(aadhaar_num) != 12 or (not aadhaar_num.isdigit()):
                validation_errors.append("Aadhaar number must be a 12-digit numeric value.")
            elif not _verhoeff_check(aadhaar_num):
                validation_errors.append("Aadhaar number failed Verhoeff checksum validation.")

        # Basic field checks (best-effort; don't hard-fail if OCR is imperfect)
        name = extracted.get("name") if isinstance(extracted, dict) else None
        if not name or not str(name).strip():
            validation_errors.append("Name not found.")

        is_valid = len(validation_errors) == 0
        last4 = aadhaar_num[-4:] if aadhaar_num and len(aadhaar_num) >= 4 else None
        masked = _mask_aadhaar(aadhaar_num) if aadhaar_num else None

        # Always avoid returning full Aadhaar number; overwrite if present.
        if isinstance(extracted, dict):
            extracted = dict(extracted)
            if "aadhaar_number" in extracted:
                extracted["aadhaar_number"] = masked or extracted.get("aadhaar_number")
            elif "aadhar_number" in extracted:
                extracted["aadhar_number"] = masked or extracted.get("aadhar_number")

        return AadhaarVerificationResponse(
            extracted_data=extracted,
            confidence=confidence,
            is_valid=is_valid,
            validation_errors=validation_errors,
            aadhaar_number_last4=last4,
            aadhaar_number_masked=masked,
        )
    except Exception as e:
        logger.error(f"Aadhaar verification error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@aadhaar_router.post("/verify-rules", response_model=AadhaarRulesVerifyResponse)
async def verify_aadhaar_rules(request: AadhaarRulesVerifyRequest):
    """
    Rule-based Aadhaar verification derived from the attached Aadhaar letter/card layout.
    This endpoint DOES NOT accept images and DOES NOT perform OCR/OpenCV.
    It parses provided text/QR-XML and validates:
    - Aadhaar number format + Verhoeff checksum
    - VID format (16 digits)
    - DOB format
    - Gender presence
    - Address/PIN presence (from back_text when provided)
    """
    validation_errors: List[str] = []

    front_text = _normalize_text(request.front_text or "")
    back_text = _normalize_text(request.back_text or "")

    extracted: Dict[str, Any] = {}

    # Prefer QR XML if provided (most reliable, still no image processing here)
    if request.qr_xml:
        x = _parse_aadhaar_printletter_xml(request.qr_xml) or {}
        extracted.update({k: v for k, v in x.items() if v})

    # Text parsing based on attached layout
    if front_text:
        extracted.setdefault("name", _extract_name_from_front_text(front_text))
        extracted.setdefault("dob", _extract_dob_from_text(front_text))
        extracted.setdefault("gender", _extract_gender_from_text(front_text))
        # Aadhaar number often printed on front too
        extracted.setdefault("aadhaar_number_raw", _extract_aadhaar_number(front_text))
        extracted.setdefault("vid_raw", _extract_vid_from_text(front_text))

    if back_text:
        extracted.setdefault("address", _extract_address_from_back_text(back_text))
        extracted.setdefault("pin_code", _extract_pin_from_text(back_text))
        extracted.setdefault("aadhaar_number_raw", _extract_aadhaar_number(back_text) or extracted.get("aadhaar_number_raw"))
        extracted.setdefault("vid_raw", _extract_vid_from_text(back_text) or extracted.get("vid_raw"))

    aadhaar_num = (extracted.get("aadhaar_number") or extracted.get("aadhaar_number_raw") or "").strip()
    aadhaar_num = re.sub(r"\s+", "", str(aadhaar_num))
    vid_num = (extracted.get("vid") or extracted.get("vid_raw") or "").strip()
    vid_num = re.sub(r"\s+", "", str(vid_num))

    if not aadhaar_num:
        validation_errors.append("Aadhaar number not found.")
    else:
        if not (aadhaar_num.isdigit() and len(aadhaar_num) == 12):
            validation_errors.append("Aadhaar number must be a 12-digit numeric value.")
        elif not _verhoeff_check(aadhaar_num):
            validation_errors.append("Aadhaar number failed Verhoeff checksum validation.")

    if vid_num:
        if not (vid_num.isdigit() and len(vid_num) == 16):
            validation_errors.append("VID must be a 16-digit numeric value.")

    if not extracted.get("name"):
        validation_errors.append("Name not found (expected on line above DOB in front text).")
    if not extracted.get("dob") and not extracted.get("yob"):
        validation_errors.append("DOB/YOB not found.")
    if not extracted.get("gender"):
        validation_errors.append("Gender not found.")
    if request.back_text and not extracted.get("address"):
        validation_errors.append("Address not found in back text (expected after 'Address:').")
    if request.back_text and not extracted.get("pin_code"):
        validation_errors.append("PIN code not found in back text.")

    # Mask sensitive values in response
    aadhaar_last4 = aadhaar_num[-4:] if aadhaar_num and len(aadhaar_num) >= 4 else None
    aadhaar_masked = _mask_aadhaar(aadhaar_num) if aadhaar_num else None
    vid_last4 = vid_num[-4:] if vid_num and len(vid_num) >= 4 else None
    vid_masked = _mask_vid(vid_num) if vid_num else None

    extracted_out = dict(extracted)
    extracted_out.pop("aadhaar_number_raw", None)
    extracted_out.pop("vid_raw", None)
    if aadhaar_masked:
        extracted_out["aadhaar_number"] = aadhaar_masked
    if vid_masked:
        extracted_out["vid"] = vid_masked

    return AadhaarRulesVerifyResponse(
        is_valid=(len(validation_errors) == 0),
        validation_errors=validation_errors,
        extracted_data=extracted_out,
        aadhaar_number_last4=aadhaar_last4,
        aadhaar_number_masked=aadhaar_masked,
        vid_last4=vid_last4,
        vid_masked=vid_masked,
    )

def _normalize_name(name: str) -> str:
    """Normalize name for comparison (remove extra spaces, convert to uppercase)"""
    if not name:
        return ""
    # Remove extra spaces, convert to uppercase, remove special characters except spaces
    normalized = re.sub(r'[^\w\s]', '', str(name).upper())
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized

def _normalize_date(date_str: str) -> Optional[str]:
    """Normalize date string to YYYY-MM-DD format for comparison"""
    if not date_str:
        return None
    try:
        # Try parsing various date formats
        date_obj = date_parser.parse(str(date_str), dayfirst=True)
        return date_obj.date().isoformat()
    except Exception:
        # If parsing fails, try to extract date pattern
        date_match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', str(date_str))
        if date_match:
            day, month, year = date_match.groups()
            try:
                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            except Exception:
                pass
        return None

def _normalize_gender(gender: str) -> str:
    """Normalize gender for comparison"""
    if not gender:
        return ""
    gender_lower = str(gender).lower().strip()
    if gender_lower in ['male', 'm', 'पुरुष', 'पुरूष']:
        return "Male"
    elif gender_lower in ['female', 'f', 'महिला', 'स्त्री']:
        return "Female"
    elif gender_lower in ['other', 'o', 'transgender', 'trans']:
        return "Other"
    return str(gender).strip()

def _normalize_aadhaar_number(aadhaar: str) -> str:
    """Normalize Aadhaar number (remove spaces)"""
    if not aadhaar:
        return ""
    return re.sub(r'\s+', '', str(aadhaar).strip())

def _compare_fields(entered: str, extracted: Optional[str], field_type: str = "text") -> bool:
    """Compare entered value with extracted value"""
    if not entered:
        return False
    if not extracted:
        return False
    
    if field_type == "name":
        return _normalize_name(entered) == _normalize_name(extracted)
    elif field_type == "dob":
        entered_norm = _normalize_date(entered)
        extracted_norm = _normalize_date(extracted)
        return entered_norm is not None and extracted_norm is not None and entered_norm == extracted_norm
    elif field_type == "gender":
        return _normalize_gender(entered) == _normalize_gender(extracted)
    elif field_type == "aadhaar_number":
        return _normalize_aadhaar_number(entered) == _normalize_aadhaar_number(extracted)
    else:
        # Default text comparison
        return str(entered).strip().lower() == str(extracted).strip().lower()

@aadhaar_router.post("/verify-with-form", response_model=AadhaarFormVerifyResponse)
async def verify_aadhaar_with_form(
    name: str = Form(...),
    dob: str = Form(...),
    aadhaar_number: str = Form(...),
    gender: str = Form(...),
    image_file: UploadFile = File(...)
):
    """
    Aadhaar verification with form inputs and document upload.
    - Accepts form inputs: Name, DoB, Aadhaar Number, Gender
    - Uploads Aadhaar image/PDF
    - Extracts details from uploaded document using OCR
    - Compares entered details with extracted details
    - Returns verification result with field-by-field comparison
    """
    try:
        # Read and process the uploaded file
        file_content = await image_file.read()
        file_extension = image_file.filename.split('.')[-1].lower() if image_file.filename else ''
        
        # Convert PDF to image if needed (basic support)
        image_bytes = file_content
        if file_extension == 'pdf':
            try:
                from pdf2image import convert_from_bytes
                images = convert_from_bytes(file_content)
                if images:
                    img_io = io.BytesIO()
                    images[0].save(img_io, format='PNG')
                    image_bytes = img_io.getvalue()
            except Exception as pdf_error:
                logger.warning(f"PDF conversion failed: {pdf_error}")
                raise HTTPException(
                    status_code=400,
                    detail="PDF processing not available. Please upload an image file (JPG, PNG)."
                )
        
        image_base64 = base64.b64encode(image_bytes).decode("utf-8")
        
        # Extract details from image using open-source OCR
        extracted: Dict[str, Any] = {}
        confidence = 0.0
        validation_errors: List[str] = []
        ocr_text = ""
        
        # Preprocess image for faster OCR (do this once, use for both OCR engines)
        if Image is None:
            raise HTTPException(
                status_code=503,
                detail="OCR processing requires PIL/Pillow. Please install: pip install pillow pytesseract"
            )
        
        img = Image.open(io.BytesIO(image_bytes))
        
        # Aggressive image optimization for faster OCR
        max_size = 1500  # Reduced from 2000 for faster processing
        width, height = img.size
        
        # Resize if image is too large
        if width > max_size or height > max_size:
            scale = min(max_size / width, max_size / height)
            new_width = int(width * scale)
            new_height = int(height * scale)
            img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            logger.info(f"Resized image from {width}x{height} to {new_width}x{new_height} for faster OCR")
        
        # Convert to grayscale for faster OCR (color not needed for text extraction)
        if img.mode != 'L':
            img = img.convert('L')
        
        # Enhance contrast for better OCR accuracy
        try:
            if ImageEnhance is not None:
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(1.5)  # Increase contrast by 50%
        except Exception:
            pass  # Continue without enhancement if it fails
        
        # Use Tesseract directly (much faster than EasyOCR)
        # Skip EasyOCR entirely for speed - Tesseract is sufficient for Aadhaar cards
        try:
            global pytesseract
            if pytesseract is None:
                import pytesseract as pt
                pytesseract = pt
                # Set Tesseract path for macOS Homebrew installation
                import os
                tesseract_paths = [
                    '/opt/homebrew/bin/tesseract',  # Homebrew on Apple Silicon
                    '/usr/local/bin/tesseract',     # Homebrew on Intel Mac
                    '/usr/bin/tesseract',           # System installation
                ]
                for path in tesseract_paths:
                    if os.path.exists(path):
                        pytesseract.pytesseract.tesseract_cmd = path
                        logger.info(f"Tesseract found at: {path}")
                        break
            
            # Use optimized Tesseract config for faster processing
            # PSM 6: Assume uniform block of text (faster)
            # OEM 3: Default OCR engine mode
            custom_config = r'--oem 3 --psm 6'
            
            # Perform OCR with English (faster than English+Hindi)
            # Aadhaar cards have English text, so English-only is sufficient
            try:
                ocr_text = pytesseract.image_to_string(img, lang='eng', config=custom_config)
            except Exception as lang_error:
                logger.warning(f"Tesseract with config failed: {lang_error}, trying default")
                # Fallback to default config
                ocr_text = pytesseract.image_to_string(img, lang='eng')
            
            extracted = {"raw_text": ocr_text}
            confidence = 0.70
            logger.info(f"Tesseract extracted text: {ocr_text[:200]}...")
        except Exception as tesseract_error:
            logger.error(f"Tesseract OCR failed: {tesseract_error}")
            raise HTTPException(
                status_code=503,
                detail="OCR processing is not available. Please install: pip install pillow pytesseract. Make sure Tesseract is installed on your system."
            )
        
        # Extract structured data from OCR result
        # Since we're using open-source OCR, we extract from raw_text using regex patterns
        raw_text = extracted.get("raw_text", "")
        extracted_name = _extract_name_from_front_text(raw_text) or ""
        extracted_dob = _extract_dob_from_text(raw_text) or ""
        extracted_aadhaar = _extract_aadhaar_number(raw_text) or ""
        extracted_gender = _extract_gender_from_text(raw_text) or ""
        
        # Also try to extract from QR code if present in the image
        try:
            if cv2 is not None and np is not None:
                # Decode QR code from image
                qr_text = _decode_qr_text_from_image_bytes(image_bytes)
                if qr_text:
                    qr_data = _parse_aadhaar_qr_payload(qr_text)
                    if qr_data:
                        # QR code data is more reliable, use it if available
                        extracted_name = extracted_name or qr_data.get("name", "")
                        extracted_dob = extracted_dob or qr_data.get("dob", "") or qr_data.get("yob", "")
                        extracted_aadhaar = extracted_aadhaar or qr_data.get("uid", "")
                        extracted_gender = extracted_gender or qr_data.get("gender", "")
                        confidence = max(confidence, 0.90)  # QR code is more reliable
                        logger.info("QR code data extracted successfully")
        except Exception as qr_error:
            logger.warning(f"QR code extraction failed: {qr_error}")
        
        # Normalize extracted Aadhaar number
        extracted_aadhaar = _normalize_aadhaar_number(extracted_aadhaar)
        
        # Compare fields
        name_matches = _compare_fields(name, extracted_name, "name")
        dob_matches = _compare_fields(dob, extracted_dob, "dob")
        aadhaar_matches = _compare_fields(aadhaar_number, extracted_aadhaar, "aadhaar_number")
        gender_matches = _compare_fields(gender, extracted_gender, "gender")
        
        # Build field comparisons
        field_comparisons = [
            FieldComparison(
                field_name="Name",
                entered_value=name,
                extracted_value=extracted_name or "Not found",
                matches=name_matches,
                confidence=confidence if name_matches else None
            ),
            FieldComparison(
                field_name="Date of Birth",
                entered_value=dob,
                extracted_value=extracted_dob or "Not found",
                matches=dob_matches,
                confidence=confidence if dob_matches else None
            ),
            FieldComparison(
                field_name="Aadhaar Number",
                entered_value=_mask_aadhaar(aadhaar_number),
                extracted_value=_mask_aadhaar(extracted_aadhaar) if extracted_aadhaar else "Not found",
                matches=aadhaar_matches,
                confidence=confidence if aadhaar_matches else None
            ),
            FieldComparison(
                field_name="Gender",
                entered_value=gender,
                extracted_value=extracted_gender or "Not found",
                matches=gender_matches,
                confidence=confidence if gender_matches else None
            ),
        ]
        
        # Validate Aadhaar number format
        normalized_aadhaar = _normalize_aadhaar_number(aadhaar_number)
        if not normalized_aadhaar.isdigit() or len(normalized_aadhaar) != 12:
            validation_errors.append("Entered Aadhaar number must be 12 digits.")
        elif not _verhoeff_check(normalized_aadhaar):
            validation_errors.append("Entered Aadhaar number failed Verhoeff checksum validation.")
        
        if extracted_aadhaar:
            if not extracted_aadhaar.isdigit() or len(extracted_aadhaar) != 12:
                validation_errors.append("Extracted Aadhaar number is not 12 digits.")
            elif not _verhoeff_check(extracted_aadhaar):
                validation_errors.append("Extracted Aadhaar number failed Verhoeff checksum validation.")
        
        # Determine overall verification status
        all_fields_match = name_matches and dob_matches and aadhaar_matches and gender_matches
        is_verified = all_fields_match and len(validation_errors) == 0
        
        # Build message
        if is_verified:
            message = "✅ All details are matching and verified!"
        else:
            mismatches = []
            if not name_matches:
                mismatches.append("Name")
            if not dob_matches:
                mismatches.append("Date of Birth")
            if not aadhaar_matches:
                mismatches.append("Aadhaar Number")
            if not gender_matches:
                mismatches.append("Gender")
            
            if mismatches:
                message = f"❌ Aadhaar details entered are not matching with the details in Aadhaar. Mismatched fields: {', '.join(mismatches)}. Please enter the details as per uploaded Aadhaar or check the uploaded Aadhaar."
            else:
                message = "⚠️ Some validation errors found. Please check the details."
        
        # Mask Aadhaar numbers in extracted data
        extracted_data_safe = dict(extracted)
        if extracted_aadhaar:
            extracted_data_safe["aadhaar_number"] = _mask_aadhaar(extracted_aadhaar)
        if "aadhar_number" in extracted_data_safe:
            extracted_data_safe["aadhar_number"] = _mask_aadhaar(extracted_aadhaar) if extracted_aadhaar else extracted_data_safe.get("aadhar_number")
        
        # Remove raw_text if present (too verbose)
        extracted_data_safe.pop("raw_text", None)
        
        last4 = normalized_aadhaar[-4:] if normalized_aadhaar and len(normalized_aadhaar) >= 4 else None
        
        return AadhaarFormVerifyResponse(
            is_verified=is_verified,
            message=message,
            field_comparisons=field_comparisons,
            extracted_data=extracted_data_safe,
            validation_errors=validation_errors,
            aadhaar_number_last4=last4,
            aadhaar_number_masked=_mask_aadhaar(normalized_aadhaar) if normalized_aadhaar else None,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Aadhaar form verification error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Verification failed: {str(e)}")

# ===================== OCR READER CACHE =====================
_easyocr_reader = None

def _get_easyocr_reader():
    """Get or initialize EasyOCR reader (cached for performance)"""
    global _easyocr_reader
    if _easyocr_reader is None:
        try:
            import easyocr
            # Initialize reader with English and Hindi support
            # This may take a few seconds on first run (downloads models)
            _easyocr_reader = easyocr.Reader(['en', 'hi'], gpu=False)
            logger.info("EasyOCR reader initialized successfully")
        except ImportError:
            logger.warning("EasyOCR not available")
            _easyocr_reader = False  # Mark as unavailable
        except Exception as e:
            logger.error(f"Failed to initialize EasyOCR: {e}")
            _easyocr_reader = False
    return _easyocr_reader if _easyocr_reader is not False else None

# ===================== FACIAL RECOGNITION ENDPOINTS =====================
_face_detector = None
_face_recognizer = None

def _require_opencv_face() -> None:
    if cv2 is None or np is None:
        raise HTTPException(
            status_code=503,
            detail="OpenCV is not available. Install backend deps: `pip install -r backend/requirements.txt` (needs opencv-contrib-python).",
        )

def _get_face_models():
    """
    Lazy-load OpenCV face detector + recognizer.
    Uses YuNet for detection and SFace for recognition (both ONNX).
    """
    global _face_detector, _face_recognizer
    _require_opencv_face()

    if _face_detector is not None and _face_recognizer is not None:
        return _face_detector, _face_recognizer

    models_dir = ROOT_DIR / "models"
    det_path = Path(os.environ.get("FACE_DETECTOR_MODEL_PATH", str(models_dir / "face_detection_yunet_2022mar.onnx")))
    rec_path = Path(os.environ.get("FACE_RECOGNIZER_MODEL_PATH", str(models_dir / "face_recognition_sface_2021dec.onnx")))

    missing = [str(p) for p in (det_path, rec_path) if not p.exists()]
    if missing:
        raise HTTPException(
            status_code=503,
            detail=(
                "OpenCV face models not found. Place the ONNX files under `backend/models/` "
                "or set FACE_DETECTOR_MODEL_PATH / FACE_RECOGNIZER_MODEL_PATH. Missing: "
                + ", ".join(missing)
            ),
        )

    # Detector input size will be set per-image via setInputSize()
    score_thr = float(os.environ.get("FACE_DETECTOR_SCORE_THRESHOLD", "0.9"))
    nms_thr = float(os.environ.get("FACE_DETECTOR_NMS_THRESHOLD", "0.3"))
    top_k = int(os.environ.get("FACE_DETECTOR_TOPK", "5000"))
    _face_detector = cv2.FaceDetectorYN.create(str(det_path), "", (320, 320), score_thr, nms_thr, top_k)
    _face_recognizer = cv2.FaceRecognizerSF.create(str(rec_path), "")
    return _face_detector, _face_recognizer

def _decode_image_bytes(image_bytes: bytes):
    _require_opencv_face()
    arr = np.frombuffer(image_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    return img

def _pick_largest_face(faces):
    # faces is Nx15 (YuNet); pick largest bbox area (w*h)
    if faces is None or len(faces) == 0:
        return None, 0
    areas = faces[:, 2] * faces[:, 3]
    idx = int(np.argmax(areas))
    return faces[idx], int(faces.shape[0])

def _face_box(face_row) -> Dict[str, float]:
    # face_row: [x, y, w, h, ...]
    return {
        "x": float(face_row[0]),
        "y": float(face_row[1]),
        "w": float(face_row[2]),
        "h": float(face_row[3]),
    }

def _extract_face_feature(img, detector, recognizer):
    h, w = img.shape[:2]
    detector.setInputSize((w, h))
    _, faces = detector.detect(img)
    if faces is None or len(faces) == 0:
        return None, None, 0
    face_row, count = _pick_largest_face(faces)
    if face_row is None:
        return None, None, count
    aligned = recognizer.alignCrop(img, face_row)
    feat = recognizer.feature(aligned)
    return feat, face_row, count

@facial_router.post("/verify", response_model=FacialVerificationResponse)
async def verify_face(
    reference_image: UploadFile = File(...),
    verify_image: UploadFile = File(...)
):
    """Verify face against reference"""
    try:
        detector, recognizer = _get_face_models()

        ref_content = await reference_image.read()
        verify_content = await verify_image.read()

        ref_img = _decode_image_bytes(ref_content)
        ver_img = _decode_image_bytes(verify_content)
        if ref_img is None or ver_img is None:
            raise HTTPException(status_code=422, detail="Invalid image file(s). Please upload valid image formats (jpg/png).")

        ref_feat, ref_face, ref_count = _extract_face_feature(ref_img, detector, recognizer)
        ver_feat, ver_face, ver_count = _extract_face_feature(ver_img, detector, recognizer)

        if ref_feat is None:
            raise HTTPException(status_code=422, detail="No face detected in reference_image.")
        if ver_feat is None:
            raise HTTPException(status_code=422, detail="No face detected in verify_image.")

        metric = os.environ.get("FACE_MATCH_METRIC", "cosine").lower().strip()
        if metric not in {"cosine", "l2"}:
            metric = "cosine"

        if metric == "l2":
            dist = float(recognizer.match(ref_feat, ver_feat, cv2.FaceRecognizerSF_FR_NORM_L2))
            threshold = float(os.environ.get("FACE_MATCH_THRESHOLD", "1.128"))
            is_match = dist <= threshold
            confidence = max(0.0, min(1.0, 1.0 - (dist / max(threshold, 1e-6))))
            similarity = 1.0 - dist  # informational
        else:
            sim = float(recognizer.match(ref_feat, ver_feat, cv2.FaceRecognizerSF_FR_COSINE))
            threshold = float(os.environ.get("FACE_MATCH_THRESHOLD", "0.363"))
            is_match = sim >= threshold
            confidence = max(0.0, min(1.0, sim))
            similarity = sim

        verification_id = str(uuid.uuid4())
        
        # Log verification
        await db.facial_verifications.insert_one({
            "verification_id": verification_id,
            "is_match": is_match,
            "confidence": float(round(confidence, 4)),
            "similarity": float(round(similarity, 4)),
            "metric": metric,
            "threshold": threshold,
            "detected_faces_ref": ref_count,
            "detected_faces_verify": ver_count,
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
        
        return FacialVerificationResponse(
            is_match=is_match,
            confidence=float(round(confidence, 4)),
            similarity=float(round(similarity, 4)),
            metric=metric,
            reference_face_box=_face_box(ref_face),
            verify_face_box=_face_box(ver_face),
            detected_faces_ref=ref_count,
            detected_faces_verify=ver_count,
            verification_id=verification_id,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Facial recognition error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== VEHICLE DETECTION ENDPOINTS =====================
_vehicle_net = None

_COCO80 = [
    "person","bicycle","car","motorcycle","airplane","bus","train","truck","boat","traffic light",
    "fire hydrant","stop sign","parking meter","bench","bird","cat","dog","horse","sheep","cow",
    "elephant","bear","zebra","giraffe","backpack","umbrella","handbag","tie","suitcase","frisbee",
    "skis","snowboard","sports ball","kite","baseball bat","baseball glove","skateboard","surfboard","tennis racket","bottle",
    "wine glass","cup","fork","knife","spoon","bowl","banana","apple","sandwich","orange",
    "broccoli","carrot","hot dog","pizza","donut","cake","chair","couch","potted plant","bed",
    "dining table","toilet","tv","laptop","mouse","remote","keyboard","cell phone","microwave","oven",
    "toaster","sink","refrigerator","book","clock","vase","scissors","teddy bear","hair drier","toothbrush",
]

def _require_opencv_vehicle() -> None:
    # Reuse the OpenCV availability check (cv2/numpy) but with vehicle-specific message.
    if cv2 is None or np is None:
        raise HTTPException(
            status_code=503,
            detail="OpenCV is not available. Install backend deps: `pip install -r backend/requirements.txt` (needs opencv-contrib-python).",
        )

def _vehicle_letterbox(img, new_shape: int = 640, color=(114, 114, 114)):
    """
    Letterbox resize while keeping aspect ratio (YOLO-style).
    Returns: resized image, scale ratio, (pad_w, pad_h)
    """
    h, w = img.shape[:2]
    r = min(new_shape / h, new_shape / w)
    new_unpad = (int(round(w * r)), int(round(h * r)))
    resized = cv2.resize(img, new_unpad, interpolation=cv2.INTER_LINEAR)
    dw = new_shape - new_unpad[0]
    dh = new_shape - new_unpad[1]
    dw /= 2
    dh /= 2
    top, bottom = int(round(dh - 0.1)), int(round(dh + 0.1))
    left, right = int(round(dw - 0.1)), int(round(dw + 0.1))
    out = cv2.copyMakeBorder(resized, top, bottom, left, right, cv2.BORDER_CONSTANT, value=color)
    return out, r, (left, top)

def _get_vehicle_net():
    global _vehicle_net
    _require_opencv_vehicle()
    if _vehicle_net is not None:
        return _vehicle_net

    models_dir = ROOT_DIR / "models"
    model_path = Path(os.environ.get("VEHICLE_DETECTOR_MODEL_PATH", str(models_dir / "vehicle_yolov8n.onnx")))
    if not model_path.exists():
        raise HTTPException(
            status_code=503,
            detail=(
                "Vehicle detector model not found. Place a YOLO ONNX model under `backend/models/` "
                "(default: `backend/models/vehicle_yolov8n.onnx`) or set VEHICLE_DETECTOR_MODEL_PATH."
            ),
        )

    net = cv2.dnn.readNetFromONNX(str(model_path))
    # Prefer CPU by default (portable); allow override
    backend = os.environ.get("OPENCV_DNN_BACKEND", "default").lower()
    target = os.environ.get("OPENCV_DNN_TARGET", "cpu").lower()
    if backend == "opencv":
        net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)
    if target == "cpu":
        net.setPreferableTarget(cv2.dnn.DNN_TARGET_CPU)
    _vehicle_net = net
    return _vehicle_net

def _decode_vehicle_output(outputs, conf_thres: float, iou_thres: float, input_size: int, ratio: float, pad):
    """
    Support common YOLO ONNX output formats:
    - (1, N, 6): [x1,y1,x2,y2,score,class] (already NMS'd)
    - (1, 84, 8400) (YOLOv8): [cx,cy,w,h] + class scores
    Returns list of (x,y,w,h,score,class_id) in original image pixels.
    """
    if isinstance(outputs, (list, tuple)):
        out = outputs[0]
    else:
        out = outputs

    out = np.array(out)
    dets = []

    # Case A: already NMS'd
    if out.ndim == 3 and out.shape[0] == 1 and out.shape[2] >= 6 and out.shape[1] >= 1:
        # Could be (1, N, 6) OR (1, 84, 8400). Disambiguate by last dim.
        if out.shape[2] == 6:
            rows = out[0]
            for r in rows:
                x1, y1, x2, y2, score, cls = r[:6]
                score = float(score)
                if score < conf_thres:
                    continue
                cls_id = int(cls)
                # map back from letterboxed input to original
                left, top = pad
                x1 = (float(x1) - left) / ratio
                y1 = (float(y1) - top) / ratio
                x2 = (float(x2) - left) / ratio
                y2 = (float(y2) - top) / ratio
                dets.append((x1, y1, max(0.0, x2 - x1), max(0.0, y2 - y1), score, cls_id))
            return dets

    # Case B: YOLOv8 raw output (1, 84, 8400) or (1, 85, 8400)
    if out.ndim == 3 and out.shape[0] == 1 and out.shape[1] >= 5:
        preds = out[0]
        # transpose to (num_preds, channels)
        preds = preds.transpose(1, 0)
        boxes = preds[:, :4]
        scores = preds[:, 4:]
        cls_ids = np.argmax(scores, axis=1)
        confs = scores[np.arange(scores.shape[0]), cls_ids]

        mask = confs >= conf_thres
        boxes = boxes[mask]
        confs = confs[mask]
        cls_ids = cls_ids[mask].astype(int)

        if boxes.shape[0] == 0:
            return []

        # cx,cy,w,h -> x,y,w,h in input space
        cx, cy, bw, bh = boxes[:, 0], boxes[:, 1], boxes[:, 2], boxes[:, 3]
        x = cx - bw / 2
        y = cy - bh / 2

        # NMS in input-space
        rects = np.stack([x, y, bw, bh], axis=1).tolist()
        conf_list = confs.astype(float).tolist()
        indices = cv2.dnn.NMSBoxes(rects, conf_list, conf_thres, iou_thres)
        if len(indices) == 0:
            return []
        if isinstance(indices, np.ndarray):
            indices = indices.flatten().tolist()
        elif isinstance(indices, (list, tuple)) and len(indices) and isinstance(indices[0], (list, tuple, np.ndarray)):
            indices = [int(i[0]) for i in indices]

        left, top = pad
        for i in indices:
            xi, yi, wi, hi = rects[i]
            score = float(conf_list[i])
            cls_id = int(cls_ids[i])
            # map back to original
            xo = (float(xi) - left) / ratio
            yo = (float(yi) - top) / ratio
            wo = float(wi) / ratio
            ho = float(hi) / ratio
            dets.append((xo, yo, max(0.0, wo), max(0.0, ho), score, cls_id))
        return dets

    return []

def _map_vehicle_class(coco_class_name: str) -> Optional[str]:
    c = coco_class_name.lower().strip()
    if c in {"motorcycle", "bicycle"}:
        return "Two Wheeler"
    if c == "car":
        return "Four Wheeler - LMV"
    if c == "bus":
        return "Bus"
    if c == "truck":
        return "Heavy Goods Vehicle"
    return None

@vehicle_router.post("/detect", response_model=VehicleDetectionResponse)
async def detect_vehicle(
    image_file: UploadFile = File(...)
):
    """Detect and classify vehicle from image"""
    try:
        net = _get_vehicle_net()
        image_content = await image_file.read()

        _require_opencv_vehicle()
        arr = np.frombuffer(image_content, dtype=np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            raise HTTPException(status_code=422, detail="Invalid image file. Please upload a valid jpg/png.")

        input_size = int(os.environ.get("VEHICLE_DETECTOR_INPUT_SIZE", "640"))
        conf_thres = float(os.environ.get("VEHICLE_DETECTOR_CONF_THRESHOLD", "0.35"))
        iou_thres = float(os.environ.get("VEHICLE_DETECTOR_NMS_IOU_THRESHOLD", "0.45"))

        lb, ratio, pad = _vehicle_letterbox(img, new_shape=input_size)
        blob = cv2.dnn.blobFromImage(lb, scalefactor=1.0 / 255.0, size=(input_size, input_size), swapRB=True, crop=False)
        net.setInput(blob)
        outputs = net.forward()

        dets = _decode_vehicle_output(outputs, conf_thres, iou_thres, input_size, ratio, pad)

        # Keep only vehicle-related detections
        vehicle_dets = []
        for x, y, w, h, score, cls_id in dets:
            if 0 <= cls_id < len(_COCO80):
                vclass = _map_vehicle_class(_COCO80[cls_id])
                if vclass:
                    vehicle_dets.append((x, y, w, h, score, cls_id, vclass))

        if not vehicle_dets:
            raise HTTPException(status_code=422, detail="No vehicle detected in the image.")

        # Choose best detection by score, then area
        vehicle_dets.sort(key=lambda r: (r[4], r[2] * r[3]), reverse=True)
        x, y, w, h, score, cls_id, vehicle_class = vehicle_dets[0]

        # Clamp bbox to image bounds
        H, W = img.shape[:2]
        x = max(0.0, min(float(x), float(W - 1)))
        y = max(0.0, min(float(y), float(H - 1)))
        w = max(1.0, min(float(w), float(W - x)))
        h = max(1.0, min(float(h), float(H - y)))

        return VehicleDetectionResponse(
            vehicle_class=vehicle_class,
            confidence=float(round(score, 4)),
            bounding_box={"x": int(round(x)), "y": int(round(y)), "width": int(round(w)), "height": int(round(h))},
            additional_info={
                "model": Path(os.environ.get("VEHICLE_DETECTOR_MODEL_PATH", "backend/models/vehicle_yolov8n.onnx")).name,
                "detector": "opencv-dnn-yolo-onnx",
                "coco_class": _COCO80[cls_id] if 0 <= cls_id < len(_COCO80) else str(cls_id),
                "image_width": W,
                "image_height": H,
                "thresholds": {"conf": conf_thres, "iou": iou_thres},
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Vehicle detection error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== EXECUTIVE DASHBOARD =====================
@dashboard_router.get("/executive-summary")
async def get_executive_summary(state_cd: Optional[str] = None, c_district: Optional[str] = None, city: Optional[str] = None):
    """Get executive summary KPIs"""
    try:
        now = datetime.now()
        match = _build_vahan_geo_match(state_cd=state_cd, c_district=c_district, city=city)

        # ================= VAHAN KPIs (data-driven) =================
        vahan_count = await db.vahan_data.count_documents(match)

        # Median vehicle value from sale_amt
        value_result = await db.vahan_data.aggregate(
            [
                {"$match": match},
                {"$match": {"sale_amt": {"$gt": 0, "$exists": True}}},
                {"$group": {"_id": None, "values": {"$push": "$sale_amt"}}},
            ]
        ).to_list(1)
        values = value_result[0]["values"] if value_result else []
        # Filter out invalid values (None, NaN, Inf, <= 0)
        valid_values = []
        for v in values:
            if v is None:
                continue
            try:
                fv = float(v)
                if math.isnan(fv) or math.isinf(fv) or fv <= 0:
                    continue
                valid_values.append(fv)
            except (ValueError, TypeError):
                continue
        
        # Use the proper _median function
        median_vehicle_value = _median(valid_values) if valid_values else 0.0
        
        # Debug logging to help diagnose issues
        if len(valid_values) > 0:
            logger.info(f"Median vehicle value calculation: {len(valid_values)} valid values, median={median_vehicle_value}, min={min(valid_values)}, max={max(valid_values)}, sample_values={valid_values[:10] if len(valid_values) >= 10 else valid_values}")
        else:
            logger.warning(f"No valid sale_amt values found for median calculation. Total values in result: {len(values)}")

        # Registration delay: avg days between purchase_dt and regn_dt where both parse
        # Limit to 10000 documents for performance
        docs_delay = await db.vahan_data.find(match, {"_id": 0, "regn_dt": 1, "purchase_dt": 1}).limit(10000).to_list(10000)
        delays = []
        for d in docs_delay:
            rd = _safe_parse_date(d.get("regn_dt"))
            pd_ = _safe_parse_date(d.get("purchase_dt"))
            if rd and pd_:
                delays.append(abs((rd - pd_).days))
        avg_registration_delay = round((sum(delays) / len(delays)) if delays else 0.0, 1)

        # Compliance: expired vs expiring soon (regn_upto)
        # Limit to 10000 documents for performance
        docs_upto = await db.vahan_data.find(match, {"_id": 0, "regn_upto": 1}).limit(10000).to_list(10000)
        expired = 0
        expiring_soon = 0
        active = 0
        for d in docs_upto:
            ru = _safe_parse_date(d.get("regn_upto"))
            if not ru:
                continue
            if ru < now:
                expired += 1
            else:
                active += 1
                if ru <= (now + timedelta(days=30)):
                    expiring_soon += 1
        active_registrations_percent = round(_pct(active, vahan_count), 1)
        compliance_risk_count = int(expired + expiring_soon)

        # Monthly growth percent from regn_dt (last month vs previous month)
        month_counts = await db.vahan_data.aggregate(
            [
                {"$match": match},
                {"$project": {"m": {"$substrBytes": [{"$toString": "$regn_dt"}, 0, 7]}}},
                {"$match": {"m": {"$regex": r"^\\d{4}-\\d{2}$"}}},
                {"$group": {"_id": "$m", "count": {"$sum": 1}}},
                {"$sort": {"_id": 1}},
            ]
        ).to_list(1000)
        monthly_growth_percent = 0.0
        if len(month_counts) >= 2:
            curr = month_counts[-1]["count"]
            prev = month_counts[-2]["count"]
            monthly_growth_percent = round(((curr - prev) / prev * 100) if prev else 0.0, 1)

        # Data quality score: completeness across key fields
        quality_fields = ["state_cd", "off_cd", "regn_no", "regn_dt", "fuel", "vch_catg", "vh_class", "norms", "body_type"]
        quality_result = await db.vahan_data.aggregate(
            [
                {"$match": match},
                {
                    "$group": {
                        "_id": None,
                        "total": {"$sum": 1},
                        **{
                            f"nn_{f}": {
                                "$sum": {
                                    "$cond": [
                                        {
                                            "$and": [
                                                {"$ne": [f"${f}", None]},
                                                {"$ne": [{"$toString": f"${f}"}, "nan"]},
                                                {"$ne": [{"$toString": f"${f}"}, ""]},
                                            ]
                                        },
                                        1,
                                        0,
                                    ]
                                }
                            }
                            for f in quality_fields
                        },
                    }
                }
            ]
        ).to_list(1)
        dq = 0.0
        if quality_result:
            total = quality_result[0].get("total", 0) or 0
            if total:
                per_field = []
                for f in quality_fields:
                    per_field.append(quality_result[0].get(f"nn_{f}", 0) / total)
                dq = round(sum(per_field) / len(per_field) * 100, 1) if per_field else 0.0

        # Top state share
        top_states = await db.vahan_data.aggregate(
            [
                {"$match": match},
                {"$group": {"_id": "$state_cd", "count": {"$sum": 1}}},
                {"$match": {"_id": {"$ne": None}}},
                {"$sort": {"count": -1}},
                {"$limit": 3},
            ]
        ).to_list(3)
        top_state = top_states[0]["_id"] if top_states else None
        top_state_share = _pct(top_states[0]["count"], vahan_count) if (top_states and vahan_count) else 0.0

        # ================= Tickets KPIs (data-driven) =================
        ticket_count = await db.tickets_data.count_documents({})
        status_counts = await db.tickets_data.aggregate([{"$group": {"_id": "$Status", "count": {"$sum": 1}}}]).to_list(50)
        by_status = {r["_id"]: r["count"] for r in status_counts if r.get("_id")}
        closed_tickets = by_status.get("Closed", 0)
        open_tickets = max(0, ticket_count - closed_tickets)
        ticket_closure_rate = round(_pct(closed_tickets, ticket_count), 1)

        # Avg resolution time (Created -> Closed) for closed tickets
        docs_closed = await db.tickets_data.find({"Status": "Closed"}, {"_id": 0, "Created": 1, "Closed": 1}).to_list(5000)
        res_days = []
        for d in docs_closed:
            c = _safe_parse_date(d.get("Created"))
            cl = _safe_parse_date(d.get("Closed"))
            if c and cl:
                res_days.append(max(0, (cl - c).days))
        avg_resolution_time = round((sum(res_days) / len(res_days)) if res_days else 0.0, 1)

        # Stale tickets: open tickets older than 30 days
        docs_open = await db.tickets_data.find({"Status": {"$ne": "Closed"}}, {"_id": 0, "Created": 1}).to_list(20000)
        stale = 0
        for d in docs_open:
            c = _safe_parse_date(d.get("Created"))
            if c and (now - c).days >= 30:
                stale += 1
        stale_ticket_percent = round(_pct(stale, open_tickets), 1) if open_tickets else 0.0

        # Hotspots: modules/categories for open tickets
        module_counts = Counter()
        cat_counts = Counter()
        open_meta = await db.tickets_data.find(
            {"Status": {"$ne": "Closed"}},
            {"_id": 0, "ModuleName": 1, "Issue Category": 1},
        ).to_list(5000)
        for d in open_meta:
            mn = d.get("ModuleName")
            ic = d.get("Issue Category")
            if mn:
                module_counts[str(mn)] += 1
            if ic:
                cat_counts[str(ic)] += 1

        # Sentiment distribution (now deterministic)
        sent_counts = await db.tickets_data.aggregate([{"$group": {"_id": "$sentiment", "count": {"$sum": 1}}}]).to_list(10)
        sentiment_distribution = {r["_id"]: r["count"] for r in sent_counts if r.get("_id")}
        neg_share = _pct(sentiment_distribution.get("negative", 0), ticket_count) if ticket_count else 0.0

        # ================= AI Insights (data-driven heuristics) =================
        insights = []

        # Theme detection from ticket text (Subject/Category/ModuleName) + 30-day trend
        theme_defs = [
            ("DL services", ["driving license", "dl", "license", "renewal", "renew"]),
            ("RC services", ["registration certificate", " rc ", "rc", "rc transfer", "rc status"]),
            ("Challan / fines", ["challan", "fine", "penalty", "e-challan", "echallan"]),
            ("Payments / finance", ["payment", "finance", "fee", "refund", "receipt"]),
            ("Vehicle registration", ["registration", "regn", "vahan"]),
        ]

        def _ticket_theme(text: str) -> str:
            if not text:
                return "Other"
            t = f" {str(text).lower()} "
            for name, kws in theme_defs:
                for kw in kws:
                    if kw in t:
                        return name
            return "Other"

        ticket_docs = await db.tickets_data.find(
            {},
            {"_id": 0, "Subject": 1, "Issue Category": 1, "ModuleName": 1, "Status": 1, "Created": 1},
        ).to_list(20000)

        theme_total = Counter()
        theme_open = Counter()
        theme_last30 = Counter()
        theme_prev30 = Counter()
        last30_start = now - timedelta(days=30)
        prev30_start = now - timedelta(days=60)

        for d in ticket_docs:
            parts = []
            for k in ("Subject", "Issue Category", "ModuleName"):
                if d.get(k):
                    parts.append(str(d.get(k)))
            theme = _ticket_theme(" ".join(parts))
            theme_total[theme] += 1
            if d.get("Status") != "Closed":
                theme_open[theme] += 1
            created = _safe_parse_date(d.get("Created"))
            if created:
                if created >= last30_start:
                    theme_last30[theme] += 1
                elif prev30_start <= created < last30_start:
                    theme_prev30[theme] += 1

        # If we have at least 1 theme with activity, add a trend insight
        if theme_total:
            # pick the most common theme overall (excluding Other if possible)
            top_theme = None
            for name, _ in theme_total.most_common():
                if name != "Other":
                    top_theme = name
                    break
            top_theme = top_theme or theme_total.most_common(1)[0][0]

            curr = theme_last30.get(top_theme, 0)
            prev = theme_prev30.get(top_theme, 0)
            if prev > 0:
                pct = round(((curr - prev) / prev) * 100, 1)
                direction = "increased" if pct >= 0 else "decreased"
                insights.append(
                    {
                        "type": "warning" if pct >= 10 else "info",
                        "message": f"{top_theme} grievances {direction} by {abs(pct)}% in the last 30 days ({curr} vs {prev}).",
                    }
                )
            else:
                insights.append(
                    {
                        "type": "info",
                        "message": f"Top grievance theme: {top_theme} ({theme_total[top_theme]} tickets; {theme_open.get(top_theme, 0)} currently open).",
                    }
                )

        if compliance_risk_count:
            insights.append(
                {
                    "type": "warning",
                    "message": f"{compliance_risk_count:,} registrations are at compliance risk ({expired:,} expired, {expiring_soon:,} expiring in 30 days).",
                }
            )

        if top_state:
            insights.append(
                {
                    "type": "info",
                    "message": f"Top registration state is {top_state} with {top_state_share}% of total registrations.",
                }
            )

        if open_tickets:
            top_modules = _top_n_list(module_counts, 2)
            top_cats = _top_n_list(cat_counts, 2)
            focus = []
            if top_modules:
                focus.append(f"modules: {', '.join(top_modules)}")
            if top_cats:
                focus.append(f"categories: {', '.join(top_cats)}")
            focus_txt = f" Focus areas: {', '.join(focus)}." if focus else ""
            insights.append(
                {
                    "type": "recommendation",
                    "message": f"Backlog detected: {open_tickets:,} open tickets; {stale:,} are >30 days old.{focus_txt}",
                }
            )

        # Sentiment insight (derived from deterministic heuristic during ticket load)
        if neg_share >= 30.0:
            insights.append(
                {
                    "type": "info",
                    "message": f"Ticket sentiment signal: {neg_share}% classified as negative based on subject/category keywords.",
                }
            )

        if month_counts:
            peak = max(month_counts, key=lambda r: r["count"])
            insights.append(
                {
                    "type": "info",
                    "message": f"Peak registration month observed: {peak['_id']} with {peak['count']:,} registrations.",
                }
            )

        # Keep only top 3 insights (prioritize warning -> recommendation -> info)
        priority = {"warning": 0, "recommendation": 1, "info": 2}
        insights = sorted(insights, key=lambda x: priority.get(x.get("type", "info"), 9))[:3]
        
        # Ensure median_vehicle_value is properly calculated and not rounded incorrectly
        # Round to 2 decimal places for precision, not 0
        final_median_value = round(median_vehicle_value, 2) if median_vehicle_value > 0 else 0.0
        
        return {
            "total_registrations": vahan_count,
            "monthly_growth_percent": monthly_growth_percent,
            "median_vehicle_value": final_median_value,
            "avg_registration_delay": avg_registration_delay,
            "active_registrations_percent": active_registrations_percent,
            "compliance_risk_count": compliance_risk_count,
            "total_tickets": ticket_count,
            "ticket_closure_rate": ticket_closure_rate,
            "avg_resolution_time": avg_resolution_time,
            "stale_ticket_percent": stale_ticket_percent,
            "data_quality_score": dq,
            "ai_insights": insights,
        }
    except Exception as e:
        logger.error(f"Error building executive summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@dashboard_router.get("/heatmap-data")
async def get_heatmap_data(month: Optional[str] = None):
    """Get state-wise data for heat map visualization"""
    try:
        # Get latest month if not specified
        if not month:
            latest_doc = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
            month = latest_doc.get("Month") if latest_doc else None
        
        if not month:
            return {"error": "No data available"}
        
        query = {"Month": month}
        
        # Fetch state-wise data
        cursor = db["kpi_state_general"].find(query)
        state_data_list = await cursor.to_list(length=100)
        
        # Aggregate data by state
        state_map = {}
        for record in state_data_list:
            state = record.get("State", "").strip()
            if not state:
                continue
            
            if state not in state_map:
                state_map[state] = {
                    "state": state,
                    "vehicle_registration": 0,
                    "accidents": 0,
                    "revenue": 0,
                    "challans": 0
                }
            
            # Use _get_field_value helper for robust field extraction
            try:
                vehicle_reg = _get_field_value(record, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0
                accidents = _get_field_value(record, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents") or 0
                revenue = _get_field_value(record, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total") or 0
                challans = _get_field_value(record, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued") or 0
                
                # Ensure values are numeric
                vehicle_reg = float(vehicle_reg) if vehicle_reg else 0
                accidents = float(accidents) if accidents else 0
                revenue = float(revenue) if revenue else 0
                challans = float(challans) if challans else 0
                
                state_map[state]["vehicle_registration"] += vehicle_reg
                state_map[state]["accidents"] += accidents
                state_map[state]["revenue"] += revenue
                state_map[state]["challans"] += challans
            except Exception as field_error:
                logger.warning(f"Error extracting field values for state {state}: {field_error}")
                continue
        
        # Convert to list
        result = list(state_map.values())
        
        return {
            "month": month,
            "data": result
        }
    except Exception as e:
        logger.error(f"Error fetching heatmap data: {e}", exc_info=True)
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== MAIN ROUTES =====================
@api_router.get("/")
async def root():
    return {"message": "Citizen Assistance Platform API", "version": "1.0.0"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    return status_checks

# Include all routers
# ===================== KPI DASHBOARD ENDPOINTS =====================

@kpi_router.get("/state/general")
async def get_state_general_kpi(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get State-level General KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month

        cursor = db["kpi_state_general"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        # Convert ObjectId to string for JSON serialization
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching state general KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/state/service-delivery")
async def get_state_service_kpi(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get State-level Service Delivery KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month

        cursor = db["kpi_state_service"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching state service KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/state/policy")
async def get_state_policy_kpi(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get State-level Policy Implementation KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month

        cursor = db["kpi_state_policy"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching state policy KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/rto/general")
async def get_rto_general_kpi(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get RTO-level General KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month

        cursor = db["kpi_rto_general"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching RTO general KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/rto/performance")
async def get_rto_performance_kpi(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get RTO-level Performance KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month

        cursor = db["kpi_rto_performance"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching RTO performance KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/rto/policy")
async def get_rto_policy_kpi(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get RTO-level Policy Implementation KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month

        cursor = db["kpi_rto_policy"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching RTO policy KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/rto/desk")
async def get_rto_desk_kpi(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get RTO-level Desk Performance KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month

        cursor = db["kpi_rto_desk"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching RTO desk KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/rto/internal")
async def get_rto_internal_kpi(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get RTO-level Internal Efficiency KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month

        cursor = db["kpi_rto_internal"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching RTO internal KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/fleet/vehicles")
async def get_fleet_vehicles_kpi(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Fleet-level Vehicles KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month

        cursor = db["kpi_fleet_vehicles"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching fleet vehicles KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/fleet/drivers")
async def get_fleet_drivers_kpi(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Fleet-level Drivers KPI data"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month

        cursor = db["kpi_fleet_drivers"].find(query).sort("Month", -1)
        records = await cursor.to_list(length=1000)
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        return {"data": records, "count": len(records)}
    except Exception as e:
        logger.error(f"Error fetching fleet drivers KPI: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/summary")
async def get_kpi_summary(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get aggregated KPI summary across all levels"""
    try:
        # Get latest month data if month not specified
        if not month:
            latest_state = await db["kpi_state_general"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest_state.get("Month") if latest_state else None

        query = {"Month": month} if month else {}
        if state:
            query["State"] = state

        summary = {
            "month": month,
            "state": state,
            "state_general": [],
            "state_service": [],
            "state_policy": [],
            "rto_general": [],
            "rto_performance": [],
            "rto_policy": [],
            "rto_desk": [],
            "rto_internal": [],
            "fleet_vehicles": [],
            "fleet_drivers": [],
        }

        # Fetch all KPI data
        collections = {
            "state_general": db["kpi_state_general"],
            "state_service": db["kpi_state_service"],
            "state_policy": db["kpi_state_policy"],
            "rto_general": db["kpi_rto_general"],
            "rto_performance": db["kpi_rto_performance"],
            "rto_policy": db["kpi_rto_policy"],
            "rto_desk": db["kpi_rto_desk"],
            "rto_internal": db["kpi_rto_internal"],
            "fleet_vehicles": db["kpi_fleet_vehicles"],
            "fleet_drivers": db["kpi_fleet_drivers"],
        }

        for key, collection in collections.items():
            cursor = collection.find(query).sort("Month", -1).limit(100)
            records = await cursor.to_list(length=100)
            # Convert ObjectId to string for JSON serialization
            for record in records:
                if "_id" in record:
                    record["_id"] = str(record["_id"])
            summary[key] = records

        return summary
    except Exception as e:
        logger.error(f"Error fetching KPI summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== ENHANCED KPI ENDPOINTS =====================

@kpi_router.get("/national/summary")
async def get_national_kpis(month: Optional[str] = None):
    """Get National-level (MoRTH) aggregated KPIs"""
    try:
        if not month:
            latest = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
            month = latest.get("Month") if latest else None

        query = {"Month": month} if month else {}
        
        # Fetch all state general records and aggregate in Python to handle field name variations
        state_gen_records = await db["kpi_state_general"].find(query).to_list(length=1000)
        
        # Aggregate using helper functions that handle field name variations
        total_vehicle_registration = _aggregate_kpi_field(
            state_gen_records, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration"
        )
        total_ll_issued = _aggregate_kpi_field(
            state_gen_records, "LL Issued", "LL Issued ", "LLIssued", "LL_Issued"
        )
        total_dl_issued = _aggregate_kpi_field(
            state_gen_records, "DL Issued", "DL Issued ", "DLIssued", "DL_Issued"
        )
        total_e_challan = _aggregate_kpi_field(
            state_gen_records, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued"
        )
        total_revenue = _aggregate_kpi_field(
            state_gen_records, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total"
        )
        total_accidents = _aggregate_kpi_field(
            state_gen_records, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents"
        )
        total_fatalities = _aggregate_kpi_field(
            state_gen_records, "Road Fatalities", "Road Fatalities ", "RoadFatalities", "Road_Fatalities"
        )
        total_transactions = _aggregate_kpi_field(
            state_gen_records, "Total Transactions (All)", "Total Transactions (All) ", "TotalTransactionsAll", "Total_Transactions_All"
        )
        state_count = len(state_gen_records)
        
        # Get service delivery metrics
        state_svc_records = await db["kpi_state_service"].find(query).to_list(length=1000)
        
        total_online_services = _aggregate_kpi_field(
            state_svc_records, "Online Service Count", "Online Service Count ", "OnlineServiceCount", "Online_Service_Count"
        )
        total_faceless_services = _aggregate_kpi_field(
            state_svc_records, "Faceless Service Count", "Faceless Service Count ", "FacelessServiceCount", "Faceless_Service_Count"
        )
        
        # Calculate averages for SLA metrics
        citizen_sla_values = []
        grievance_sla_values = []
        for record in state_svc_records:
            citizen_sla = _get_field_value(
                record, "Citizen Service SLA % (within SLA)", "Citizen Service SLA % (within SLA) ", 
                "CitizenServiceSLA", "Citizen_Service_SLA"
            )
            grievance_sla = _get_field_value(
                record, "Grievance SLA % (within SLA)", "Grievance SLA % (within SLA) ", 
                "GrievanceSLA", "Grievance_SLA"
            )
            if citizen_sla is not None:
                citizen_sla_values.append(citizen_sla)
            if grievance_sla is not None:
                grievance_sla_values.append(grievance_sla)
        
        avg_citizen_sla = sum(citizen_sla_values) / len(citizen_sla_values) if citizen_sla_values else 0.0
        avg_grievance_sla = sum(grievance_sla_values) / len(grievance_sla_values) if grievance_sla_values else 0.0
        
        # Calculate derived KPIs
        total_licenses = total_ll_issued + total_dl_issued
        digital_adoption = 0.0
        if total_transactions > 0:
            online_transactions = total_online_services * 1000  # Estimate
            digital_adoption = (online_transactions / total_transactions) * 100
        
        return {
            "month": month,
            "core_metrics": {
                "vehicle_registration": total_vehicle_registration,
                "ll_issued": total_ll_issued,
                "dl_issued": total_dl_issued,
                "e_challan_issued": total_e_challan,
                "revenue_collected": total_revenue,
                "road_accidents": total_accidents,
                "road_fatalities": total_fatalities,
                "total_transactions": total_transactions
            },
            "derived_kpis": {
                "national_vehicle_registration_volume": total_vehicle_registration,
                "national_license_issuance_index": total_licenses,
                "national_transport_revenue": total_revenue,
                "compliance_enforcement_index": total_e_challan,
                "digital_adoption_ratio": round(digital_adoption, 2),
                "online_service_count": total_online_services,
                "faceless_service_count": total_faceless_services,
                "avg_citizen_sla": round(avg_citizen_sla, 2),
                "avg_grievance_sla": round(avg_grievance_sla, 2)
            },
            "state_count": state_count
        }
    except Exception as e:
        logger.error(f"Error fetching national KPIs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/state/derived")
async def get_state_derived_kpis(state: Optional[str] = None, month: Optional[str] = None):
    """Get State-level derived KPIs"""
    try:
        if not month:
            latest = await db["kpi_state_general"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {"Month": month} if month else {}
        if state:
            query["State"] = state

        # Get state general data
        state_gen = await db["kpi_state_general"].find_one(query)
        state_svc = await db["kpi_state_service"].find_one(query)
        state_pol = await db["kpi_state_policy"].find_one(query)
        
        if not state_gen:
            return {"error": "No data found"}
        
        # Extract values using helper functions to handle field name variations
        vehicle_reg = _safe_aggregate_field(
            state_gen, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration"
        )
        ll_issued = _safe_aggregate_field(
            state_gen, "LL Issued", "LL Issued ", "LLIssued", "LL_Issued"
        )
        dl_issued = _safe_aggregate_field(
            state_gen, "DL Issued", "DL Issued ", "DLIssued", "DL_Issued"
        )
        revenue_total = _safe_aggregate_field(
            state_gen, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total"
        )
        
        # Service delivery metrics
        online_service_count = 0
        faceless_service_count = 0
        citizen_sla = 0
        grievance_sla = 0
        
        if state_svc:
            online_service_count = _safe_aggregate_field(
                state_svc, "Online Service Count", "Online Service Count ", "OnlineServiceCount", "Online_Service_Count"
            )
            faceless_service_count = _safe_aggregate_field(
                state_svc, "Faceless Service Count", "Faceless Service Count ", "FacelessServiceCount", "Faceless_Service_Count"
            )
            citizen_sla = _safe_aggregate_field(
                state_svc, "Citizen Service SLA % (within SLA)", "Citizen Service SLA % (within SLA) ", 
                "CitizenServiceSLA", "Citizen_Service_SLA"
            )
            grievance_sla = _safe_aggregate_field(
                state_svc, "Grievance SLA % (within SLA)", "Grievance SLA % (within SLA) ", 
                "GrievanceSLA", "Grievance_SLA"
            )
        
        # Calculate derived KPIs
        total_services = online_service_count + faceless_service_count
        faceless_adoption = (faceless_service_count / total_services * 100) if total_services > 0 else 0.0
        
        # Enforcement infrastructure index
        enforcement_index = 0.0
        if state_pol:
            ats = _safe_aggregate_field(state_pol, "No. of ATS", "No. of ATS ", "NoOfATS", "No_of_ATS")
            adtt = _safe_aggregate_field(state_pol, "No. of ADTT", "No. of ADTT ", "NoOfADTT", "No_of_ADTT")
            rvsf = _safe_aggregate_field(state_pol, "No. of RVSF", "No. of RVSF ", "NoOfRVSF", "No_of_RVSF")
            enforcement_index = ats + adtt + rvsf
        
        # License issuance efficiency (licenses per day - assuming 30 days)
        total_licenses = ll_issued + dl_issued
        licenses_per_day = (total_licenses / 30.0) if total_licenses > 0 else 0.0
        
        # Service delivery ranking
        service_delivery_ranking = (citizen_sla * 0.6 + grievance_sla * 0.4) if (citizen_sla > 0 or grievance_sla > 0) else 0.0
        
        return {
            "month": month,
            "state": state_gen.get("State"),
            "derived_kpis": {
                "state_service_delivery_ranking": round(service_delivery_ranking, 2),
                "faceless_services_adoption_pct": round(faceless_adoption, 2),
                "state_revenue_contribution": revenue_total,
                "enforcement_infrastructure_index": enforcement_index,
                "vehicle_growth_rate": 0,  # Would need YoY comparison
                "license_issuance_efficiency": round(licenses_per_day, 2),
                "online_service_count": online_service_count,
                "faceless_service_count": faceless_service_count
            }
        }
    except Exception as e:
        logger.error(f"Error fetching state derived KPIs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/rto/derived")
async def get_rto_derived_kpis(state: Optional[str] = None, rto: Optional[str] = None, month: Optional[str] = None):
    """Get RTO-level derived KPIs"""
    try:
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month

        rto_gen = await db["kpi_rto_general"].find_one(query, sort=[("Month", -1)])
        rto_perf = await db["kpi_rto_performance"].find_one(query, sort=[("Month", -1)])
        
        if not rto_gen:
            return {"error": "No data found"}
        
        # Extract values using helper functions to handle field name variations
        revenue = _safe_aggregate_field(
            rto_gen, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total"
        )
        vehicle_reg = _safe_aggregate_field(
            rto_gen, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration"
        )
        e_challan = _safe_aggregate_field(
            rto_gen, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued"
        )
        
        revenue_per_rto = revenue
        enforcement_effectiveness = (e_challan / vehicle_reg * 100) if vehicle_reg > 0 else 0.0
        
        # RTO Performance Index (composite score)
        performance_index = 0.0
        faceless_pct = 0.0
        citizen_sla = 0.0
        grievance_sla = 0.0
        revenue_actual = 0.0
        revenue_target = 0.0
        
        if rto_perf:
            faceless_pct = _safe_aggregate_field(
                rto_perf, "Faceless Application %", "Faceless Application % ", "FacelessApplicationPct", "Faceless_Application_Pct"
            )
            citizen_sla = _safe_aggregate_field(
                rto_perf, "Citizen Service SLA % (within SLA)", "Citizen Service SLA % (within SLA) ", 
                "CitizenServiceSLA", "Citizen_Service_SLA"
            )
            grievance_sla = _safe_aggregate_field(
                rto_perf, "Grievance SLA % (within SLA)", "Grievance SLA % (within SLA) ", 
                "GrievanceSLA", "Grievance_SLA"
            )
            revenue_actual = _safe_aggregate_field(
                rto_perf, "Revenue - Actual", "Revenue - Actual ", "RevenueActual", "Revenue_Actual"
            )
            revenue_target = _safe_aggregate_field(
                rto_perf, "Revenue - Target", "Revenue - Target ", "RevenueTarget", "Revenue_Target"
            )
            
            revenue_target_ratio = 0.0
            if revenue_target > 0:
                revenue_target_ratio = (revenue_actual / revenue_target) * 100
            
            performance_index = (
                faceless_pct * 0.25 +
                citizen_sla * 0.30 +
                grievance_sla * 0.25 +
                min(revenue_target_ratio, 100) * 0.20
            )
        
        return {
            "month": rto_gen.get("Month"),
            "state": rto_gen.get("State"),
            "rto": rto_gen.get("RTO"),
            "derived_kpis": {
                "rto_performance_index": round(performance_index, 2),
                "revenue_per_rto": revenue_per_rto,
                "enforcement_effectiveness": round(enforcement_effectiveness, 2),
                "fitness_compliance_ratio": 0,  # Would need fitness test data
                "faceless_application_pct": round(faceless_pct, 2),
                "citizen_service_sla": round(citizen_sla, 2),
                "grievance_sla": round(grievance_sla, 2),
                "revenue_actual": revenue_actual,
                "revenue_target": revenue_target
            }
        }
    except Exception as e:
        logger.error(f"Error fetching RTO derived KPIs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/executive/summary")
async def get_executive_summary_kpis(month: Optional[str] = None):
    """Get Executive Summary KPIs for CM/CS/MoRTH Review"""
    try:
        if not month:
            latest = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
            month = latest.get("Month") if latest else None

        query = {"Month": month} if month else {}
        
        # Fetch all records and aggregate in Python to handle field name variations
        state_gen_records = await db["kpi_state_general"].find(query).to_list(length=1000)
        state_svc_records = await db["kpi_state_service"].find(query).to_list(length=1000)
        state_pol_records = await db["kpi_state_policy"].find(query).to_list(length=1000)
        
        # Aggregate national data using helper functions
        total_vehicles = _aggregate_kpi_field(
            state_gen_records, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration"
        )
        total_revenue = _aggregate_kpi_field(
            state_gen_records, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total"
        )
        total_accidents = _aggregate_kpi_field(
            state_gen_records, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents"
        )
        total_fatalities = _aggregate_kpi_field(
            state_gen_records, "Road Fatalities", "Road Fatalities ", "RoadFatalities", "Road_Fatalities"
        )
        total_challans = _aggregate_kpi_field(
            state_gen_records, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued"
        )
        
        # Service delivery metrics
        total_online = _aggregate_kpi_field(
            state_svc_records, "Online Service Count", "Online Service Count ", "OnlineServiceCount", "Online_Service_Count"
        )
        total_faceless = _aggregate_kpi_field(
            state_svc_records, "Faceless Service Count", "Faceless Service Count ", "FacelessServiceCount", "Faceless_Service_Count"
        )
        
        # Calculate averages for SLA metrics
        citizen_sla_values = []
        grievance_sla_values = []
        for record in state_svc_records:
            citizen_sla = _get_field_value(
                record, "Citizen Service SLA % (within SLA)", "Citizen Service SLA % (within SLA) ", 
                "CitizenServiceSLA", "Citizen_Service_SLA"
            )
            grievance_sla = _get_field_value(
                record, "Grievance SLA % (within SLA)", "Grievance SLA % (within SLA) ", 
                "GrievanceSLA", "Grievance_SLA"
            )
            if citizen_sla is not None:
                citizen_sla_values.append(citizen_sla)
            if grievance_sla is not None:
                grievance_sla_values.append(grievance_sla)
        
        avg_citizen_sla = sum(citizen_sla_values) / len(citizen_sla_values) if citizen_sla_values else 0.0
        avg_grievance_sla = sum(grievance_sla_values) / len(grievance_sla_values) if grievance_sla_values else 0.0
        
        # Policy implementation metrics
        total_ats = _aggregate_kpi_field(
            state_pol_records, "No. of ATS", "No. of ATS ", "NoOfATS", "No_of_ATS"
        )
        total_adtt = _aggregate_kpi_field(
            state_pol_records, "No. of ADTT", "No. of ADTT ", "NoOfADTT", "No_of_ADTT"
        )
        total_rvsf = _aggregate_kpi_field(
            state_pol_records, "No. of RVSF", "No. of RVSF ", "NoOfRVSF", "No_of_RVSF"
        )
        total_vltd = _aggregate_kpi_field(
            state_pol_records, "Count of Vehicles fitted with VLTD", "Count of Vehicles fitted with VLTD ", 
            "CountOfVehiclesFittedWithVLTD", "Count_of_Vehicles_fitted_with_VLTD"
        )
        total_hsrp = _aggregate_kpi_field(
            state_pol_records, "Count of Vehicles fitted with HSRP", "Count of Vehicles fitted with HSRP ", 
            "CountOfVehiclesFittedWithHSRP", "Count_of_Vehicles_fitted_with_HSRP"
        )
        
        # Calculate executive KPIs
        # National Mobility Growth Index (simplified - would need YoY)
        mobility_growth_index = 100  # Placeholder
        
        # Digital Governance Index
        total_services = total_online + total_faceless
        digital_governance_index = round(
            avg_citizen_sla * 0.5 +
            avg_grievance_sla * 0.3 +
            min((total_services / 1000) * 10, 20), 2
        )
        
        # Road Safety Risk Index
        safety_risk_index = 0.0
        if total_vehicles > 0:
            accident_rate = (total_accidents / total_vehicles) * 1000
            violation_rate = (total_challans / total_vehicles) * 1000
            safety_risk_index = round((accident_rate * 0.6 + violation_rate * 0.4), 2)
        
        # Enforcement ROI (Revenue vs Devices)
        total_devices = total_ats + total_adtt + total_rvsf
        enforcement_roi = round(total_revenue / total_devices, 2) if total_devices > 0 else 0.0
        
        # Citizen Service Efficiency Score
        citizen_efficiency = round(
            avg_citizen_sla * 0.7 +
            avg_grievance_sla * 0.3, 2
        )
        
        return {
            "month": month,
            "executive_kpis": {
                "national_mobility_growth_index": mobility_growth_index,
                "digital_governance_index": digital_governance_index,
                "road_safety_risk_index": safety_risk_index,
                "enforcement_roi_kpi": enforcement_roi,
                "citizen_service_efficiency_score": citizen_efficiency
            },
            "supporting_metrics": {
                "total_vehicles": total_vehicles,
                "total_revenue": total_revenue,
                "total_accidents": total_accidents,
                "total_fatalities": total_fatalities,
                "total_challans": total_challans,
                "total_online_services": total_online,
                "total_faceless_services": total_faceless,
                "total_enforcement_devices": total_devices,
                "vehicles_with_vltd": total_vltd,
                "vehicles_with_hsrp": total_hsrp
            }
        }
    except Exception as e:
        logger.error(f"Error fetching executive summary KPIs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ===================== DRILL-DOWN ENDPOINTS =====================

@kpi_router.get("/drilldown/national/vehicle-registration")
async def get_national_vehicle_registration_drilldown(month: Optional[str] = None):
    """Drill-down for National Vehicle Registration - State breakdown"""
    try:
        if not month:
            latest = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
            month = latest.get("Month") if latest else None

        query = {"Month": month} if month else {}
        cursor = db["kpi_state_general"].find(query).sort("Vehicle Registration", -1)
        records = await cursor.to_list(length=100)
        
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        # Get trend data (last 12 months)
        trend_query = {}
        trend_cursor = db["kpi_state_general"].find(trend_query).sort("Month", -1).limit(12)
        trend_records = await trend_cursor.to_list(length=12)
        
        # Aggregate by month
        monthly_totals = {}
        for record in trend_records:
            month_key = record.get("Month")
            if month_key:
                if month_key not in monthly_totals:
                    monthly_totals[month_key] = 0
                monthly_totals[month_key] += record.get("Vehicle Registration", 0) or 0
        
        trend_data = [
            {"month": m, "value": v}
            for m, v in sorted(monthly_totals.items(), reverse=True)[:12]
        ]

        total = sum(r.get("Vehicle Registration", 0) or 0 for r in records)
        
        return {
            "state_breakdown": records,
            "trend_data": trend_data,
            "summary": {
                "total_registrations": total,
                "state_count": len(records),
                "avg_per_state": round(total / len(records), 0) if records else 0
            }
        }
    except Exception as e:
        logger.error(f"Error fetching national vehicle registration drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/national/revenue")
async def get_national_revenue_drilldown(month: Optional[str] = None):
    """Drill-down for National Revenue - State breakdown"""
    try:
        if not month:
            latest = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
            month = latest.get("Month") if latest else None

        query = {"Month": month} if month else {}
        cursor = db["kpi_state_general"].find(query).sort("Revenue - Total", -1)
        records = await cursor.to_list(length=100)
        
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        # Get trend data
        trend_query = {}
        trend_cursor = db["kpi_state_general"].find(trend_query).sort("Month", -1).limit(12)
        trend_records = await trend_cursor.to_list(length=12)
        
        monthly_totals = {}
        for record in trend_records:
            month_key = record.get("Month")
            if month_key:
                if month_key not in monthly_totals:
                    monthly_totals[month_key] = 0
                monthly_totals[month_key] += record.get("Revenue - Total", 0) or 0
        
        trend_data = [
            {"month": m, "value": v}
            for m, v in sorted(monthly_totals.items(), reverse=True)[:12]
        ]

        total = sum(r.get("Revenue - Total", 0) or 0 for r in records)
        
        return {
            "state_breakdown": records,
            "trend_data": trend_data,
            "summary": {
                "total_revenue": total,
                "state_count": len(records),
                "avg_per_state": round(total / len(records), 0) if records else 0
            }
        }
    except Exception as e:
        logger.error(f"Error fetching national revenue drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/state/breakdown")
async def get_state_breakdown_drilldown(
    state: Optional[str] = None,
    month: Optional[str] = None,
    metric: Optional[str] = None
):
    """Drill-down for State level - RTO breakdown"""
    try:
        # Get latest month if not specified
        if not month:
            latest = await db["kpi_rto_general"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {}
        if state:
            # Ensure state filter is always applied when state is provided
            state_clean = state.strip() if isinstance(state, str) else str(state)
            if state_clean:  # Only add if not empty after strip
                query["State"] = state_clean
        if month:
            query["Month"] = month

        # Get RTO-level data for the state - ensure only RTOs from the specified state are returned
        cursor = db["kpi_rto_general"].find(query).sort("Vehicle Registration", -1).limit(100)
        records = await cursor.to_list(length=100)
        
        # Additional validation: filter out any records that don't match the state (case-insensitive)
        if state:
            state_normalized = state.strip().lower() if isinstance(state, str) else str(state).lower()
            records = [r for r in records if r.get("State", "").strip().lower() == state_normalized]
        
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        # Get trend data for the state
        state_query = {"State": state} if state else {}
        trend_cursor = db["kpi_state_general"].find(state_query).sort("Month", -1).limit(12)
        trend_records = await trend_cursor.to_list(length=12)
        
        # Determine field name based on metric
        field_name = "Vehicle Registration" if metric == "vehicle_registration" else "Revenue - Total"
        trend_data = [
            {
                "month": r.get("Month"),
                "value": r.get(field_name, 0) or 0
            }
            for r in trend_records
        ]
        trend_data.reverse()

        return {
            "rto_breakdown": records,
            "trend_data": trend_data,
            "summary": {
                "rto_count": len(records),
                "total_vehicles": sum(r.get("Vehicle Registration", 0) or 0 for r in records),
                "total_revenue": sum(r.get("Revenue - Total", 0) or 0 for r in records)
            }
        }
    except Exception as e:
        logger.error(f"Error fetching state breakdown drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/rto/breakdown")
async def get_rto_breakdown_drilldown(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Drill-down for RTO level - Detailed metrics"""
    try:
        # Get latest month if not specified
        if not month:
            latest = await db["kpi_rto_general"].find_one(
                {"State": state, "RTO": rto} if (state and rto) else ({"State": state} if state else {}),
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {}
        if state:
            # Ensure state filter is always applied when state is provided
            query["State"] = state.strip() if isinstance(state, str) else state
        if rto:
            query["RTO"] = rto.strip() if isinstance(rto, str) else rto
        if month:
            query["Month"] = month

        # Get RTO general data - ensure state filter is applied
        rto_gen = await db["kpi_rto_general"].find_one(query, sort=[("Month", -1)])
        rto_perf = await db["kpi_rto_performance"].find_one(query, sort=[("Month", -1)])
        rto_desk = await db["kpi_rto_desk"].find_one(query, sort=[("Month", -1)])
        rto_int = await db["kpi_rto_internal"].find_one(query, sort=[("Month", -1)])

        if rto_gen and "_id" in rto_gen:
            rto_gen["_id"] = str(rto_gen["_id"])

        # Get trend data - filter by both state and RTO if provided
        trend_query = {}
        if state:
            trend_query["State"] = state
        if rto:
            trend_query["RTO"] = rto
        trend_cursor = db["kpi_rto_general"].find(trend_query).sort("Month", -1).limit(12)
        trend_records = await trend_cursor.to_list(length=12)
        
        trend_data = [
            {
                "month": r.get("Month"),
                "vehicle_registration": r.get("Vehicle Registration", 0) or 0,
                "revenue": r.get("Revenue - Total", 0) or 0
            }
            for r in trend_records
        ].reverse()

        return {
            "rto_data": rto_gen,
            "performance_data": rto_perf,
            "desk_data": rto_desk,
            "internal_data": rto_int,
            "trend_data": trend_data
        }
    except Exception as e:
        logger.error(f"Error fetching RTO breakdown drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/service-delivery")
async def get_service_delivery_drilldown(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Drill-down for Service Delivery metrics"""
    try:
        # Get latest month if not specified
        if not month:
            latest = await db["kpi_state_service"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month

        # Get state service data - get last 12 months
        if month:
            # Get last 12 months of data
            all_months_cursor = db["kpi_state_service"].find(
                {"State": state} if state else {}
            ).sort("Month", -1).limit(12)
        else:
            all_months_cursor = db["kpi_state_service"].find(query).sort("Month", -1).limit(12)
        
        records = await all_months_cursor.to_list(length=12)
        
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        # Get RTO-level service data
        rto_query = {}
        if state:
            rto_query["State"] = state
        if month:
            rto_query["Month"] = month
        rto_cursor = db["kpi_rto_performance"].find(rto_query).sort("Citizen Service SLA % (within SLA)", -1).limit(50)
        rto_records = await rto_cursor.to_list(length=50)
        
        for record in rto_records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        trend_data = [
            {
                "month": r.get("Month"),
                "citizen_sla": r.get("Citizen Service SLA % (within SLA)", 0) or 0,
                "grievance_sla": r.get("Grievance SLA % (within SLA)", 0) or 0,
                "online_services": r.get("Online Service Count", 0) or 0,
                "faceless_services": r.get("Faceless Service Count", 0) or 0
            }
            for r in records
        ]
        trend_data.reverse()

        return {
            "state_data": records or [],
            "rto_breakdown": rto_records or [],
            "trend_data": trend_data or []
        }
    except Exception as e:
        logger.error(f"Error fetching service delivery drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/revenue-trend")
async def get_revenue_trend_drilldown(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Drill-down for Revenue Trend analysis"""
    try:
        # Get latest month if not specified
        if not month:
            latest = await db["kpi_state_general"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {}
        if state:
            query["State"] = state

        # Get state revenue data - get last 12 months
        cursor = db["kpi_state_general"].find(query).sort("Month", -1).limit(12)
        records = await cursor.to_list(length=12)
        
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        # Get RTO revenue breakdown
        rto_query = {}
        if state:
            rto_query["State"] = state
        if month:
            rto_query["Month"] = month
        rto_cursor = db["kpi_rto_general"].find(rto_query).sort("Revenue - Total", -1).limit(50)
        rto_records = await rto_cursor.to_list(length=50)
        
        for record in rto_records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        trend_data = [
            {
                "month": r.get("Month"),
                "revenue_total": r.get("Revenue - Total", 0) or 0,
                "revenue_taxes": r.get("Revenue - Taxes", 0) or 0,
                "revenue_fees": r.get("Revenue - Fees", 0) or 0,
                "revenue_penalties": r.get("Revenue - Penalties", 0) or 0
            }
            for r in records
        ]
        trend_data.reverse()

        return {
            "state_data": records or [],
            "rto_breakdown": rto_records or [],
            "trend_data": trend_data or [],
            "summary": {
                "total_revenue": sum(r.get("Revenue - Total", 0) or 0 for r in records) if records else 0,
                "avg_monthly_revenue": sum(r.get("Revenue - Total", 0) or 0 for r in records) / len(records) if records and len(records) > 0 else 0
            }
        }
    except Exception as e:
        logger.error(f"Error fetching revenue trend drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/enforcement")
async def get_enforcement_drilldown(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Drill-down for Enforcement metrics"""
    try:
        # Get latest month if not specified
        if not month:
            latest = await db["kpi_state_general"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {}
        if state:
            query["State"] = state

        # Get state general data (for e-Challan) - get last 12 months
        state_cursor = db["kpi_state_general"].find(query).sort("Month", -1).limit(12)
        state_records = await state_cursor.to_list(length=12)
        
        # Get policy data (for enforcement devices) - get last 12 months
        policy_query = {}
        if state:
            policy_query["State"] = state
        policy_cursor = db["kpi_state_policy"].find(policy_query).sort("Month", -1).limit(12)
        policy_records = await policy_cursor.to_list(length=12)

        for record in state_records:
            if "_id" in record:
                record["_id"] = str(record["_id"])
        for record in policy_records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        trend_data = [
            {
                "month": state_records[i].get("Month") if i < len(state_records) else (policy_records[i].get("Month") if i < len(policy_records) else None),
                "e_challan": state_records[i].get("e-Challan Issued", 0) or 0 if i < len(state_records) else 0,
                "accidents": state_records[i].get("Road Accidents", 0) or 0 if i < len(state_records) else 0,
                "fatalities": state_records[i].get("Road Fatalities", 0) or 0 if i < len(state_records) else 0,
                "ats_count": policy_records[i].get("No. of ATS", 0) or 0 if i < len(policy_records) else 0,
                "adtt_count": policy_records[i].get("No. of ADTT", 0) or 0 if i < len(policy_records) else 0
            }
            for i in range(max(len(state_records), len(policy_records)))
        ]
        trend_data.reverse()

        return {
            "state_data": state_records or [],
            "policy_data": policy_records or [],
            "trend_data": trend_data or []
        }
    except Exception as e:
        logger.error(f"Error fetching enforcement drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/drilldown/fleet-vehicles")
async def get_fleet_vehicles_drilldown(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Drill-down for Fleet Vehicles metrics"""
    try:
        # Get latest month if not specified
        if not month:
            latest = await db["kpi_fleet_vehicles"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest.get("Month") if latest else None

        query = {}
        if state:
            query["State"] = state

        # Get fleet vehicles data - get last 12 months
        cursor = db["kpi_fleet_vehicles"].find(query).sort("Month", -1).limit(12)
        records = await cursor.to_list(length=12)
        
        for record in records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        # Get fleet drivers data
        drivers_query = {}
        if state:
            drivers_query["State"] = state
        if month:
            drivers_query["Month"] = month
        drivers_cursor = db["kpi_fleet_drivers"].find(drivers_query).sort("Month", -1).limit(12)
        drivers_records = await drivers_cursor.to_list(length=12)
        
        for record in drivers_records:
            if "_id" in record:
                record["_id"] = str(record["_id"])

        trend_data = [
            {
                "month": records[i].get("Month") if i < len(records) else (drivers_records[i].get("Month") if i < len(drivers_records) else None),
                "vehicles_owned": records[i].get("Vehicle Owned", 0) or 0 if i < len(records) else 0,
                "tax_due_count": records[i].get("Tax Due - Count", 0) or 0 if i < len(records) else 0,
                "insurance_due_count": records[i].get("Insurance Due - Count", 0) or 0 if i < len(records) else 0,
                "pucc_due_count": records[i].get("PUCC Due - Count", 0) or 0 if i < len(records) else 0,
                "fitness_due_count": records[i].get("Fitness Due - Count", 0) or 0 if i < len(records) else 0,
                "drivers_count": drivers_records[i].get("Driver Count", 0) or 0 if i < len(drivers_records) else 0
            }
            for i in range(max(len(records), len(drivers_records)))
        ]
        trend_data.reverse()

        return {
            "fleet_vehicles_data": records or [],
            "fleet_drivers_data": drivers_records or [],
            "trend_data": trend_data or [],
            "summary": {
                "total_vehicles": sum(r.get("Vehicle Owned", 0) or 0 for r in records) if records else 0,
                "total_tax_due": sum(r.get("Tax Due - Amount", 0) or 0 for r in records) if records else 0,
                "total_insurance_due": sum(r.get("Insurance Due - Count", 0) or 0 for r in records) if records else 0,
                "total_drivers": sum(r.get("Driver Count", 0) or 0 for r in drivers_records)
            }
        }
    except Exception as e:
        logger.error(f"Error fetching fleet vehicles drill-down: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ===================== ADVANCED/DERIVED KPI ENDPOINTS =====================

@kpi_router.get("/advanced/mobility-growth")
async def get_mobility_growth_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Mobility & Growth Intelligence KPIs"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        
        # Get current and previous month data
        current_cursor = db["kpi_state_general"].find(query).sort("Month", -1).limit(1)
        current_data = await current_cursor.to_list(1)
        if not current_data:
            # Try without filters to get latest data
            current_cursor = db["kpi_state_general"].find({}).sort("Month", -1).limit(1)
            current_data = await current_cursor.to_list(1)
            if not current_data:
                return {"error": "No data found"}
        
        current = current_data[0]
        current_month = current.get("Month")
        
        # If no state filter, aggregate across all states for current month
        if not state:
            current_agg_cursor = db["kpi_state_general"].find({"Month": current_month}).sort("Month", -1)
            current_agg_data = await current_agg_cursor.to_list(100)
            current_reg = sum(_get_field_value(r, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0 for r in current_agg_data)
            current_trans = sum(_get_field_value(r, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0 for r in current_agg_data)
        else:
            # Try multiple possible field names
            current_reg = _get_field_value(current, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0
            current_trans = _get_field_value(current, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0
        
        # Get previous month - need to find records with different month
        prev_query = {}
        if state:
            prev_query["State"] = state
        # Get distinct months to find previous one
        all_months_cursor = db["kpi_state_general"].find(prev_query, {"Month": 1}).sort("Month", -1).limit(12)
        all_months_data = await all_months_cursor.to_list(12)
        distinct_months = sorted(set(r.get("Month") for r in all_months_data if r.get("Month")), reverse=True)
        
        prev_reg = current_reg
        if len(distinct_months) > 1 and current_month in distinct_months:
            prev_month_idx = distinct_months.index(current_month) + 1
            if prev_month_idx < len(distinct_months):
                prev_month = distinct_months[prev_month_idx]
                # If no state filter, aggregate across all states for previous month
                if not state:
                    prev_agg_cursor = db["kpi_state_general"].find({"Month": prev_month})
                    prev_agg_data = await prev_agg_cursor.to_list(100)
                    prev_reg = sum(_get_field_value(r, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0 for r in prev_agg_data)
                else:
                    prev_query["Month"] = prev_month
                    prev_cursor = db["kpi_state_general"].find(prev_query).sort("Month", -1).limit(1)
                    prev_data = await prev_cursor.to_list(1)
                    if prev_data:
                        prev_reg = _get_field_value(prev_data[0], "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or current_reg
        
        # Calculate KPIs
        vehicle_demand_momentum = round(((current_reg - prev_reg) / prev_reg * 100) if prev_reg > 0 else 0, 2)
        reg_to_trans_efficiency = round((current_reg / current_trans * 100) if current_trans > 0 else 0, 2)
        
        # Get monthly trend for Z-score calculation
        trend_query = {}
        if state:
            trend_query["State"] = state
        trend_cursor = db["kpi_state_general"].find(trend_query).sort("Month", -1).limit(12)
        trend_data = await trend_cursor.to_list(12)
        
        # Aggregate by month if no state filter
        if not state:
            monthly_agg = defaultdict(float)
            for r in trend_data:
                month = r.get("Month")
                if month:
                    reg = _get_field_value(r, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0
                    monthly_agg[month] += reg
            monthly_regs = sorted([monthly_agg[m] for m in monthly_agg.keys()], reverse=True)[:12]
        else:
            monthly_regs = [_get_field_value(r, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0 for r in trend_data]
        
        # Calculate Z-score (simplified)
        if len(monthly_regs) > 1:
            mean_reg = sum(monthly_regs) / len(monthly_regs)
            std_reg = _stddev_pop(monthly_regs) if len(monthly_regs) > 1 else 1
            z_score = round((current_reg - mean_reg) / std_reg if std_reg > 0 else 0, 2)
        else:
            z_score = 0
        
        seasonal_spike_indicator = "High" if abs(z_score) > 2 else "Normal" if abs(z_score) > 1 else "Low"
        
        return {
            "month": current_month,
            "state": state,
            "kpis": {
                "vehicle_demand_momentum_index": vehicle_demand_momentum,
                "registration_to_transaction_efficiency": reg_to_trans_efficiency,
                "seasonal_mobility_spike_indicator": seasonal_spike_indicator,
                "z_score": z_score
            },
            "supporting_metrics": {
                "current_month_registrations": current_reg,
                "previous_month_registrations": prev_reg,
                "current_transactions": current_trans
            }
        }
    except Exception as e:
        logger.error(f"Error fetching mobility growth KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/digital-governance")
async def get_digital_governance_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Digital Governance & Service Maturity KPIs"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        
        # Get latest month first
        latest_month_doc = await db["kpi_state_service"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        # Update query with latest month if month not specified
        if not month:
            query["Month"] = current_month
        
        cursor = db["kpi_state_service"].find(query).sort("Month", -1).limit(12)
        records = await cursor.to_list(12)
        
        if not records:
            # Try with latest month
            query["Month"] = current_month
            cursor = db["kpi_state_service"].find(query).sort("Month", -1).limit(12)
            records = await cursor.to_list(12)
            if not records:
                return {"error": "No data found"}
        
        current = records[0]
        if not current_month:
            current_month = current.get("Month")
        
        # If no state filter, aggregate across all states for current month
        if not state:
            current_agg_cursor = db["kpi_state_service"].find({"Month": current_month})
            current_agg_data = await current_agg_cursor.to_list(100)
            online = sum(_get_field_value(r, "Online Service Count", "Online Service Count ", "OnlineServiceCount", "Online_Service_Count") or 0 for r in current_agg_data)
            faceless = sum(_get_field_value(r, "Faceless Service Count", "Faceless Service Count ", "FacelessServiceCount", "Faceless_Service_Count") or 0 for r in current_agg_data)
            # Try to get total transactions from general data if not in service data
            general_month = await db["kpi_state_general"].find_one({"Month": current_month}, sort=[("Month", -1)])
            total_trans = _get_field_value(general_month, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0 if general_month else 0
            if total_trans == 0:
                # Aggregate from general data
                general_agg = await db["kpi_state_general"].find({"Month": current_month}).to_list(100)
                total_trans = sum(_get_field_value(r, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0 for r in general_agg)
            # Average SLA values
            citizen_sla = sum(_get_field_value(r, "Citizen Service SLA % (within SLA)", "Citizen Service SLA %", "CitizenServiceSLA", "Citizen_Service_SLA") or 0 for r in current_agg_data) / len(current_agg_data) if current_agg_data else 0
            grievance_sla = sum(_get_field_value(r, "Grievance SLA % (within SLA)", "Grievance SLA %", "GrievanceSLA", "Grievance_SLA") or 0 for r in current_agg_data) / len(current_agg_data) if current_agg_data else 0
        else:
            # Try multiple possible field name variations
            online = _get_field_value(current, "Online Service Count", "Online Service Count ", "OnlineServiceCount", "Online_Service_Count") or 0
            faceless = _get_field_value(current, "Faceless Service Count", "Faceless Service Count ", "FacelessServiceCount", "Faceless_Service_Count") or 0
            total_trans = _get_field_value(current, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0
            if total_trans == 0:
                # Try to get from general data
                general_month = await db["kpi_state_general"].find_one({"Month": current_month, "State": state}, sort=[("Month", -1)])
                total_trans = _get_field_value(general_month, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0 if general_month else 0
            citizen_sla = _get_field_value(current, "Citizen Service SLA % (within SLA)", "Citizen Service SLA %", "CitizenServiceSLA", "Citizen_Service_SLA") or 0
            grievance_sla = _get_field_value(current, "Grievance SLA % (within SLA)", "Grievance SLA %", "GrievanceSLA", "Grievance_SLA") or 0
        
        # Calculate KPIs
        digital_service_penetration = round(((online + faceless) / total_trans * 100) if total_trans > 0 else 0, 2)
        faceless_transformation_index = round((faceless / online * 100) if online > 0 else 0, 2)
        service_reliability_index = round((citizen_sla + grievance_sla) / 2, 2)
        
        # Calculate SLA Volatility (std dev of SLA %)
        sla_values = [_get_field_value(r, "Citizen Service SLA % (within SLA)", "Citizen Service SLA %", "CitizenServiceSLA", "Citizen_Service_SLA") or 0 for r in records]
        sla_volatility = round(_stddev_pop(sla_values), 2)
        
        return {
            "month": current.get("Month"),
            "state": state,
            "kpis": {
                "digital_service_penetration_score": digital_service_penetration,
                "faceless_transformation_index": faceless_transformation_index,
                "service_reliability_index": service_reliability_index,
                "sla_volatility_score": sla_volatility
            },
            "supporting_metrics": {
                "online_services": online,
                "faceless_services": faceless,
                "total_transactions": total_trans,
                "citizen_sla": citizen_sla,
                "grievance_sla": grievance_sla
            }
        }
    except Exception as e:
        logger.error(f"Error fetching digital governance KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/revenue-intelligence")
async def get_revenue_intelligence_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Revenue Intelligence & Leak Detection KPIs"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        
        # Get latest month first
        latest_month_doc = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        if not month:
            query["Month"] = current_month
        
        current = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
        if not current:
            query["Month"] = current_month
            current = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
            if not current:
                return {"error": "No data found"}
        
        current_month = current.get("Month")
        
        # If no state filter, aggregate across all states for current month
        if not state:
            current_agg_cursor = db["kpi_state_general"].find({"Month": current_month})
            current_agg_data = await current_agg_cursor.to_list(100)
            revenue_total = sum(_get_field_value(r, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total") or 0 for r in current_agg_data)
            revenue_tax = sum(_get_field_value(r, "Revenue - Taxes", "Revenue - Taxes ", "RevenueTaxes", "Revenue_Taxes") or 0 for r in current_agg_data)
            revenue_fees = sum(_get_field_value(r, "Revenue - Fees", "Revenue - Fees ", "RevenueFees", "Revenue_Fees") or 0 for r in current_agg_data)
            revenue_penalty = sum(_get_field_value(r, "Revenue - Penalties", "Revenue - Penalties ", "RevenuePenalties", "Revenue_Penalties") or 0 for r in current_agg_data)
            revenue_target = sum(_get_field_value(r, "Revenue Target", "Revenue Target ", "RevenueTarget", "Revenue_Target") or 0 for r in current_agg_data)
            if revenue_target == 0:
                revenue_target = revenue_total * 1.1 if revenue_total > 0 else 0
        else:
            revenue_total = _get_field_value(current, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total") or 0
            revenue_tax = _get_field_value(current, "Revenue - Taxes", "Revenue - Taxes ", "RevenueTaxes", "Revenue_Taxes") or 0
            revenue_fees = _get_field_value(current, "Revenue - Fees", "Revenue - Fees ", "RevenueFees", "Revenue_Fees") or 0
            revenue_penalty = _get_field_value(current, "Revenue - Penalties", "Revenue - Penalties ", "RevenuePenalties", "Revenue_Penalties") or 0
            revenue_target = _get_field_value(current, "Revenue Target", "Revenue Target ", "RevenueTarget", "Revenue_Target") or (revenue_total * 1.1 if revenue_total > 0 else 0)
        
        # Get defaulter data (if available in RTO or state data)
        defaulter_query = query.copy()
        defaulter_cursor = db["kpi_rto_general"].find(defaulter_query)
        defaulter_records = await defaulter_cursor.to_list(100)
        defaulter_count = sum(_as_float(r.get("Tax Defaulter Count")) or 0 for r in defaulter_records)
        defaulter_amount = sum(_as_float(r.get("Tax Defaulter Amount")) or 0 for r in defaulter_records)
        
        # Calculate KPIs
        revenue_quality_index = round(((revenue_tax + revenue_fees) / revenue_total * 100) if revenue_total > 0 else 0, 2)
        penalty_dependency_ratio = round((revenue_penalty / revenue_total * 100) if revenue_total > 0 else 0, 2)
        revenue_achievement_score = round((revenue_total / revenue_target * 100) if revenue_target > 0 else 0, 2)
        revenue_leakage_risk_score = round((defaulter_amount / revenue_total * 100) if revenue_total > 0 else 0, 2)
        
        return {
            "month": current.get("Month"),
            "state": state,
            "kpis": {
                "revenue_quality_index": revenue_quality_index,
                "penalty_dependency_ratio": penalty_dependency_ratio,
                "revenue_achievement_score": revenue_achievement_score,
                "revenue_leakage_risk_score": revenue_leakage_risk_score
            },
            "supporting_metrics": {
                "revenue_total": revenue_total,
                "revenue_taxes": revenue_tax,
                "revenue_fees": revenue_fees,
                "revenue_penalties": revenue_penalty,
                "revenue_target": revenue_target,
                "defaulter_count": defaulter_count,
                "defaulter_amount": defaulter_amount
            }
        }
    except Exception as e:
        logger.error(f"Error fetching revenue intelligence KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/enforcement-safety")
async def get_enforcement_safety_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Enforcement & Road Safety KPIs"""
    try:
        # Get latest month first
        latest_month_doc = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        else:
            query["Month"] = current_month
        
        current = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
        if not current:
            # Try with latest month
            query["Month"] = current_month
            current = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
            if not current:
                return {"error": "No data found"}
        
        current_month = current.get("Month")
        
        # If no state filter, aggregate across all states for current month
        if not state:
            current_agg_cursor = db["kpi_state_general"].find({"Month": current_month})
            current_agg_data = await current_agg_cursor.to_list(100)
            e_challan = sum(_get_field_value(r, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued") or 0 for r in current_agg_data)
            accidents = sum(_get_field_value(r, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents") or 0 for r in current_agg_data)
            fatalities = sum(_get_field_value(r, "Road Fatalities", "Road Fatalities ", "RoadFatalities", "Road_Fatalities") or 0 for r in current_agg_data)
            registrations = sum(_get_field_value(r, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0 for r in current_agg_data)
        else:
            e_challan = _get_field_value(current, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued") or 0
            accidents = _get_field_value(current, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents") or 0
            fatalities = _get_field_value(current, "Road Fatalities", "Road Fatalities ", "RoadFatalities", "Road_Fatalities") or 0
            registrations = _get_field_value(current, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0
        
        # Calculate KPIs
        # Enforcement Effectiveness: Lower accidents per challan is better, so we calculate as (1 - accidents/e_challan) * 100
        # But if e_challan is 0, we can't calculate. Let's use a different metric: e_challan per accident
        # Actually, let's keep it as accidents/e_challan but multiply by 1000 to get a more readable number
        enforcement_effectiveness_ratio = round((accidents / e_challan * 1000) if e_challan > 0 else 0, 2)
        fatality_severity_index = round((fatalities / accidents * 100) if accidents > 0 else 0, 2)
        preventive_enforcement_score = round((e_challan / registrations * 100) if registrations > 0 else 0, 2)
        
        # Get trend for Z-score calculation
        trend_query = {}
        if state:
            trend_query["State"] = state
        trend_cursor = db["kpi_state_general"].find(trend_query).sort("Month", -1).limit(12)
        trend_records = await trend_cursor.to_list(12)
        monthly_accidents = [_get_field_value(r, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents") or 0 for r in trend_records]
        
        if len(monthly_accidents) > 1:
            mean_acc = sum(monthly_accidents) / len(monthly_accidents)
            std_acc = _stddev_pop(monthly_accidents) if len(monthly_accidents) > 1 else 1
            z_score = round((accidents - mean_acc) / std_acc if std_acc > 0 else 0, 2)
        else:
            z_score = 0
        
        high_risk_period_indicator = "High Risk" if abs(z_score) > 2 else "Normal" if abs(z_score) > 1 else "Low Risk"
        
        return {
            "month": current.get("Month"),
            "state": state,
            "kpis": {
                "enforcement_effectiveness_ratio": enforcement_effectiveness_ratio,
                "fatality_severity_index": fatality_severity_index,
                "preventive_enforcement_score": preventive_enforcement_score,
                "high_risk_period_indicator": high_risk_period_indicator,
                "accident_z_score": z_score
            },
            "supporting_metrics": {
                "e_challan_issued": e_challan,
                "road_accidents": accidents,
                "road_fatalities": fatalities,
                "vehicle_registrations": registrations
            }
        }
    except Exception as e:
        logger.error(f"Error fetching enforcement safety KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/policy-effectiveness")
async def get_policy_effectiveness_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Policy Implementation Effectiveness KPIs"""
    try:
        # Get latest month first
        latest_month_doc = await db["kpi_state_policy"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        else:
            query["Month"] = current_month
        
        policy_data = await db["kpi_state_policy"].find_one(query, sort=[("Month", -1)])
        general_data = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
        
        if not policy_data:
            query["Month"] = current_month
            policy_data = await db["kpi_state_policy"].find_one(query, sort=[("Month", -1)])
        if not general_data:
            query["Month"] = current_month
            general_data = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
        
        if not policy_data or not general_data:
            return {"error": "No data found"}
        
        current_month = policy_data.get("Month")
        
        # If no state filter, aggregate across all states
        if not state:
            policy_agg_cursor = db["kpi_state_policy"].find({"Month": current_month})
            policy_agg_data = await policy_agg_cursor.to_list(100)
            general_agg_cursor = db["kpi_state_general"].find({"Month": current_month})
            general_agg_data = await general_agg_cursor.to_list(100)
            
            ats = sum(_get_field_value(r, "No. of ATS", "No. of ATS ", "NoOfATS", "No_of_ATS") or 0 for r in policy_agg_data)
            adtt = sum(_get_field_value(r, "No. of ADTT", "No. of ADTT ", "NoOfADTT", "No_of_ADTT") or 0 for r in policy_agg_data)
            rvsf = sum(_get_field_value(r, "No. of RVSF", "No. of RVSF ", "NoOfRVSF", "No_of_RVSF") or 0 for r in policy_agg_data)
            vltd = sum(_get_field_value(r, "Count of Vehicles fitted with VLTD", "Count of Vehicles fitted with VLTD ", "CountOfVehiclesFittedWithVLTD", "Count_of_Vehicles_fitted_with_VLTD") or 0 for r in policy_agg_data)
            hsrp = sum(_get_field_value(r, "Count of Vehicles fitted with HSRP", "Count of Vehicles fitted with HSRP ", "CountOfVehiclesFittedWithHSRP", "Count_of_Vehicles_fitted_with_HSRP") or 0 for r in policy_agg_data)
            ev_reg = sum(_get_field_value(r, "EV Registrations", "EV Registrations ", "EVRegistrations", "EV_Registrations") or 0 for r in policy_agg_data)
            ev_incentive = sum(_get_field_value(r, "EV Incentives Disbursed", "EV Incentives Disbursed ", "EVIncentivesDisbursed", "EV_Incentives_Disbursed") or 0 for r in policy_agg_data)
            total_vehicles = sum(_get_field_value(r, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0 for r in general_agg_data)
        else:
            ats = _get_field_value(policy_data, "No. of ATS", "No. of ATS ", "NoOfATS", "No_of_ATS") or 0
            adtt = _get_field_value(policy_data, "No. of ADTT", "No. of ADTT ", "NoOfADTT", "No_of_ADTT") or 0
            rvsf = _get_field_value(policy_data, "No. of RVSF", "No. of RVSF ", "NoOfRVSF", "No_of_RVSF") or 0
            vltd = _get_field_value(policy_data, "Count of Vehicles fitted with VLTD", "Count of Vehicles fitted with VLTD ", "CountOfVehiclesFittedWithVLTD", "Count_of_Vehicles_fitted_with_VLTD") or 0
            hsrp = _get_field_value(policy_data, "Count of Vehicles fitted with HSRP", "Count of Vehicles fitted with HSRP ", "CountOfVehiclesFittedWithHSRP", "Count_of_Vehicles_fitted_with_HSRP") or 0
            ev_reg = _get_field_value(policy_data, "EV Registrations", "EV Registrations ", "EVRegistrations", "EV_Registrations") or 0
            ev_incentive = _get_field_value(policy_data, "EV Incentives Disbursed", "EV Incentives Disbursed ", "EVIncentivesDisbursed", "EV_Incentives_Disbursed") or 0
            total_vehicles = _get_field_value(general_data, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0
        
        # Calculate KPIs
        enforcement_infra_density = round(((ats + adtt + rvsf) / total_vehicles * 1000) if total_vehicles > 0 else 0, 2)
        vehicle_traceability_index = round(((vltd + hsrp) / total_vehicles * 100) if total_vehicles > 0 else 0, 2)
        ev_adoption_efficiency = round((ev_reg / ev_incentive) if ev_incentive > 0 else 0, 2)
        green_mobility_transition_score = round((ev_reg / total_vehicles * 100) if total_vehicles > 0 else 0, 2)
        
        return {
            "month": policy_data.get("Month"),
            "state": state,
            "kpis": {
                "enforcement_infrastructure_density": enforcement_infra_density,
                "vehicle_traceability_index": vehicle_traceability_index,
                "ev_adoption_efficiency": ev_adoption_efficiency,
                "green_mobility_transition_score": green_mobility_transition_score
            },
            "supporting_metrics": {
                "ats_count": ats,
                "adtt_count": adtt,
                "rvsf_count": rvsf,
                "vltd_count": vltd,
                "hsrp_count": hsrp,
                "ev_registrations": ev_reg,
                "ev_incentives": ev_incentive,
                "total_vehicles": total_vehicles
            }
        }
    except Exception as e:
        logger.error(f"Error fetching policy effectiveness KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/rto-performance")
async def get_rto_performance_kpis(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get RTO Performance Intelligence KPIs"""
    try:
        # Get latest month first
        latest_month_doc = await db["kpi_rto_performance"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month
        else:
            query["Month"] = current_month
        
        rto_perf = await db["kpi_rto_performance"].find_one(query, sort=[("Month", -1)])
        rto_gen = await db["kpi_rto_general"].find_one(query, sort=[("Month", -1)])
        
        if not rto_perf:
            query["Month"] = current_month
            rto_perf = await db["kpi_rto_performance"].find_one(query, sort=[("Month", -1)])
        if not rto_gen:
            query["Month"] = current_month
            rto_gen = await db["kpi_rto_general"].find_one(query, sort=[("Month", -1)])
        
        if not rto_perf or not rto_gen:
            return {"error": "No data found"}
        
        current_month = rto_perf.get("Month")
        
        # If no state/rto filter, aggregate across all RTOs
        if not state and not rto:
            perf_agg_cursor = db["kpi_rto_performance"].find({"Month": current_month})
            perf_agg_data = await perf_agg_cursor.to_list(100)
            gen_agg_cursor = db["kpi_rto_general"].find({"Month": current_month})
            gen_agg_data = await gen_agg_cursor.to_list(100)
            
            # Average faceless and SLA percentages
            faceless_values = [_as_float(r.get("Faceless %")) or 0 for r in perf_agg_data if r.get("Faceless %")]
            sla_values = [_as_float(r.get("Citizen Service SLA % (within SLA)")) or 0 for r in perf_agg_data if r.get("Citizen Service SLA % (within SLA)")]
            faceless_pct = sum(faceless_values) / len(faceless_values) if faceless_values else 0
            sla_pct = sum(sla_values) / len(sla_values) if sla_values else 0
            
            revenue = sum(_as_float(r.get("Revenue - Total")) or 0 for r in gen_agg_data)
            defaulter_amount = sum(_as_float(r.get("Tax Defaulter Amount")) or 0 for r in gen_agg_data)
            defaulter_count = sum(_as_float(r.get("Tax Defaulter Count")) or 0 for r in gen_agg_data)
        else:
            faceless_pct = _as_float(rto_perf.get("Faceless %")) or 0
            sla_pct = _as_float(rto_perf.get("Citizen Service SLA % (within SLA)")) or 0
            revenue = _as_float(rto_gen.get("Revenue - Total")) or 0
            defaulter_amount = _as_float(rto_gen.get("Tax Defaulter Amount")) or 0
            defaulter_count = _as_float(rto_gen.get("Tax Defaulter Count")) or 0
        
        # Calculate KPIs
        rto_digital_maturity_score = round((faceless_pct * sla_pct) / 100, 2)
        rto_financial_health_index = round(revenue - defaulter_amount, 2)
        rto_risk_flag = "High Risk" if defaulter_count > 100 or defaulter_amount > revenue * 0.1 else "Normal"
        
        # Composite RTO Performance Index (weighted)
        composite_score = round(
            (rto_digital_maturity_score * 0.4) +
            (min(rto_financial_health_index / 1000000, 100) * 0.4) +
            (sla_pct * 0.2), 2
        )
        
        return {
            "month": rto_perf.get("Month"),
            "state": state,
            "rto": rto,
            "kpis": {
                "rto_digital_maturity_score": rto_digital_maturity_score,
                "rto_financial_health_index": rto_financial_health_index,
                "rto_risk_flag": rto_risk_flag,
                "composite_rto_performance_index": composite_score
            },
            "supporting_metrics": {
                "faceless_percentage": faceless_pct,
                "sla_percentage": sla_pct,
                "revenue": revenue,
                "defaulter_count": defaulter_count,
                "defaulter_amount": defaulter_amount
            }
        }
    except Exception as e:
        logger.error(f"Error fetching RTO performance KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/internal-efficiency")
async def get_internal_efficiency_kpis(
    state: Optional[str] = None,
    rto: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Internal Efficiency & Fraud Detection KPIs"""
    try:
        # Get latest month first
        latest_month_doc = await db["kpi_rto_internal"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        query = {}
        if state:
            query["State"] = state
        if rto:
            query["RTO"] = rto
        if month:
            query["Month"] = month
        else:
            query["Month"] = current_month
        
        rto_int = await db["kpi_rto_internal"].find_one(query, sort=[("Month", -1)])
        rto_gen = await db["kpi_rto_general"].find_one(query, sort=[("Month", -1)])
        
        if not rto_int:
            query["Month"] = current_month
            rto_int = await db["kpi_rto_internal"].find_one(query, sort=[("Month", -1)])
        if not rto_gen:
            query["Month"] = current_month
            rto_gen = await db["kpi_rto_general"].find_one(query, sort=[("Month", -1)])
        
        if not rto_int or not rto_gen:
            return {"error": "No data found"}
        
        current_month = rto_int.get("Month")
        
        # If no state/rto filter, aggregate across all RTOs
        if not state and not rto:
            int_agg_cursor = db["kpi_rto_internal"].find({"Month": current_month})
            int_agg_data = await int_agg_cursor.to_list(100)
            gen_agg_cursor = db["kpi_rto_general"].find({"Month": current_month})
            gen_agg_data = await gen_agg_cursor.to_list(100)
            
            total_staff = sum(_get_field_value(r, "Staff - Total", "Total Staff", "Total Staff ", "TotalStaff", "Total_Staff") or 0 for r in int_agg_data)
            field_staff = sum(_get_field_value(r, "Staff - Field Enforcement", "Field Staff", "Field Staff ", "FieldStaff", "Field_Staff") or 0 for r in int_agg_data)
            back_office_staff = sum(_get_field_value(r, "Staff - Back Office", "Back Office Staff", "Back Office Staff ", "BackOfficeStaff", "Back_Office_Staff") or 0 for r in int_agg_data)
            total_transactions = sum(_get_field_value(r, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0 for r in gen_agg_data)
            e_challan = sum(_get_field_value(r, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued") or 0 for r in gen_agg_data)
            anomalies = sum(_get_field_value(r, "Anomalies Detected", "Anomalies Detected ", "AnomaliesDetected", "Anomalies_Detected") or 0 for r in int_agg_data)
        else:
            total_staff = _get_field_value(rto_int, "Staff - Total", "Total Staff", "Total Staff ", "TotalStaff", "Total_Staff") or 0
            field_staff = _get_field_value(rto_int, "Staff - Field Enforcement", "Field Staff", "Field Staff ", "FieldStaff", "Field_Staff") or 0
            back_office_staff = _get_field_value(rto_int, "Staff - Back Office", "Back Office Staff", "Back Office Staff ", "BackOfficeStaff", "Back_Office_Staff") or 0
            total_transactions = _get_field_value(rto_gen, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0
            e_challan = _get_field_value(rto_gen, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued") or 0
            anomalies = _get_field_value(rto_int, "Anomalies Detected", "Anomalies Detected ", "AnomaliesDetected", "Anomalies_Detected") or 0
        
        # Calculate KPIs
        staff_utilization_efficiency = round((total_transactions / total_staff) if total_staff > 0 else 0, 2)
        enforcement_load_ratio = round((e_challan / field_staff) if field_staff > 0 else 0, 2)
        anomaly_density_index = round((anomalies / total_transactions * 100) if total_transactions > 0 else 0, 2)
        governance_stress_indicator = round((back_office_staff / total_staff * 100) if total_staff > 0 else 0, 2)
        
        return {
            "month": rto_int.get("Month"),
            "state": state,
            "rto": rto,
            "kpis": {
                "staff_utilization_efficiency": staff_utilization_efficiency,
                "enforcement_load_ratio": enforcement_load_ratio,
                "anomaly_density_index": anomaly_density_index,
                "governance_stress_indicator": governance_stress_indicator
            },
            "supporting_metrics": {
                "total_staff": total_staff,
                "field_staff": field_staff,
                "back_office_staff": back_office_staff,
                "total_transactions": total_transactions,
                "e_challan_issued": e_challan,
                "anomalies_detected": anomalies
            }
        }
    except Exception as e:
        logger.error(f"Error fetching internal efficiency KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/fleet-compliance")
async def get_fleet_compliance_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Fleet Compliance & Risk KPIs"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        
        fleet_data = await db["kpi_fleet_vehicles"].find_one(query, sort=[("Month", -1)])
        if not fleet_data:
            return {"error": "No data found"}
        
        vehicles_owned = _as_float(fleet_data.get("Vehicle Owned")) or 0
        tax_due_count = _as_float(fleet_data.get("Tax Due - Count")) or 0
        tax_due_amount = _as_float(fleet_data.get("Tax Due - Amount")) or 0
        insurance_due_count = _as_float(fleet_data.get("Insurance Due - Count")) or 0
        insurance_due_amount = _as_float(fleet_data.get("Insurance Due - Amount")) or 0
        fitness_due_count = _as_float(fleet_data.get("Fitness Due - Count")) or 0
        fitness_due_amount = _as_float(fleet_data.get("Fitness Due - Amount")) or 0
        pucc_due_count = _as_float(fleet_data.get("PUCC Due - Count")) or 0
        pucc_due_amount = _as_float(fleet_data.get("PUCC Due - Amount")) or 0
        challan_due_count = _as_float(fleet_data.get("e-Challan Due - Count")) or 0
        challan_due_amount = _as_float(fleet_data.get("e-Challan Due - Amount")) or 0
        
        total_dues = tax_due_amount + insurance_due_amount + fitness_due_amount + pucc_due_amount + challan_due_amount
        
        # Calculate KPIs
        fleet_compliance_score = round((1 - (total_dues / (vehicles_owned * 10000))) * 100 if vehicles_owned > 0 else 0, 2)
        revenue_at_risk_fleet = round(tax_due_amount + challan_due_amount, 2)
        fitness_risk_index = round((fitness_due_count / vehicles_owned * 100) if vehicles_owned > 0 else 0, 2)
        insurance_exposure_score = round((insurance_due_count / vehicles_owned * 100) if vehicles_owned > 0 else 0, 2)
        
        return {
            "month": fleet_data.get("Month"),
            "state": state,
            "kpis": {
                "fleet_compliance_score": max(0, fleet_compliance_score),
                "revenue_at_risk_fleet": revenue_at_risk_fleet,
                "fitness_risk_index": fitness_risk_index,
                "insurance_exposure_score": insurance_exposure_score
            },
            "supporting_metrics": {
                "vehicles_owned": vehicles_owned,
                "tax_due_count": tax_due_count,
                "tax_due_amount": tax_due_amount,
                "insurance_due_count": insurance_due_count,
                "insurance_due_amount": insurance_due_amount,
                "fitness_due_count": fitness_due_count,
                "fitness_due_amount": fitness_due_amount,
                "pucc_due_count": pucc_due_count,
                "challan_due_count": challan_due_count,
                "challan_due_amount": challan_due_amount,
                "total_dues": total_dues
            }
        }
    except Exception as e:
        logger.error(f"Error fetching fleet compliance KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/driver-risk")
async def get_driver_risk_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get Driver Risk & Behaviour Analytics KPIs"""
    try:
        # Get latest month first
        latest_month_doc = await db["kpi_fleet_drivers"].find_one({}, sort=[("Month", -1)])
        if not latest_month_doc:
            return {"error": "No data found"}
        
        current_month = latest_month_doc.get("Month")
        
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        else:
            query["Month"] = current_month
        
        driver_data = await db["kpi_fleet_drivers"].find_one(query, sort=[("Month", -1)])
        if not driver_data:
            query["Month"] = current_month
            driver_data = await db["kpi_fleet_drivers"].find_one(query, sort=[("Month", -1)])
            if not driver_data:
                return {"error": "No data found"}
        
        current_month = driver_data.get("Month")
        
        # If no state filter, aggregate across all states
        if not state:
            driver_agg_cursor = db["kpi_fleet_drivers"].find({"Month": current_month})
            driver_agg_data = await driver_agg_cursor.to_list(100)
            
            driver_count = sum(_get_field_value(r, "Driver Count", "Driver Count ", "DriverCount", "Driver_Count") or 0 for r in driver_agg_data)
            dl_renewal_due_count = sum(_get_field_value(r, "DL Due for Renewal - Count", "DL Renewal Due - Count", "DL Renewal Due - Count ", "DLRenewalDueCount", "DL_Renewal_Due_Count") or 0 for r in driver_agg_data)
            dl_renewal_due_amount = sum(_get_field_value(r, "DL Due for Renewal - Amount", "DL Renewal Due - Amount", "DL Renewal Due - Amount ", "DLRenewalDueAmount", "DL_Renewal_Due_Amount") or 0 for r in driver_agg_data)
            challans_on_dl_count = sum(_get_field_value(r, "e-Challan on DL Due - Count", "Challans on DL - Count", "Challans on DL - Count ", "ChallansOnDLCount", "Challans_on_DL_Count") or 0 for r in driver_agg_data)
            challans_on_dl_amount = sum(_get_field_value(r, "e-Challan on DL Due - Amount", "Challans on DL - Amount", "Challans on DL - Amount ", "ChallansOnDLAmount", "Challans_on_DL_Amount") or 0 for r in driver_agg_data)
        else:
            driver_count = _get_field_value(driver_data, "Driver Count", "Driver Count ", "DriverCount", "Driver_Count") or 0
            dl_renewal_due_count = _get_field_value(driver_data, "DL Due for Renewal - Count", "DL Renewal Due - Count", "DL Renewal Due - Count ", "DLRenewalDueCount", "DL_Renewal_Due_Count") or 0
            dl_renewal_due_amount = _get_field_value(driver_data, "DL Due for Renewal - Amount", "DL Renewal Due - Amount", "DL Renewal Due - Amount ", "DLRenewalDueAmount", "DL_Renewal_Due_Amount") or 0
            challans_on_dl_count = _get_field_value(driver_data, "e-Challan on DL Due - Count", "Challans on DL - Count", "Challans on DL - Count ", "ChallansOnDLCount", "Challans_on_DL_Count") or 0
            challans_on_dl_amount = _get_field_value(driver_data, "e-Challan on DL Due - Amount", "Challans on DL - Amount", "Challans on DL - Amount ", "ChallansOnDLAmount", "Challans_on_DL_Amount") or 0
        
        # Calculate KPIs
        driver_compliance_ratio = round((1 - (dl_renewal_due_count / driver_count)) * 100 if driver_count > 0 else 0, 2)
        habitual_driver_risk_score = round((challans_on_dl_count / driver_count * 100) if driver_count > 0 else 0, 2)
        license_enforcement_yield = round((challans_on_dl_amount / dl_renewal_due_amount) if dl_renewal_due_amount > 0 else 0, 2)
        high_risk_driver_flag = "High Risk" if (dl_renewal_due_count > driver_count * 0.1 or challans_on_dl_count > driver_count * 0.2) else "Normal"
        
        return {
            "month": driver_data.get("Month"),
            "state": state,
            "kpis": {
                "driver_compliance_ratio": max(0, driver_compliance_ratio),
                "habitual_driver_risk_score": habitual_driver_risk_score,
                "license_enforcement_yield": license_enforcement_yield,
                "high_risk_driver_identification_flag": high_risk_driver_flag
            },
            "supporting_metrics": {
                "driver_count": driver_count,
                "dl_renewal_due_count": dl_renewal_due_count,
                "dl_renewal_due_amount": dl_renewal_due_amount,
                "challans_on_dl_count": challans_on_dl_count,
                "challans_on_dl_amount": challans_on_dl_amount
            }
        }
    except Exception as e:
        logger.error(f"Error fetching driver risk KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/super-kpis")
async def get_super_kpis(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Get CM/CS/MoRTH One-Slide Super KPIs"""
    try:
        query = {}
        if state:
            query["State"] = state
        if month:
            query["Month"] = month
        
        # Get all necessary data
        general_data = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
        service_data = await db["kpi_state_service"].find_one(query, sort=[("Month", -1)])
        policy_data = await db["kpi_state_policy"].find_one(query, sort=[("Month", -1)])
        
        if not general_data:
            general_data = await db["kpi_state_general"].find_one({}, sort=[("Month", -1)])
        if not service_data:
            service_data = await db["kpi_state_service"].find_one({}, sort=[("Month", -1)])
        if not policy_data:
            policy_data = await db["kpi_state_policy"].find_one({}, sort=[("Month", -1)])
        
        if not general_data:
            return {"error": "No data found"}
        
        # Extract metrics using helper function
        vehicles = _get_field_value(general_data, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or 0
        revenue = _get_field_value(general_data, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total") or 0
        accidents = _get_field_value(general_data, "Road Accidents", "Road Accidents ", "RoadAccidents", "Road_Accidents") or 0
        fatalities = _get_field_value(general_data, "Road Fatalities", "Road Fatalities ", "RoadFatalities", "Road_Fatalities") or 0
        challans = _get_field_value(general_data, "e-Challan Issued", "e-Challan Issued ", "eChallanIssued", "e_Challan_Issued") or 0
        online = _get_field_value(service_data, "Online Service Count", "Online Service Count ", "OnlineServiceCount", "Online_Service_Count") or 0 if service_data else 0
        faceless = _get_field_value(service_data, "Faceless Service Count", "Faceless Service Count ", "FacelessServiceCount", "Faceless_Service_Count") or 0 if service_data else 0
        citizen_sla = _get_field_value(service_data, "Citizen Service SLA % (within SLA)", "Citizen Service SLA %", "CitizenServiceSLA", "Citizen_Service_SLA") or 0 if service_data else 0
        ats = _get_field_value(policy_data, "No. of ATS", "No. of ATS ", "NoOfATS", "No_of_ATS") or 0 if policy_data else 0
        adtt = _get_field_value(policy_data, "No. of ADTT", "No. of ADTT ", "NoOfADTT", "No_of_ADTT") or 0 if policy_data else 0
        rvsf = _get_field_value(policy_data, "No. of RVSF", "No. of RVSF ", "NoOfRVSF", "No_of_RVSF") or 0 if policy_data else 0
        
        # Calculate Super KPIs
        # 1. Mobility Health Index (Growth + Safety + Compliance)
        # Get distinct months to find previous one
        all_months_cursor = db["kpi_state_general"].find(query, {"Month": 1}).sort("Month", -1).limit(12)
        all_months_data = await all_months_cursor.to_list(12)
        distinct_months = sorted(set(r.get("Month") for r in all_months_data if r.get("Month")), reverse=True)
        
        prev_vehicles = vehicles
        current_month = general_data.get("Month")
        if len(distinct_months) > 1 and current_month in distinct_months:
            prev_month_idx = distinct_months.index(current_month) + 1
            if prev_month_idx < len(distinct_months):
                prev_month = distinct_months[prev_month_idx]
                prev_query = query.copy()
                prev_query["Month"] = prev_month
                prev_cursor = db["kpi_state_general"].find(prev_query).sort("Month", -1).limit(1)
                prev_data = await prev_cursor.to_list(1)
                if prev_data:
                    prev_vehicles = _get_field_value(prev_data[0], "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration") or vehicles
        growth_rate = ((vehicles - prev_vehicles) / prev_vehicles * 100) if prev_vehicles > 0 else 0
        safety_score = 100 - min((accidents / vehicles * 1000) if vehicles > 0 else 0, 100)
        compliance_score = min((challans / vehicles * 100) if vehicles > 0 else 0, 100)
        mobility_health_index = round((growth_rate * 0.3 + safety_score * 0.4 + compliance_score * 0.3), 2)
        
        # 2. Digital Transport Governance Index
        total_trans = _get_field_value(general_data, "Total Transactions (All)", "Total Transactions", "TotalTransactions", "Total_Transactions") or 0
        digital_penetration = ((online + faceless) / total_trans * 100) if total_trans > 0 else 0
        digital_governance_index = round((digital_penetration * 0.6 + citizen_sla * 0.4), 2)
        
        # 3. Road Safety Risk Index
        accident_rate = (accidents / vehicles * 1000) if vehicles > 0 else 0
        fatality_rate = (fatalities / accidents * 100) if accidents > 0 else 0
        road_safety_risk_index = round((accident_rate * 0.6 + fatality_rate * 0.4), 2)
        
        # 4. Revenue Sustainability Index
        revenue_tax = _get_field_value(general_data, "Revenue - Taxes", "Revenue - Taxes ", "RevenueTaxes", "Revenue_Taxes") or 0
        revenue_fees = _get_field_value(general_data, "Revenue - Fees", "Revenue - Fees ", "RevenueFees", "Revenue_Fees") or 0
        revenue_penalty = _get_field_value(general_data, "Revenue - Penalties", "Revenue - Penalties ", "RevenuePenalties", "Revenue_Penalties") or 0
        sustainable_revenue = revenue_tax + revenue_fees
        revenue_sustainability_index = round((sustainable_revenue / revenue * 100) if revenue > 0 else 0, 2)
        
        # 5. Enforcement ROI Index
        total_devices = ats + adtt + rvsf
        enforcement_roi_index = round((revenue / total_devices) if total_devices > 0 else 0, 2)
        
        return {
            "month": general_data.get("Month"),
            "state": state,
            "super_kpis": {
                "mobility_health_index": mobility_health_index,
                "digital_transport_governance_index": digital_governance_index,
                "road_safety_risk_index": road_safety_risk_index,
                "revenue_sustainability_index": revenue_sustainability_index,
                "enforcement_roi_index": enforcement_roi_index
            },
            "supporting_metrics": {
                "total_vehicles": vehicles,
                "total_revenue": revenue,
                "total_accidents": accidents,
                "total_fatalities": fatalities,
                "total_challans": challans,
                "digital_services": online + faceless,
                "enforcement_devices": total_devices
            }
        }
    except Exception as e:
        logger.error(f"Error fetching super KPIs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/insights")
async def get_kpi_insights(
    state: Optional[str] = None,
    month: Optional[str] = None
):
    """Generate insights, recommendations, and action items based on KPI data"""
    try:
        # Get latest month if not specified
        if not month:
            latest_state = await db["kpi_state_general"].find_one(
                {"State": state} if state else {},
                sort=[("Month", -1)]
            )
            month = latest_state.get("Month") if latest_state else None

        query = {"Month": month} if month else {}
        if state:
            query["State"] = state

        insights = []
        recommendations = []
        action_items = []

        # Fetch key KPI data - aggregate if no state filter
        if state:
            # Single state - get one record
            state_gen = await db["kpi_state_general"].find_one(query, sort=[("Month", -1)])
            state_svc = await db["kpi_state_service"].find_one(query, sort=[("Month", -1)])
            state_pol = await db["kpi_state_policy"].find_one(query, sort=[("Month", -1)])
            rto_perf = await db["kpi_rto_performance"].find_one(query, sort=[("Month", -1)])
            fleet_veh = await db["kpi_fleet_vehicles"].find_one(query, sort=[("Month", -1)])
        else:
            # National level - aggregate across all states
            state_gen_cursor = db["kpi_state_general"].find(query, sort=[("Month", -1)])
            state_gen_list = await state_gen_cursor.to_list(length=1000)
            state_svc_cursor = db["kpi_state_service"].find(query, sort=[("Month", -1)])
            state_svc_list = await state_svc_cursor.to_list(length=1000)
            state_pol_cursor = db["kpi_state_policy"].find(query, sort=[("Month", -1)])
            state_pol_list = await state_pol_cursor.to_list(length=1000)
            rto_perf_cursor = db["kpi_rto_performance"].find(query, sort=[("Month", -1)])
            rto_perf_list = await rto_perf_cursor.to_list(length=1000)
            fleet_veh_cursor = db["kpi_fleet_vehicles"].find(query, sort=[("Month", -1)])
            fleet_veh_list = await fleet_veh_cursor.to_list(length=1000)
            
            # Aggregate data
            state_gen = {}
            state_svc = {}
            state_pol = {}
            rto_perf = {}
            fleet_veh = {}
            
            for record in state_gen_list:
                for field in ["Vehicle Registration", "Revenue - Total", "e-Challan - Count", "e-Challan - Amount"]:
                    val = _get_field_value(record, field, f"{field} ", field.replace(" - ", ""), field.replace(" ", "_"))
                    if val is not None:
                        state_gen[field] = state_gen.get(field, 0) + val
            
            for record in state_svc_list:
                for field in ["Citizen Service SLA % (within SLA)", "Grievance SLA % (within SLA)"]:
                    val = _get_field_value(record, field, f"{field} ", field.replace(" %", ""))
                    if val is not None:
                        # For percentages, collect values in a list first
                        if field not in state_svc:
                            state_svc[field] = []
                        if isinstance(state_svc[field], list):
                            state_svc[field].append(val)
            # Calculate averages for percentages after collecting all values
            for field in ["Citizen Service SLA % (within SLA)", "Grievance SLA % (within SLA)"]:
                if field in state_svc and isinstance(state_svc[field], list):
                    state_svc[field] = sum(state_svc[field]) / len(state_svc[field]) if state_svc[field] else 0
            
            for record in state_pol_list:
                val = _get_field_value(record, "EV - Count", "EV - Count ", "EV Count", "EVCount")
                if val is not None:
                    state_pol["EV - Count"] = state_pol.get("EV - Count", 0) + val
            
            for record in rto_perf_list:
                val = _get_field_value(record, "Faceless %", "Faceless % ", "Faceless", "FacelessPercent")
                if val is not None:
                    if "Faceless %" not in rto_perf:
                        rto_perf["Faceless %"] = []
                    if isinstance(rto_perf["Faceless %"], list):
                        rto_perf["Faceless %"].append(val)
            # Calculate average for percentage after collecting all values
            if "Faceless %" in rto_perf and isinstance(rto_perf["Faceless %"], list):
                rto_perf["Faceless %"] = sum(rto_perf["Faceless %"]) / len(rto_perf["Faceless %"]) if rto_perf["Faceless %"] else 0
            
            for record in fleet_veh_list:
                for field in ["Tax Due - Count", "Insurance Due - Count", "Total Vehicles"]:
                    val = _get_field_value(record, field, f"{field} ", field.replace(" - ", ""), field.replace(" ", "_"))
                    if val is not None:
                        fleet_veh[field] = fleet_veh.get(field, 0) + val

        # Get previous month for comparison
        if month:
            prev_month = month[:-2] + str(int(month[-2:]) - 1).zfill(2) if int(month[-2:]) > 1 else str(int(month[:4]) - 1) + "-12"
            prev_query = {"Month": prev_month}
            if state:
                prev_query["State"] = state
            
            if state:
                prev_state_gen = await db["kpi_state_general"].find_one(prev_query)
            else:
                # Aggregate previous month data
                prev_state_gen_cursor = db["kpi_state_general"].find(prev_query)
                prev_state_gen_list = await prev_state_gen_cursor.to_list(length=1000)
                prev_state_gen = {}
                for record in prev_state_gen_list:
                    val = _get_field_value(record, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration")
                    if val is not None:
                        prev_state_gen["Vehicle Registration"] = prev_state_gen.get("Vehicle Registration", 0) + val
                    val = _get_field_value(record, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total")
                    if val is not None:
                        prev_state_gen["Revenue - Total"] = prev_state_gen.get("Revenue - Total", 0) + val
        else:
            prev_state_gen = None

        # Helper function to get field value using _get_field_value for consistency
        def get_val(record, *field_names, default=0):
            if not record:
                return default
            result = _get_field_value(record, *field_names)
            return result if result is not None else default

        # 1. Vehicle Registration Insights
        curr_reg = get_val(state_gen, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration", default=0)
        prev_reg = get_val(prev_state_gen, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration", default=0)
        reg_change = ((curr_reg - prev_reg) / prev_reg * 100) if prev_reg > 0 else 0

        if reg_change > 5:
            insights.append({
                "type": "positive",
                "category": "Mobility & Growth",
                "title": "Strong Vehicle Registration Growth",
                "description": f"Vehicle registrations increased by {reg_change:.1f}% compared to previous month, indicating healthy mobility demand.",
                "metric": f"{curr_reg:,.0f} registrations",
                "trend": "up"
            })
        elif reg_change < -5:
            insights.append({
                "type": "warning",
                "category": "Mobility & Growth",
                "title": "Declining Vehicle Registrations",
                "description": f"Vehicle registrations decreased by {abs(reg_change):.1f}% compared to previous month. This may indicate economic slowdown or policy impact.",
                "metric": f"{curr_reg:,.0f} registrations",
                "trend": "down"
            })

        # 2. Revenue Insights
        curr_rev = get_val(state_gen, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total", default=0)
        prev_rev = get_val(prev_state_gen, "Revenue - Total", "Revenue - Total ", "RevenueTotal", "Revenue_Total", default=0)
        rev_change = ((curr_rev - prev_rev) / prev_rev * 100) if prev_rev > 0 else 0

        if curr_rev > 0:
            insights.append({
                "type": "info",
                "category": "Revenue Intelligence",
                "title": f"Total Revenue: ₹{curr_rev:,.0f}",
                "description": f"Revenue {'increased' if rev_change > 0 else 'decreased'} by {abs(rev_change):.1f}% compared to previous month." if prev_rev > 0 else "Revenue data available for current period.",
                "metric": f"₹{curr_rev:,.0f}",
                "trend": "up" if rev_change > 0 else "down"
            })

        # 3. Service Delivery Insights
        curr_sla = get_val(state_svc, "Citizen Service SLA % (within SLA)", "Citizen Service SLA % (within SLA) ", "CitizenServiceSLA", default=0)
        curr_grievance = get_val(state_svc, "Grievance SLA % (within SLA)", "Grievance SLA % (within SLA) ", "GrievanceSLA", default=0)

        if curr_sla < 80:
            insights.append({
                "type": "critical",
                "category": "Digital Governance",
                "title": "Low Service SLA Performance",
                "description": f"Citizen Service SLA is at {curr_sla:.1f}%, below the 80% target threshold. This impacts citizen satisfaction.",
                "metric": f"{curr_sla:.1f}%",
                "trend": "down"
            })
            recommendations.append({
                "priority": "high",
                "category": "Digital Governance",
                "title": "Improve Service Delivery Performance",
                "description": "Investigate bottlenecks in service delivery processes and implement automation to reduce processing time.",
                "impact": "High",
                "effort": "Medium"
            })
            action_items.append({
                "priority": "high",
                "category": "Digital Governance",
                "title": "Conduct Service Delivery Audit",
                "description": "Review all service delivery workflows and identify delays. Target: Achieve 85% SLA within 2 months.",
                "due_date": "30 days",
                "owner": "Service Delivery Team"
            })

        if curr_grievance < 70:
            insights.append({
                "type": "warning",
                "category": "Digital Governance",
                "title": "Grievance Resolution Below Target",
                "description": f"Grievance SLA is at {curr_grievance:.1f}%, indicating need for improved grievance handling processes.",
                "metric": f"{curr_grievance:.1f}%",
                "trend": "down"
            })

        # 4. Faceless Services Insights
        curr_faceless = get_val(rto_perf, "Faceless %", "Faceless % ", "Faceless", "FacelessPercent", default=0)
        if curr_faceless < 50:
            insights.append({
                "type": "info",
                "category": "Digital Governance",
                "title": "Digital Transformation Opportunity",
                "description": f"Faceless service adoption is at {curr_faceless:.1f}%. There's significant opportunity to increase digital service penetration.",
                "metric": f"{curr_faceless:.1f}%",
                "trend": "neutral"
            })
            recommendations.append({
                "priority": "medium",
                "category": "Digital Governance",
                "title": "Accelerate Faceless Service Adoption",
                "description": "Promote faceless services through awareness campaigns and streamline the digital service experience.",
                "impact": "High",
                "effort": "Low"
            })
            action_items.append({
                "priority": "medium",
                "category": "Digital Governance",
                "title": "Launch Faceless Service Campaign",
                "description": "Create awareness materials and training programs to increase faceless service adoption by 20% in next quarter.",
                "due_date": "45 days",
                "owner": "Digital Transformation Team"
            })

        # 5. Enforcement Insights
        curr_challan = get_val(state_gen, "e-Challan - Count", "e-Challan - Count ", "e-Challan Count", "eChallanCount", default=0)
        curr_challan_amt = get_val(state_gen, "e-Challan - Amount", "e-Challan - Amount ", "e-Challan Amount", "eChallanAmount", default=0)
        
        if curr_challan > 0:
            avg_challan = curr_challan_amt / curr_challan if curr_challan > 0 else 0
            insights.append({
                "type": "info",
                "category": "Enforcement & Road Safety",
                "title": f"Enforcement Activity: {curr_challan:,.0f} Challans",
                "description": f"Average challan amount: ₹{avg_challan:,.0f}. Total enforcement revenue: ₹{curr_challan_amt:,.0f}.",
                "metric": f"{curr_challan:,.0f} challans",
                "trend": "neutral"
            })

        # 6. Fleet Compliance Insights
        tax_due = get_val(fleet_veh, "Tax Due - Count", "Tax Due - Count ", "Tax Due Count", "TaxDueCount", default=0)
        insurance_due = get_val(fleet_veh, "Insurance Due - Count", "Insurance Due - Count ", "Insurance Due Count", "InsuranceDueCount", default=0)
        total_vehicles = get_val(fleet_veh, "Total Vehicles", "Total Vehicles ", "TotalVehicles", "Total_Vehicles", default=0)

        if total_vehicles > 0:
            compliance_rate = ((total_vehicles - tax_due - insurance_due) / total_vehicles * 100) if total_vehicles > 0 else 0
            if compliance_rate < 85:
                insights.append({
                    "type": "warning",
                    "category": "Fleet Compliance",
                    "title": "Fleet Compliance Below Target",
                    "description": f"Fleet compliance rate is {compliance_rate:.1f}%. {tax_due:,.0f} vehicles have tax due, {insurance_due:,.0f} have insurance due.",
                    "metric": f"{compliance_rate:.1f}%",
                    "trend": "down"
                })
                recommendations.append({
                    "priority": "high",
                    "category": "Fleet Compliance",
                    "title": "Improve Fleet Compliance",
                    "description": "Implement automated reminders and streamline payment processes to improve tax and insurance compliance rates.",
                    "impact": "High",
                    "effort": "Medium"
                })
                action_items.append({
                    "priority": "high",
                    "category": "Fleet Compliance",
                    "title": "Deploy Compliance Reminder System",
                    "description": "Set up automated SMS/email reminders for tax and insurance renewals. Target: Achieve 90% compliance within 3 months.",
                    "due_date": "60 days",
                    "owner": "Fleet Management Team"
                })

        # 7. Policy Implementation Insights
        ev_count = get_val(state_pol, "EV - Count", "EV - Count ", "EV Count", "EVCount", default=0)
        total_reg = get_val(state_gen, "Vehicle Registration", "Vehicle Registration ", "VehicleRegistration", "Vehicle_Registration", default=0)
        ev_percentage = (ev_count / total_reg * 100) if total_reg > 0 else 0

        if ev_percentage < 5 and total_reg > 0:
            insights.append({
                "type": "info",
                "category": "Policy Implementation",
                "title": "EV Adoption Below Target",
                "description": f"EV adoption is at {ev_percentage:.2f}% ({ev_count:,.0f} EVs). Government target is typically 5-10%.",
                "metric": f"{ev_percentage:.2f}%",
                "trend": "neutral"
            })
            recommendations.append({
                "priority": "medium",
                "category": "Policy Implementation",
                "title": "Promote EV Adoption",
                "description": "Enhance incentives and infrastructure for EV adoption to meet policy targets.",
                "impact": "Medium",
                "effort": "High"
            })

        # 8. Revenue Optimization Recommendations
        if curr_rev > 0 and prev_rev > 0:
            if rev_change < 0:
                recommendations.append({
                    "priority": "medium",
                    "category": "Revenue Intelligence",
                    "title": "Revenue Decline - Investigate Causes",
                    "description": "Revenue decreased compared to previous month. Review fee structures, enforcement activities, and collection efficiency.",
                    "impact": "High",
                    "effort": "Low"
                })
                action_items.append({
                    "priority": "medium",
                    "category": "Revenue Intelligence",
                    "title": "Revenue Analysis Review",
                    "description": "Conduct detailed analysis of revenue streams to identify decline causes and implement corrective measures.",
                    "due_date": "21 days",
                    "owner": "Finance Team"
                })

        # 9. General Recommendations
        if not state:
            recommendations.append({
                "priority": "low",
                "category": "General",
                "title": "State-Level Analysis Recommended",
                "description": "For more targeted insights and recommendations, filter by specific state to get state-level analysis.",
                "impact": "Medium",
                "effort": "Low"
            })

        # 10. Data Quality Action Items
        if not month:
            action_items.append({
                "priority": "low",
                "category": "Data Quality",
                "title": "Ensure Monthly Data Updates",
                "description": "Verify that all KPI data is updated monthly for accurate insights and trend analysis.",
                "due_date": "Ongoing",
                "owner": "Data Management Team"
            })

        return {
            "insights": insights,
            "recommendations": recommendations,
            "action_items": action_items,
            "summary": {
                "total_insights": len(insights),
                "total_recommendations": len(recommendations),
                "total_action_items": len(action_items),
                "month": month,
                "state": state
            }
        }
    except Exception as e:
        logger.error(f"Error generating insights: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@kpi_router.get("/advanced/insights")
async def get_advanced_kpi_insights(
    state: Optional[str] = None,
    month: Optional[str] = None,
    section: Optional[str] = None
):
    """Generate insights, recommendations, and action items based on Advanced KPI data for a specific section"""
    try:
        insights = []
        recommendations = []
        action_items = []
        
        # Fetch only the relevant KPI data based on section
        if section == "mobility_growth" or section is None:
            mobility_res = await get_mobility_growth_kpis(state, month)
        else:
            mobility_res = {}
            
        if section == "digital_governance" or section is None:
            digital_res = await get_digital_governance_kpis(state, month)
        else:
            digital_res = {}
            
        if section == "revenue_intelligence" or section is None:
            revenue_res = await get_revenue_intelligence_kpis(state, month)
        else:
            revenue_res = {}
            
        if section == "enforcement_safety" or section is None:
            enforcement_res = await get_enforcement_safety_kpis(state, month)
        else:
            enforcement_res = {}
            
        if section == "policy_effectiveness" or section is None:
            policy_res = await get_policy_effectiveness_kpis(state, month)
        else:
            policy_res = {}
            
        if section == "rto_performance" or section is None:
            rto_res = await get_rto_performance_kpis(state, None, month)
        else:
            rto_res = {}
            
        if section == "internal" or section is None:
            internal_res = await get_internal_efficiency_kpis(state, None, month)
        else:
            internal_res = {}
            
        if section == "fleet_compliance" or section is None:
            fleet_res = await get_fleet_compliance_kpis(state, month)
        else:
            fleet_res = {}
            
        if section == "driver" or section is None:
            driver_res = await get_driver_risk_kpis(state, month)
        else:
            driver_res = {}

        insights = []
        recommendations = []
        action_items = []

        # Helper to safely get KPI value from response structure
        def get_kpi_val(data, key, default=0):
            if not data or not isinstance(data, dict):
                return default
            # Check in kpis dictionary first
            kpis = data.get("kpis", {})
            if isinstance(kpis, dict):
                val = kpis.get(key)
                if val is not None:
                    # Handle string values (remove %, commas, etc.)
                    if isinstance(val, str):
                        val = val.replace('%', '').replace(',', '').strip()
                    try:
                        result = float(val) if val else default
                        # Return None if result is NaN or invalid, otherwise return the value
                        if math.isnan(result) or math.isinf(result):
                            return default
                        return result
                    except (ValueError, TypeError):
                        return default
            # If not found in kpis, return default
            return default

        # Generate section-specific insights
        # 1. Mobility Growth Insights (for mobility_growth section)
        if section is None or section == "mobility_growth":
            momentum = get_kpi_val(mobility_res, "vehicle_demand_momentum_index", 0)
            if momentum < -5:
                insights.append({
                    "type": "warning",
                    "category": "Mobility & Growth",
                    "title": "Negative Vehicle Demand Momentum",
                    "description": f"Vehicle Demand Momentum Index is {momentum:.1f}%, indicating declining demand. This may signal economic concerns or policy impact.",
                    "metric": f"{momentum:.1f}%",
                    "trend": "down"
                })
                recommendations.append({
                    "priority": "high",
                    "category": "Mobility & Growth",
                    "title": "Investigate Demand Decline",
                    "description": "Analyze root causes of declining vehicle demand - economic factors, policy changes, or market saturation.",
                    "impact": "High",
                    "effort": "Medium"
                })

        # 2. Digital Governance Insights
        penetration = get_kpi_val(digital_res, "digital_service_penetration_score", 0)
        faceless = get_kpi_val(digital_res, "faceless_transformation_index", 0)
        
        if penetration < 50:
            insights.append({
                "type": "warning",
                "category": "Digital Governance",
                "title": "Low Digital Service Penetration",
                "description": f"Digital Service Penetration Score is {penetration:.1f}%, indicating significant opportunity for digital transformation.",
                "metric": f"{penetration:.1f}%",
                "trend": "neutral"
            })
            recommendations.append({
                "priority": "high",
                "category": "Digital Governance",
                "title": "Accelerate Digital Adoption",
                "description": "Implement user-friendly digital interfaces, reduce friction in online processes, and promote digital services through campaigns.",
                "impact": "High",
                "effort": "Medium"
            })
            action_items.append({
                "priority": "high",
                "category": "Digital Governance",
                "title": "Digital Service Enhancement Program",
                "description": "Redesign digital service workflows to improve user experience. Target: Increase penetration by 15% in 3 months.",
                "due_date": "90 days",
                "owner": "Digital Team"
            })

        if faceless < 40:
            insights.append({
                "type": "info",
                "category": "Digital Governance",
                "title": "Faceless Transformation Opportunity",
                "description": f"Faceless Transformation Index is {faceless:.1f}%. Increasing automation can reduce operational costs and improve efficiency.",
                "metric": f"{faceless:.1f}%",
                "trend": "neutral"
            })

        # 3. Revenue Intelligence Insights
        revenue_efficiency = get_kpi_val(revenue_res, "revenue_collection_efficiency", 0)
        if revenue_efficiency < 80:
            insights.append({
                "type": "warning",
                "category": "Revenue Intelligence",
                "title": "Revenue Collection Efficiency Below Target",
                "description": f"Revenue Collection Efficiency is {revenue_efficiency:.1f}%, indicating potential gaps in collection processes.",
                "metric": f"{revenue_efficiency:.1f}%",
                "trend": "down"
            })
            recommendations.append({
                "priority": "medium",
                "category": "Revenue Intelligence",
                "title": "Improve Collection Processes",
                "description": "Streamline payment processes, implement automated reminders, and enhance collection tracking systems.",
                "impact": "High",
                "effort": "Medium"
            })

        # 4. Enforcement & Safety Insights
        enforcement_ratio = get_kpi_val(enforcement_res, "enforcement_effectiveness_ratio", 0)
        if enforcement_ratio < 0.5:
            insights.append({
                "type": "warning",
                "category": "Enforcement & Road Safety",
                "title": "Low Enforcement Effectiveness",
                "description": f"Enforcement Effectiveness Ratio is {enforcement_ratio:.2f}, suggesting need for improved enforcement strategies.",
                "metric": f"{enforcement_ratio:.2f}",
                "trend": "down"
            })
            recommendations.append({
                "priority": "high",
                "category": "Enforcement & Road Safety",
                "title": "Enhance Enforcement Capabilities",
                "description": "Deploy more enforcement devices, improve training, and implement data-driven enforcement strategies.",
                "impact": "High",
                "effort": "High"
            })
            action_items.append({
                "priority": "high",
                "category": "Enforcement & Road Safety",
                "title": "Enforcement Infrastructure Upgrade",
                "description": "Deploy additional enforcement devices and train staff on effective enforcement techniques. Target: Increase effectiveness ratio by 30% in 6 months.",
                "due_date": "180 days",
                "owner": "Enforcement Team"
            })

        # 5. Policy Effectiveness Insights
        ev_efficiency = get_kpi_val(policy_res, "ev_adoption_efficiency", 0)
        green_score = get_kpi_val(policy_res, "green_mobility_transition_score", 0)
        
        if ev_efficiency < 30:
            insights.append({
                "type": "info",
                "category": "Policy Implementation",
                "title": "EV Adoption Needs Boost",
                "description": f"EV Adoption Efficiency is {ev_efficiency:.1f}%, indicating need for stronger incentives and infrastructure.",
                "metric": f"{ev_efficiency:.1f}%",
                "trend": "neutral"
            })
            recommendations.append({
                "priority": "medium",
                "category": "Policy Implementation",
                "title": "Enhance EV Support Infrastructure",
                "description": "Increase EV charging infrastructure, provide better incentives, and streamline EV registration processes.",
                "impact": "Medium",
                "effort": "High"
            })

        if green_score < 40:
            insights.append({
                "type": "info",
                "category": "Policy Implementation",
                "title": "Green Mobility Transition Progress",
                "description": f"Green Mobility Transition Score is {green_score:.1f}%. Continued focus on sustainable mobility policies is needed.",
                "metric": f"{green_score:.1f}%",
                "trend": "neutral"
            })

        # 6. RTO Performance Insights
        rto_sla = get_kpi_val(rto_res, "citizen_service_sla", 0)
        if rto_sla < 75:
            insights.append({
                "type": "critical",
                "category": "RTO Performance",
                "title": "RTO Service SLA Below Target",
                "description": f"RTO Citizen Service SLA is {rto_sla:.1f}%, significantly below the 80% target. Immediate action required.",
                "metric": f"{rto_sla:.1f}%",
                "trend": "down"
            })
            action_items.append({
                "priority": "critical",
                "category": "RTO Performance",
                "title": "RTO Service Delivery Improvement Plan",
                "description": "Conduct comprehensive review of RTO operations, identify bottlenecks, and implement process improvements. Target: Achieve 85% SLA within 2 months.",
                "due_date": "60 days",
                "owner": "RTO Operations Team"
            })

        # 7. Internal Efficiency Insights
        staff_productivity = get_kpi_val(internal_res, "staff_productivity_index", 0)
        if staff_productivity < 50:
            insights.append({
                "type": "warning",
                "category": "Internal Efficiency",
                "title": "Staff Productivity Below Optimal",
                "description": f"Staff Productivity Index is {staff_productivity:.1f}%, indicating potential for process optimization and training.",
                "metric": f"{staff_productivity:.1f}%",
                "trend": "down"
            })
            recommendations.append({
                "priority": "medium",
                "category": "Internal Efficiency",
                "title": "Optimize Staff Productivity",
                "description": "Implement automation, provide training, and optimize workflows to improve staff productivity.",
                "impact": "Medium",
                "effort": "Medium"
            })

        # 8. Fleet Compliance Insights
        fleet_risk = get_kpi_val(fleet_res, "fleet_compliance_risk_score", 0)
        if fleet_risk > 30:
            insights.append({
                "type": "warning",
                "category": "Fleet Compliance",
                "title": "High Fleet Compliance Risk",
                "description": f"Fleet Compliance Risk Score is {fleet_risk:.1f}%, indicating significant number of non-compliant vehicles.",
                "metric": f"{fleet_risk:.1f}%",
                "trend": "down"
            })
            action_items.append({
                "priority": "high",
                "category": "Fleet Compliance",
                "title": "Fleet Compliance Drive",
                "description": "Launch targeted campaign to improve fleet compliance. Implement automated reminders and streamline renewal processes.",
                "due_date": "45 days",
                "owner": "Fleet Management Team"
            })

        # 9. Driver Risk Insights
        driver_risk = get_kpi_val(driver_res, "habitual_driver_risk_score", 0)
        if driver_risk > 25:
            insights.append({
                "type": "warning",
                "category": "Driver Risk",
                "title": "High Driver Risk Identified",
                "description": f"Habitual Driver Risk Score is {driver_risk:.1f}%, indicating drivers with multiple violations requiring attention.",
                "metric": f"{driver_risk:.1f}%",
                "trend": "down"
            })
            recommendations.append({
                "priority": "high",
                "category": "Driver Risk",
                "title": "Implement Driver Risk Management",
                "description": "Develop targeted interventions for high-risk drivers, including mandatory training and enhanced monitoring.",
                "impact": "High",
                "effort": "Medium"
            })
            action_items.append({
                "priority": "high",
                "category": "Driver Risk",
                "title": "Driver Risk Intervention Program",
                "description": "Identify high-risk drivers and implement mandatory training programs. Target: Reduce risk score by 20% in 4 months.",
                "due_date": "120 days",
                "owner": "Driver Safety Team"
            })

        return {
            "insights": insights,
            "recommendations": recommendations,
            "action_items": action_items,
            "summary": {
                "total_insights": len(insights),
                "total_recommendations": len(recommendations),
                "total_action_items": len(action_items),
                "month": month,
                "state": state,
                "section": section
            }
        }
    except Exception as e:
        logger.error(f"Error generating advanced insights: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint to verify server and database connectivity"""
    try:
        # Check MongoDB connection
        mongo_status = "disconnected"
        mongo_error = None
        collections_status = {}
        
        try:
            await asyncio.wait_for(client.admin.command("ping"), timeout=2.0)
            mongo_status = "connected"
            
            # Check if collections exist and have data
            collections_to_check = [
                "vahan_data", "tickets_data", "kpi_state_general", 
                "kpi_state_service", "kpi_rto_general", "kpi_fleet_vehicles"
            ]
            for coll_name in collections_to_check:
                try:
                    count = await db[coll_name].count_documents({})
                    collections_status[coll_name] = {"exists": True, "count": count}
                except Exception as e:
                    collections_status[coll_name] = {"exists": False, "error": str(e)}
        except Exception as e:
            mongo_error = str(e)
        
        return {
            "status": "ok" if mongo_status == "connected" else "degraded",
            "server": "running",
            "mongodb": {
                "status": mongo_status,
                "url": mongo_url,
                "database": os.environ.get('DB_NAME', 'citizen_assistance'),
                "error": mongo_error
            },
            "collections": collections_status if mongo_status == "connected" else {},
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        logger.error(f"Health check error: {e}")
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

@app.get("/api/health")
async def api_health_check():
    """API health check endpoint"""
    return await health_check()

api_router.include_router(dashboard_router)
api_router.include_router(kpi_router)
api_router.include_router(tickets_router)
api_router.include_router(chatbot_router)
api_router.include_router(stt_router)
api_router.include_router(ocr_router)
api_router.include_router(aadhaar_router)
api_router.include_router(facial_router)
api_router.include_router(vehicle_router)

app.include_router(api_router)

_cors_origins_raw = os.environ.get("CORS_ORIGINS", "").strip()
# Browser-safe defaults:
# - if CORS_ORIGINS not set: allow localhost + common private LAN ranges on port 3000 and allow credentials
# - if CORS_ORIGINS contains "*": allow any origin but DISABLE credentials (required by CORS spec)
_cors_allow_origin_regex = None
if not _cors_origins_raw:
    _cors_allow_origins = ["http://localhost:3000", "http://127.0.0.1:3000"]
    # Allow common private LAN ranges for dev (e.g., http://192.168.1.10:3000)
    _cors_allow_origin_regex = r"^http://((localhost)|(127\.0\.0\.1)|((10|192\.168|172\.(1[6-9]|2\d|3[0-1]))\.\d+\.\d+\.\d+))(:3000)?$"
    _cors_allow_credentials = True
else:
    _cors_allow_origins = [o.strip() for o in _cors_origins_raw.split(",") if o.strip()]
    if "*" in _cors_allow_origins:
        _cors_allow_origins = ["*"]
        _cors_allow_credentials = False
    else:
        _cors_allow_credentials = True

# Add rate limiting middleware (optional, can be disabled via env var)
# Disable rate limiting by default in development, enable in production
# Set RATE_LIMIT_ENABLED=true to enable in development
rate_limit_enabled = os.environ.get("RATE_LIMIT_ENABLED", "false").lower() == "true"
rate_limit_rpm = int(os.environ.get("RATE_LIMIT_RPM", "100"))

if rate_limit_enabled:
    app.add_middleware(RateLimitMiddleware, requests_per_minute=rate_limit_rpm, enabled=True)
    logger.info(f"Rate limiting enabled: {rate_limit_rpm} requests per minute")
else:
    logger.info("Rate limiting disabled (development mode - set RATE_LIMIT_ENABLED=true to enable)")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=_cors_allow_credentials,
    allow_origins=_cors_allow_origins,
    allow_origin_regex=_cors_allow_origin_regex,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """Load data on startup"""
    logger.info("Starting Citizen Assistance Platform...")
    logger.info(f"MongoDB URL: {mongo_url}")
    logger.info(f"Database: {os.environ.get('DB_NAME', 'citizen_assistance')}")
    
    # If MongoDB isn't available, don't block server startup.
    try:
        await asyncio.wait_for(client.admin.command("ping"), timeout=5.0)
        logger.info("MongoDB connection successful")
        
        # Check existing data counts
        vahan_count = await db.vahan_data.count_documents({})
        tickets_count = await db.tickets_data.count_documents({})
        kpi_count = await db["kpi_state_general"].count_documents({})
        
        logger.info(f"Existing data counts - Vahan: {vahan_count}, Tickets: {tickets_count}, KPI: {kpi_count}")
        
        # Only load data if collections are empty
        if vahan_count == 0:
            logger.info("Loading Vahan data...")
            await load_vahan_data()
        else:
            logger.info(f"Skipping Vahan data load - {vahan_count} records already exist")
            
        if tickets_count == 0:
            logger.info("Loading Tickets data...")
            await load_tickets_data()
        else:
            logger.info(f"Skipping Tickets data load - {tickets_count} records already exist")
            
        if kpi_count == 0:
            logger.info("Loading KPI data...")
            await load_kpi_data()
        else:
            logger.info(f"Skipping KPI data load - {kpi_count} records already exist")
        
        logger.info("Data loading complete")
    except asyncio.TimeoutError:
        logger.error("MongoDB connection timeout - server will start but data endpoints may fail")
        logger.error("Please ensure MongoDB is running and accessible at: " + mongo_url)
    except Exception as e:
        logger.error(f"MongoDB connection failed; skipping data bootstrap. Error: {e}")
        logger.error("Server will start but data endpoints may not work. Check MongoDB connection.")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
